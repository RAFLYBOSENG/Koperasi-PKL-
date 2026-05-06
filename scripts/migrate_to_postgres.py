import json
import os
import sys
from datetime import datetime
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    try:
        for raw_line in path.read_text(encoding='utf-8').splitlines():
            line = raw_line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
    except Exception:
        pass


_load_env_file(ROOT_DIR / '.env')

from openpyxl import load_workbook
from sqlalchemy import text

from koperasi_system.db import db_session, init_db_schema, ping_database
from koperasi_system.settings import (
    ANGGOTA_FIELDNAMES,
    BASE_DIR,
    CICILAN_FIELDNAMES,
    DATA_DIR,
    FILE_ANGGOTA,
    FILE_IMPORT_LOG,
    FILE_IURAN_SOSIAL,
    FILE_PENDAFTARAN_ANGGOTA,
    FILE_PINJAMAN,
    FILE_PINJAMAN_CICILAN,
    FILE_SIMPANAN,
    FILE_SIMPANAN_PENGAJUAN,
    FILE_SIMPANAN_TRANSAKSI,
    FILE_USERS,
    IURAN_SOSIAL_FIELDNAMES,
    PENDAFTARAN_FIELDNAMES,
    PINJAMAN_FIELDNAMES,
    SIMPANAN_FIELDNAMES,
    SIMPANAN_PENGAJUAN_FIELDNAMES,
    SIMPANAN_TRANSAKSI_FIELDNAMES,
)

SCHEMA_PATH = Path(BASE_DIR) / "db" / "schema.sql"
BLUEPRINT_USER_ROLES = [
    'super_admin',
    'admin_koperasi',
    'bendahara',
    'ketua_pengurus',
    'anggota',
    'auditor',
    'admin',
    'user',
]


def read_rows(filepath: str):
    path = Path(filepath)
    if not path.exists():
        return []
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        out = []
        for row in rows[1:]:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                item[header] = "" if value is None else str(value).strip()
            if any(item.values()):
                out.append(item)
        return out
    raise ValueError(f"Unsupported file type: {path.suffix}")


def normalize_num(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def normalize_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        return None


def import_table(conn, table_name: str, rows: list, fieldnames: list):
    if not rows:
        return 0
    columns = [c for c in fieldnames if c != "id"]
    placeholders = ", ".join(f":{c}" for c in columns)
    columns_sql = ", ".join(columns)
    stmt = text(f"INSERT INTO {table_name} ({columns_sql}) VALUES ({placeholders}) ON CONFLICT DO NOTHING")
    count = 0
    for row in rows:
        payload = {}
        for k in columns:
            value = row.get(k, None)
            # Empty strings from XLSX should become SQL NULL to avoid DATE/TIMESTAMP cast errors.
            if isinstance(value, str):
                value = value.strip()
            payload[k] = None if value == "" else value
        conn.execute(stmt, payload)
        count += 1
    return count


def ensure_users_role_constraint(conn) -> None:
    role_list_sql = ", ".join(f"'{role}'" for role in BLUEPRINT_USER_ROLES)
    conn.execute(text("ALTER TABLE users DROP CONSTRAINT IF EXISTS users_role_check"))
    conn.execute(text(f"ALTER TABLE users ADD CONSTRAINT users_role_check CHECK (role IN ({role_list_sql}))"))


def main():
    if not ping_database():
        raise RuntimeError("DATABASE_URL belum valid atau database Neon tidak dapat diakses.")

    init_db_schema(str(SCHEMA_PATH))

    with db_session() as conn:
        ensure_users_role_constraint(conn)

        # urutan penting karena foreign key
        users = read_rows(FILE_USERS)
        anggota = read_rows(FILE_ANGGOTA)
        simpanan = read_rows(FILE_SIMPANAN)
        simpanan_transaksi = read_rows(FILE_SIMPANAN_TRANSAKSI)
        simpanan_pengajuan = read_rows(FILE_SIMPANAN_PENGAJUAN)
        iuran_sosial = read_rows(FILE_IURAN_SOSIAL)
        pinjaman = read_rows(FILE_PINJAMAN)
        pinjaman_cicilan = read_rows(FILE_PINJAMAN_CICILAN)
        pendaftaran = read_rows(FILE_PENDAFTARAN_ANGGOTA)
        import_log = read_rows(FILE_IMPORT_LOG)

        # truncate supaya import ulang aman saat migrasi awal
        for tbl in [
            "pinjaman_cicilan",
            "pinjaman",
            "iuran_sosial",
            "simpanan_pengajuan",
            "simpanan_transaksi",
            "simpanan",
            "users",
            "pendaftaran_anggota",
            "anggota",
            "import_log",
        ]:
            conn.execute(text(f"TRUNCATE TABLE {tbl} RESTART IDENTITY CASCADE"))

        # insert anggota dulu karena dipakai FK
        import_table(conn, "anggota", anggota, ANGGOTA_FIELDNAMES)
        import_table(conn, "simpanan", simpanan, SIMPANAN_FIELDNAMES)
        import_table(conn, "simpanan_transaksi", simpanan_transaksi, SIMPANAN_TRANSAKSI_FIELDNAMES)
        import_table(conn, "simpanan_pengajuan", simpanan_pengajuan, SIMPANAN_PENGAJUAN_FIELDNAMES)
        import_table(conn, "iuran_sosial", iuran_sosial, IURAN_SOSIAL_FIELDNAMES)
        import_table(conn, "pinjaman", pinjaman, PINJAMAN_FIELDNAMES)
        import_table(conn, "pinjaman_cicilan", pinjaman_cicilan, CICILAN_FIELDNAMES)
        import_table(conn, "pendaftaran_anggota", pendaftaran, PENDAFTARAN_FIELDNAMES)
        import_table(conn, "users", users, ["id_user", "username", "password_hash", "role", "id_anggota", "created_at"])
        if import_log:
            import_table(conn, "import_log", import_log, ["id", "waktu", "user", "mode", "nama_file", "berhasil", "gagal", "catatan"])

    print("Migrasi awal ke PostgreSQL selesai.")


if __name__ == "__main__":
    main()
