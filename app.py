import os
import csv
import re
import json
import uuid
import secrets
import hmac
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, session, flash, abort
import io
import sys
try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:
    Workbook = None
    load_workbook = None
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)

# ──────────────────────────────────────────────
#  PATH DATABASE CSV
# ──────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
os.makedirs(DATA_DIR, exist_ok=True)

FILE_ANGGOTA = os.path.join(DATA_DIR, 'anggota.csv')
FILE_SIMPANAN = os.path.join(DATA_DIR, 'simpanan.csv')
FILE_PINJAMAN = os.path.join(DATA_DIR, 'pinjaman.csv')
FILE_PINJAMAN_CICILAN = os.path.join(DATA_DIR, 'pinjaman_cicilan.csv')
FILE_USERS = os.path.join(DATA_DIR, 'users.csv')
FILE_PENDAFTARAN_ANGGOTA = os.path.join(DATA_DIR, 'pendaftaran_anggota.csv')
FILE_IMPORT_LOG = os.path.join(DATA_DIR, 'import_log.csv')
DSR_DEFAULT = 0.35
JENIS_SIMPANAN_IMPORT = 'Manasuka'
DEFAULT_TENOR_IMPORT_PINJAMAN = 12
METODE_BAYAR_CHOICES = (
    'Indomaret',
    'Alfamart',
    'Dana',
    'Transfer Bank',
    'Lainnya',
)
CICILAN_FIELDNAMES = [
    'id_cicilan', 'id_pinjaman', 'id_anggota', 'no_anggota',
    'nama_anggota', 'jumlah', 'tanggal_pengajuan', 'status',
    'tanggal_konfirmasi', 'dikonfirmasi_oleh', 'diajukan_oleh',
    'keterangan', 'metode_pembayaran', 'detail_pembayaran',
]
# Skema saldo per anggota (satu baris per id_anggota)
SIMPANAN_FIELDNAMES = ['id_anggota', 'total_simpanan']
# Pinjaman: banyak baris per anggota; tidak menggabungkan jenis berbeda (bunga berbeda)
PINJAMAN_FIELDNAMES = [
    'id_pinjaman', 'id_anggota', 'nama_anggota', 'no_anggota',
    'jenis_pinjaman', 'plafon', 'tenor_bulan', 'bunga_persen',
    'total_bayar', 'cicilan_per_bulan', 'sisa_pinjaman',
    'tanggal_pengajuan', 'status', 'tanggal_lunas',
]
JENIS_IMPORT_CSV = 'Import CSV'
IMPORT_PREVIEW_DIR = os.path.join(DATA_DIR, 'import_preview')
ANGGOTA_FIELDNAMES = ['id_anggota', 'no_anggota', 'nik', 'nama', 'alamat', 'no_telp', 'tgl_bergabung',
                      'penghasilan_bersih', 'cicilan_lain']

PENDAFTARAN_FIELDNAMES = [
    'id_pengajuan', 'nama', 'alamat', 'no_telp', 'penghasilan_bersih', 'cicilan_lain',
    'status', 'tanggal_pengajuan', 'catatan_admin', 'id_anggota_dibuat', 'no_anggota_dibuat',
]


# ──────────────────────────────────────────────
#  HELPER: Baca & Tulis CSV
# ──────────────────────────────────────────────
def baca_csv(filepath):
    """Baca file CSV dan kembalikan list of dict."""
    if not os.path.exists(filepath):
        return []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)


def tulis_csv(filepath, data, fieldnames):
    """Tulis list of dict ke file CSV."""
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)


def init_csv():
    """Inisialisasi file CSV jika belum ada."""
    if not os.path.exists(FILE_ANGGOTA):
        tulis_csv(FILE_ANGGOTA, [], ANGGOTA_FIELDNAMES)
    if not os.path.exists(FILE_SIMPANAN):
        tulis_csv(FILE_SIMPANAN, [], SIMPANAN_FIELDNAMES)
    if not os.path.exists(FILE_PINJAMAN):
        tulis_csv(FILE_PINJAMAN, [], PINJAMAN_FIELDNAMES)
    if not os.path.exists(FILE_PINJAMAN_CICILAN):
        tulis_csv(FILE_PINJAMAN_CICILAN, [], CICILAN_FIELDNAMES)
    if not os.path.exists(FILE_USERS):
        admin_default_password = os.environ.get('DEFAULT_ADMIN_PASSWORD') or (
            f"Adm!{secrets.token_hex(6)}"
        )
        user_default_password = os.environ.get('DEFAULT_USER_PASSWORD') or (
            f"Usr!{secrets.token_hex(6)}"
        )
        default_users = [
            {
                'id_user': str(uuid.uuid4()),
                'username': 'admin',
                'password_hash': generate_password_hash(admin_default_password),
                'role': 'admin',
                'id_anggota': '',
                'created_at': datetime.now().strftime('%Y-%m-%d')
            },
            {
                'id_user': str(uuid.uuid4()),
                'username': 'user',
                'password_hash': generate_password_hash(user_default_password),
                'role': 'user',
                'id_anggota': '',
                'created_at': datetime.now().strftime('%Y-%m-%d')
            }
        ]
        tulis_csv(FILE_USERS, default_users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
        app.logger.warning(
            'users.csv dibuat otomatis. Password default admin/user di-generate acak. '
            'Atur DEFAULT_ADMIN_PASSWORD dan DEFAULT_USER_PASSWORD agar tidak berubah.'
        )
    if not os.path.exists(FILE_PENDAFTARAN_ANGGOTA):
        tulis_csv(FILE_PENDAFTARAN_ANGGOTA, [], PENDAFTARAN_FIELDNAMES)


init_csv()


def migrate_simpanan_ke_saldo():
    """Migrasi skema lama (transaksi) ke saldo per id_anggota."""
    rows = baca_csv(FILE_SIMPANAN)
    if not rows:
        return
    if 'total_simpanan' in rows[0] and 'id_anggota' in rows[0]:
        merged = _dedupe_rows_simpanan(rows)
        tulis_csv(FILE_SIMPANAN, merged, SIMPANAN_FIELDNAMES)
        return
    agg = {}
    for s in rows:
        id_a = (s.get('id_anggota') or '').strip()
        if not id_a:
            continue
        if (s.get('status') or 'Disetujui') != 'Disetujui':
            continue
        try:
            agg[id_a] = agg.get(id_a, 0.0) + float(s.get('jumlah') or 0)
        except (TypeError, ValueError):
            continue
    out = [{'id_anggota': k, 'total_simpanan': str(round(v, 2))} for k, v in sorted(agg.items())]
    tulis_csv(FILE_SIMPANAN, out, SIMPANAN_FIELDNAMES)


def migrate_pinjaman_ke_saldo():
    """Migrasi berbagai skema lama ke format multi-baris + jenis pinjaman."""
    rows = baca_csv(FILE_PINJAMAN)
    if not rows:
        return
    sample = rows[0]

    if 'id_pinjaman' in sample:
        normalized = []
        for r in rows:
            row = {k: (r.get(k) if r.get(k) is not None else '') for k in PINJAMAN_FIELDNAMES}
            if not row.get('id_pinjaman'):
                row['id_pinjaman'] = str(uuid.uuid4())
            normalized.append(row)
        tulis_csv(FILE_PINJAMAN, normalized, PINJAMAN_FIELDNAMES)
        return

    # Skema saldo 3 kolom (id_anggota, total_pinjaman, tenor)
    if 'total_pinjaman' in sample and 'tenor' in sample and 'plafon' not in sample:
        amap = {a.get('id_anggota'): a for a in baca_csv(FILE_ANGGOTA)}
        out = []
        for r in rows:
            id_a = (r.get('id_anggota') or '').strip()
            if not id_a:
                continue
            try:
                plaf = float(r.get('total_pinjaman') or 0)
                ten = int(float(r.get('tenor') or 0))
            except (TypeError, ValueError):
                continue
            ag = amap.get(id_a) or {}
            bunga_p = bunga_dari_tenor(ten) if ten > 0 else 0.0
            i = bunga_p / 100.0
            total_bayar = plaf + (plaf * i * ten) if ten > 0 else plaf
            cic = (total_bayar / ten) if ten > 0 else 0.0
            out.append({
                'id_pinjaman': str(uuid.uuid4()),
                'id_anggota': id_a,
                'nama_anggota': ag.get('nama', ''),
                'no_anggota': ag.get('no_anggota', ''),
                'jenis_pinjaman': 'Saldo (migrasi)',
                'plafon': str(round(plaf, 2)),
                'tenor_bulan': str(ten),
                'bunga_persen': str(bunga_p),
                'total_bayar': str(round(total_bayar, 2)),
                'cicilan_per_bulan': str(round(cic, 2)),
                'sisa_pinjaman': str(round(total_bayar, 2)),
                'tanggal_pengajuan': '-',
                'status': 'Disetujui',
                'tanggal_lunas': '',
            })
        tulis_csv(FILE_PINJAMAN, out, PINJAMAN_FIELDNAMES)
        return

    # Skema transaksi lama (plafon, tenor_bulan, ...)
    agg_tot = {}
    agg_tenor = {}
    for p in rows:
        id_a = (p.get('id_anggota') or '').strip()
        if not id_a:
            continue
        st = p.get('status', '')
        if st not in ('Disetujui', 'Lunas', 'Menunggu'):
            continue
        try:
            plaf = float(p.get('plafon') or 0)
            t = int(float(p.get('tenor_bulan') or 0))
        except (TypeError, ValueError):
            continue
        agg_tot[id_a] = agg_tot.get(id_a, 0.0) + plaf
        agg_tenor[id_a] = max(agg_tenor.get(id_a, 0), t)
    amap = {a.get('id_anggota'): a for a in baca_csv(FILE_ANGGOTA)}
    out = []
    for id_a in sorted(agg_tot.keys()):
        ag = amap.get(id_a) or {}
        plaf = agg_tot[id_a]
        ten = max(agg_tenor.get(id_a, 0), 0)
        bunga_p = bunga_dari_tenor(ten) if ten > 0 else 0.0
        i = bunga_p / 100.0
        total_bayar = plaf + (plaf * i * ten) if ten > 0 else plaf
        cic = (total_bayar / ten) if ten > 0 else 0.0
        out.append({
            'id_pinjaman': str(uuid.uuid4()),
            'id_anggota': id_a,
            'nama_anggota': ag.get('nama', ''),
            'no_anggota': ag.get('no_anggota', ''),
            'jenis_pinjaman': 'Gabungan (migrasi)',
            'plafon': str(round(plaf, 2)),
            'tenor_bulan': str(ten),
            'bunga_persen': str(bunga_p),
            'total_bayar': str(round(total_bayar, 2)),
            'cicilan_per_bulan': str(round(cic, 2)),
            'sisa_pinjaman': str(round(total_bayar, 2)),
            'tanggal_pengajuan': '-',
            'status': 'Disetujui',
            'tanggal_lunas': '',
        })
    tulis_csv(FILE_PINJAMAN, out, PINJAMAN_FIELDNAMES)


def _dedupe_rows_simpanan(rows):
    seen = {}
    for r in rows:
        id_a = (r.get('id_anggota') or '').strip()
        if not id_a:
            continue
        try:
            v = float(r.get('total_simpanan') or 0)
        except (TypeError, ValueError):
            v = 0.0
        seen[id_a] = seen.get(id_a, 0.0) + v
    return [{'id_anggota': k, 'total_simpanan': str(round(v, 2))} for k, v in sorted(seen.items())]


def _dedupe_rows_pinjaman(rows):
    """Satukan baris dengan (id_anggota, jenis_pinjaman) sama — untuk impor CSV."""
    if not rows:
        return []
    if 'jenis_pinjaman' not in rows[0]:
        return rows
    merged = {}
    order = []
    for r in rows:
        id_a = (r.get('id_anggota') or '').strip()
        jn = (r.get('jenis_pinjaman') or JENIS_IMPORT_CSV).strip()
        if not id_a:
            continue
        key = (id_a, jn)
        if key not in merged:
            merged[key] = dict(r)
            order.append(key)
            continue
        m = merged[key]
        try:
            plaf = float(m.get('plafon') or 0) + float(r.get('plafon') or 0)
            ten = max(int(float(m.get('tenor_bulan') or 0)), int(float(r.get('tenor_bulan') or 0)))
        except (TypeError, ValueError):
            continue
        bp = float(m.get('bunga_persen') or 0)
        i = bp / 100.0
        tb = plaf + (plaf * i * ten) if ten > 0 else plaf
        cic = (tb / ten) if ten > 0 else 0.0
        m['plafon'] = str(round(plaf, 2))
        m['tenor_bulan'] = str(ten)
        m['total_bayar'] = str(round(tb, 2))
        m['cicilan_per_bulan'] = str(round(cic, 2))
        if m.get('status') == 'Disetujui':
            m['sisa_pinjaman'] = str(round(tb, 2))
    return [merged[k] for k in order]


def merge_akumulasi(simpanan_rows: list, id_anggota: str, tambah: float) -> None:
    """Akumulasi simpanan per id_anggota (satu baris per anggota)."""
    tambah = max(float(tambah or 0), 0.0)
    for r in simpanan_rows:
        if r.get('id_anggota') == id_anggota:
            cur = float(r.get('total_simpanan') or 0)
            r['total_simpanan'] = str(round(cur + tambah, 2))
            return
    simpanan_rows.append({'id_anggota': id_anggota, 'total_simpanan': str(round(tambah, 2))})


def merge_pinjaman_akumulasi(
    pinjaman_rows: list,
    id_anggota: str,
    tambah_pinjaman: float,
    tenor_baru: int,
    jenis: str = None,
) -> None:
    """Akumulasi pinjaman per (anggota, jenis). Jangka pendek & panjang tidak digabung."""
    jenis_key = (jenis or JENIS_IMPORT_CSV).strip()
    tambah_pinjaman = max(float(tambah_pinjaman or 0), 0.0)
    tenor_baru = max(int(tenor_baru or 0), 0)
    for r in pinjaman_rows:
        if r.get('id_anggota') != id_anggota:
            continue
        if (r.get('jenis_pinjaman') or '').strip() != jenis_key:
            continue
        plaf = float(r.get('plafon') or 0) + tambah_pinjaman
        ten = max(int(float(r.get('tenor_bulan') or 0)), tenor_baru)
        bp = float(r.get('bunga_persen') or 0)
        i = bp / 100.0
        tb = plaf + (plaf * i * ten) if ten > 0 else plaf
        cic = (tb / ten) if ten > 0 else 0.0
        r['plafon'] = str(round(plaf, 2))
        r['tenor_bulan'] = str(ten)
        r['total_bayar'] = str(round(tb, 2))
        r['cicilan_per_bulan'] = str(round(cic, 2))
        if r.get('status') == 'Disetujui':
            r['sisa_pinjaman'] = str(round(tb, 2))
        return
    bp = bunga_dari_tenor(tenor_baru) if tenor_baru > 0 else 0.0
    i = bp / 100.0
    plaf = tambah_pinjaman
    tb = plaf + (plaf * i * tenor_baru) if tenor_baru > 0 else plaf
    cic = (tb / tenor_baru) if tenor_baru > 0 else 0.0
    pinjaman_rows.append({
        'id_pinjaman': str(uuid.uuid4()),
        'id_anggota': id_anggota,
        'nama_anggota': '',
        'no_anggota': '',
        'jenis_pinjaman': jenis_key,
        'plafon': str(round(plaf, 2)),
        'tenor_bulan': str(tenor_baru),
        'bunga_persen': str(bp),
        'total_bayar': str(round(tb, 2)),
        'cicilan_per_bulan': str(round(cic, 2)),
        'sisa_pinjaman': str(round(tb, 2)),
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Disetujui',
        'tanggal_lunas': '',
    })


def get_pinjaman_saldo(pinjaman_rows: list, id_anggota: str):
    for p in pinjaman_rows:
        if p.get('id_anggota') == id_anggota:
            return p
    return None


def cicilan_per_bulan_saldo(p: dict) -> float:
    if p.get('cicilan_per_bulan') not in (None, ''):
        try:
            return float(p['cicilan_per_bulan'])
        except (TypeError, ValueError):
            pass
    tot = float(p.get('sisa_pinjaman') or p.get('total_pinjaman') or 0)
    ten = int(float(p.get('tenor_bulan') or p.get('tenor') or 0))
    if ten <= 0 or tot <= 0:
        return 0.0
    return tot / ten


def enrich_simpanan_untuk_tampilan(simpanan_rows: list, anggota_rows: list) -> list:
    amap = {a.get('id_anggota'): a for a in anggota_rows}
    out = []
    for s in simpanan_rows:
        a = amap.get(s.get('id_anggota'), {})
        out.append({
            **s,
            'no_anggota': a.get('no_anggota', ''),
            'nama_anggota': a.get('nama', ''),
            'jumlah': s.get('total_simpanan', '0'),
            'tanggal': '-',
            'status': 'Disetujui',
            'jenis_simpanan': 'Saldo',
            'keterangan': 'Total saldo per anggota',
            'id_simpanan': s.get('id_anggota', ''),
        })
    return out


def enrich_pinjaman_untuk_tampilan(pinjaman_rows: list, anggota_rows: list) -> list:
    amap = {a.get('id_anggota'): a for a in anggota_rows}
    out = []
    for p in pinjaman_rows:
        a = amap.get(p.get('id_anggota'), {})
        if p.get('plafon') is not None or p.get('id_pinjaman'):
            plaf = float(p.get('plafon') or 0)
            sisa = float(p.get('sisa_pinjaman') or 0)
            cic = float(p.get('cicilan_per_bulan') or 0) or cicilan_per_bulan_saldo(p)
            st = p.get('status') or 'Menunggu'
            out.append({
                **p,
                'no_anggota': p.get('no_anggota') or a.get('no_anggota', ''),
                'nama_anggota': p.get('nama_anggota') or a.get('nama', ''),
                'plafon': str(plaf),
                'tenor_bulan': p.get('tenor_bulan', p.get('tenor', '0')),
                'cicilan_per_bulan': str(round(cic, 2)),
                'sisa_pinjaman': str(round(sisa, 2)),
                'status': st,
                'tanggal_pengajuan': p.get('tanggal_pengajuan') or '-',
                'id_pinjaman': p.get('id_pinjaman', ''),
                'id_anggota': p.get('id_anggota', ''),
            })
            continue
        tot = float(p.get('total_pinjaman') or 0)
        ten = int(float(p.get('tenor') or 0))
        cic = cicilan_per_bulan_saldo(p)
        st = 'Disetujui' if tot > 0 else 'Lunas'
        out.append({
            **p,
            'no_anggota': a.get('no_anggota', ''),
            'nama_anggota': a.get('nama', ''),
            'plafon': p.get('total_pinjaman', '0'),
            'tenor_bulan': p.get('tenor', '0'),
            'cicilan_per_bulan': str(round(cic, 2)),
            'sisa_pinjaman': str(round(tot, 2)),
            'status': st,
            'tanggal_pengajuan': '-',
            'jenis_pinjaman': 'Saldo',
            'id_pinjaman': p.get('id_anggota', ''),
        })
    return out


def ensure_anggota_schema():
    """Pastikan data anggota punya kolom nik, penghasilan, dan cicilan lain."""
    rows = baca_csv(FILE_ANGGOTA)
    if not rows:
        return
    if 'nik' in rows[0] and 'penghasilan_bersih' in rows[0] and 'cicilan_lain' in rows[0]:
        return
    for r in rows:
        r['nik'] = r.get('nik', '')
        r['penghasilan_bersih'] = r.get('penghasilan_bersih', '0')
        r['cicilan_lain'] = r.get('cicilan_lain', '0')
    tulis_csv(FILE_ANGGOTA, rows, ANGGOTA_FIELDNAMES)


def ensure_pinjaman_plafon_schema():
    """Kompatibilitas: migrasi skema lama ke saldo per id_anggota."""
    migrate_pinjaman_ke_saldo()


def ensure_pinjaman_cicilan_schema():
    """Pastikan file pinjaman_cicilan.csv memiliki fieldnames terbaru."""
    if not os.path.exists(FILE_PINJAMAN_CICILAN):
        tulis_csv(FILE_PINJAMAN_CICILAN, [], CICILAN_FIELDNAMES)
        return
    rows = baca_csv(FILE_PINJAMAN_CICILAN)
    normalized = []
    for r in rows:
        normalized.append({k: (r.get(k) or '') for k in CICILAN_FIELDNAMES})
    tulis_csv(FILE_PINJAMAN_CICILAN, normalized, CICILAN_FIELDNAMES)


def ensure_import_log_schema():
    if not os.path.exists(FILE_IMPORT_LOG):
        tulis_csv(
            FILE_IMPORT_LOG,
            [],
            ['waktu', 'user', 'mode', 'nama_file', 'berhasil', 'gagal', 'catatan'],
        )


def ensure_import_preview_dir():
    os.makedirs(IMPORT_PREVIEW_DIR, exist_ok=True)


def ensure_pendaftaran_schema():
    """Migrasi kolom pengajuan anggota (penghasilan, cicilan, id anggota dibuat, dll.)."""
    if not os.path.exists(FILE_PENDAFTARAN_ANGGOTA):
        return
    rows = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    if rows:
        sample = rows[0]
        if all(k in sample for k in PENDAFTARAN_FIELDNAMES):
            return
        for r in rows:
            for k in PENDAFTARAN_FIELDNAMES:
                if k not in r:
                    r[k] = '0' if k in ('penghasilan_bersih', 'cicilan_lain') else ''
        tulis_csv(FILE_PENDAFTARAN_ANGGOTA, rows, PENDAFTARAN_FIELDNAMES)
        return
    with open(FILE_PENDAFTARAN_ANGGOTA, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        old_fields = reader.fieldnames or []
    if all(k in old_fields for k in PENDAFTARAN_FIELDNAMES):
        return
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, [], PENDAFTARAN_FIELDNAMES)


ensure_pendaftaran_schema()

PASSWORD_POLICY_MSG = (
    'Password minimal 8 karakter dan wajib memuat huruf besar (A-Z), huruf kecil (a-z), '
    'angka (0-9), dan simbol khusus (contoh: !@#$).'
)


def validate_password_policy(password: str):
    """Minimal 8 karakter, kombinasi huruf besar/kecil, angka, dan simbol khusus."""
    if not password or len(password) < 8:
        return False, PASSWORD_POLICY_MSG
    if not re.search(r'[A-Z]', password):
        return False, PASSWORD_POLICY_MSG
    if not re.search(r'[a-z]', password):
        return False, PASSWORD_POLICY_MSG
    if not re.search(r'\d', password):
        return False, PASSWORD_POLICY_MSG
    simbol = set('!@#$%^&*()_+-=[]{}|;:,.<>?/`~')
    if not any(c in password for c in simbol):
        return False, PASSWORD_POLICY_MSG
    return True, ''


def get_anggota_by_id(id_anggota: str):
    for a in baca_csv(FILE_ANGGOTA):
        if a.get('id_anggota') == id_anggota:
            return a
    return None


def generate_no_anggota_berikutnya(data_anggota):
    """Generate nomor AGT-XXXX berbasis nomor tertinggi yang sudah ada."""
    max_no = 0
    for a in data_anggota:
        no = (a.get('no_anggota') or '').strip()
        if no.startswith('AGT-'):
            bagian = no.replace('AGT-', '')
            if bagian.isdigit():
                max_no = max(max_no, int(bagian))
    return f"AGT-{(max_no + 1):04d}"


def ensure_simpanan_schema():
    """Kompatibilitas: migrasi skema lama ke saldo per id_anggota."""
    migrate_simpanan_ke_saldo()


def bunga_dari_tenor(tenor_bulan: int) -> float:
    """Skema bunga per bulan.
    1-3 bulan   -> 2.0%
    4-12 bulan  -> 1.5%
    13-24 bulan -> 1.0%
    """
    if tenor_bulan <= 0:
        return 0.0
    if 1 <= tenor_bulan <= 3:
        return 2.0
    if 4 <= tenor_bulan <= 12:
        return 1.5
    if 13 <= tenor_bulan <= 24:
        return 1.0
    raise ValueError('Tenor harus antara 1 sampai 24 bulan.')


migrate_simpanan_ke_saldo()
migrate_pinjaman_ke_saldo()
ensure_pinjaman_cicilan_schema()
ensure_anggota_schema()
ensure_import_log_schema()
ensure_import_preview_dir()


def hitung_kapasitas_cicilan(penghasilan_bersih: float, cicilan_lain: float, dsr: float) -> float:
    return max((penghasilan_bersih * dsr) - cicilan_lain, 0.0)


def dsr_otomatis_dari_penghasilan(penghasilan_bersih: float) -> float:
    """DSR 30–40% mengikuti penghasilan bersih (otomatis, tanpa pilihan manual)."""
    if penghasilan_bersih <= 0:
        return DSR_DEFAULT
    if penghasilan_bersih <= 5_000_000:
        return 0.40
    if penghasilan_bersih <= 15_000_000:
        return 0.35
    return 0.30


def hitung_plafon_maks_anuitas(cicilan_bulanan: float, bunga_persen_bulanan: float, tenor_bulan: int) -> float:
    i = bunga_persen_bulanan / 100.0
    n = tenor_bulan
    if n <= 0:
        raise ValueError('Tenor harus lebih dari 0.')
    if i <= 0:
        return cicilan_bulanan * n
    return (cicilan_bulanan * (1 - (1 + i) ** (-n))) / i


def info_pinjaman_dsr_anggota(anggota_row: dict) -> dict:
    """Ringkasan DSR otomatis dan kapasitas cicilan untuk satu baris anggota (tampilan admin)."""
    ph = max(float(anggota_row.get('penghasilan_bersih') or 0), 0.0)
    cl = max(float(anggota_row.get('cicilan_lain') or 0), 0.0)
    dsr = dsr_otomatis_dari_penghasilan(ph)
    kap = hitung_kapasitas_cicilan(ph, cl, dsr)
    return {
        'dsr_persen': int(round(dsr * 100)),
        'kapasitas_cicilan': kap,
    }


def parse_rupiah_to_float(value):
    return float(str(value or '0').replace(',', '').replace('.', ''))


def get_user_by_username(username: str):
    users = baca_csv(FILE_USERS)
    for u in users:
        if u.get('username', '').lower() == (username or '').lower():
            return u
    return None


def get_user_by_id(id_user: str):
    users = baca_csv(FILE_USERS)
    for u in users:
        if u.get('id_user') == id_user:
            return u
    return None


def get_current_user_id_anggota():
    return session.get('id_anggota') or ''


def is_current_user_admin():
    return session.get('role') == 'admin'


def ensure_users_schema():
    """Pastikan users.csv punya kolom id_anggota (upgrade schema lama)."""
    users = baca_csv(FILE_USERS)
    if not users:
        return
    if 'id_anggota' in users[0]:
        return
    for u in users:
        u['id_anggota'] = ''
    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


def ensure_csrf_token() -> str:
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


def validate_csrf_token() -> bool:
    token_form = request.form.get('csrf_token') or request.headers.get('X-CSRF-Token')
    token_session = session.get('_csrf_token') or ''
    if not token_form or not token_session:
        return False
    return hmac.compare_digest(token_form, token_session)


def csrf_protect(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not validate_csrf_token():
            flash('Permintaan ditolak: token keamanan (CSRF) tidak valid.', 'danger')
            return redirect(request.referrer or url_for('dashboard'))
        return view_func(*args, **kwargs)
    return wrapper


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        if session.get('role') != 'admin':
            abort(403)
        return view_func(*args, **kwargs)
    return wrapper


def restrict_id_anggota_or_forbid(id_anggota: str):
    """User biasa hanya boleh akses data id_anggota dirinya. Admin boleh semua."""
    if is_current_user_admin():
        return
    current = get_current_user_id_anggota()
    if not current or current != id_anggota:
        abort(403)


@app.context_processor
def inject_globals():
    return {
        'now': datetime.now().strftime('%d %B %Y'),
        'current_user': session.get('user'),
        'current_role': session.get('role'),
        'is_admin': is_current_user_admin(),
        'current_id_anggota': get_current_user_id_anggota(),
        'csrf_token': ensure_csrf_token(),
    }


# ══════════════════════════════════════════════
#  ROUTE: LANDING & AUTH
# ══════════════════════════════════════════════
@app.route('/')
def landing():
    if session.get('user'):
        return redirect(url_for('dashboard'))
    ensure_simpanan_schema()
    ensure_pinjaman_plafon_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    simpanan = baca_csv(FILE_SIMPANAN)
    pinjaman = baca_csv(FILE_PINJAMAN)
    n_anggota = len(anggota)
    total_simpanan_disetujui = sum(float(s.get('total_simpanan') or 0) for s in simpanan)
    total_pinjaman_disalurkan = sum(
        float(p.get('plafon') or p.get('total_pinjaman') or 0)
        for p in pinjaman
        if p.get('status') in ('Disetujui', 'Lunas')
    )
    # Estimasi ilustratif (bukan pembukuan resmi SHU): proporsi dari total simpanan disetujui
    shu_estimasi_ilustratif = int(total_simpanan_disetujui * 0.08)
    return render_template(
        'landing.html',
        n_anggota=n_anggota,
        total_pinjaman_disalurkan=total_pinjaman_disalurkan,
        shu_estimasi_ilustratif=shu_estimasi_ilustratif,
        total_simpanan_disetujui=total_simpanan_disetujui,
    )


@app.route('/pendaftaran-anggota', methods=['POST'])
@csrf_protect
def pengajuan_anggota_baru():
    """Pengajuan anggota baru dari landing page."""
    ensure_pendaftaran_schema()
    nama = (request.form.get('nama') or '').strip()
    alamat = (request.form.get('alamat') or '').strip()
    no_telp = (request.form.get('no_telp') or '').strip()
    penghasilan = parse_rupiah_to_float(request.form.get('penghasilan_bersih', '0'))
    cicilan_lain = parse_rupiah_to_float(request.form.get('cicilan_lain', '0'))

    if not nama or not alamat:
        flash('Nama dan alamat wajib diisi untuk pengajuan anggota.', 'danger')
        return redirect(url_for('landing'))
    if penghasilan <= 0:
        flash('Penghasilan bersih per bulan wajib diisi (lebih dari Rp 0) untuk pengajuan anggota.', 'danger')
        return redirect(url_for('landing'))

    data = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    data.append({
        'id_pengajuan': str(uuid.uuid4()),
        'nama': nama,
        'alamat': alamat,
        'no_telp': no_telp,
        'penghasilan_bersih': str(int(penghasilan)),
        'cicilan_lain': str(int(cicilan_lain)),
        'status': 'Menunggu',
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'catatan_admin': '',
        'id_anggota_dibuat': '',
        'no_anggota_dibuat': ''
    })
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, data, PENDAFTARAN_FIELDNAMES)
    flash('Pengajuan anggota berhasil dikirim. Mohon tunggu konfirmasi admin.', 'success')
    return redirect(url_for('landing'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    ensure_users_schema()
    if request.method == 'POST':
        if not validate_csrf_token():
            flash('Permintaan ditolak: token keamanan (CSRF) tidak valid.', 'danger')
            return render_template('login.html', next=request.form.get('next') or url_for('dashboard')), 400
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        next_url = request.form.get('next') or url_for('dashboard')

        user = get_user_by_username(username)
        if not user or not check_password_hash(user.get('password_hash', ''), password):
            flash('Username atau password salah.', 'danger')
            return render_template('login.html', next=next_url), 401

        session['user'] = user.get('username')
        session['role'] = user.get('role')
        session['id_user'] = user.get('id_user')
        session['id_anggota'] = user.get('id_anggota', '')
        if session.get('role') == 'user' and not session.get('id_anggota'):
            session.clear()
            flash('Akun ini belum dihubungkan ke anggota. Hubungi admin.', 'danger')
            return render_template('login.html', next=next_url), 403
        flash(f"Selamat datang, {session['user']}!", 'success')
        return redirect(next_url)

    if session.get('user'):
        return redirect(url_for('dashboard'))
    return render_template('login.html', next=request.args.get('next', url_for('dashboard')))


@app.route('/lupa-password', methods=['GET', 'POST'])
def lupa_password():
    """Reset password mandiri untuk akun user (verifikasi nomor anggota yang terhubung)."""
    ensure_users_schema()
    if request.method == 'POST':
        if not validate_csrf_token():
            flash('Permintaan ditolak: token keamanan (CSRF) tidak valid.', 'danger')
            return render_template('lupa_password.html'), 400
        username = (request.form.get('username') or '').strip()
        no_anggota = (request.form.get('no_anggota') or '').strip()
        pwd1 = request.form.get('password') or ''
        pwd2 = request.form.get('password_confirm') or ''

        user = get_user_by_username(username)
        if not user or user.get('role') != 'user' or not user.get('id_anggota'):
            flash('Data tidak cocok atau akun tidak memenuhi syarat reset mandiri. Hubungi admin.', 'danger')
            return render_template('lupa_password.html'), 400

        anggota = get_anggota_by_id(user.get('id_anggota'))
        if (not anggota or
                (anggota.get('no_anggota') or '').strip().upper() != no_anggota.strip().upper()):
            flash('Data tidak cocok atau akun tidak memenuhi syarat reset mandiri. Hubungi admin.', 'danger')
            return render_template('lupa_password.html'), 400

        if pwd1 != pwd2:
            flash('Konfirmasi password tidak sama.', 'danger')
            return render_template('lupa_password.html'), 400

        ok, msg = validate_password_policy(pwd1)
        if not ok:
            flash(msg, 'danger')
            return render_template('lupa_password.html'), 400

        users = baca_csv(FILE_USERS)
        for u in users:
            if u.get('id_user') == user.get('id_user'):
                u['password_hash'] = generate_password_hash(pwd1)
                break
        tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
        flash('Password berhasil diubah. Silakan login.', 'success')
        return redirect(url_for('login'))

    return render_template('lupa_password.html')


@app.route('/logout')
@login_required
@csrf_protect
def logout():
    session.clear()
    flash('Anda berhasil logout.', 'success')
    return redirect(url_for('landing'))


# ══════════════════════════════════════════════
#  ROUTE: MANAJEMEN USER (ADMIN)
# ══════════════════════════════════════════════
@app.route('/users')
@admin_required
def users_index():
    ensure_users_schema()
    users = baca_csv(FILE_USERS)
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_map = {a.get('id_anggota'): a for a in anggota}
    users.sort(key=lambda u: (u.get('role') != 'admin', (u.get('username') or '').lower()))
    return render_template('users.html', users=users, anggota=anggota, anggota_map=anggota_map)


@app.route('/users/tambah', methods=['POST'])
@admin_required
@csrf_protect
def users_tambah():
    ensure_users_schema()
    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    role = request.form.get('role') or 'user'
    id_anggota = request.form.get('id_anggota') or ''

    if not username or not password:
        flash('Username dan password wajib diisi.', 'danger')
        return redirect(url_for('users_index'))
    ok_pol, msg_pol = validate_password_policy(password)
    if not ok_pol:
        flash(msg_pol, 'danger')
        return redirect(url_for('users_index'))
    if role not in ('admin', 'user'):
        flash('Role tidak valid.', 'danger')
        return redirect(url_for('users_index'))
    if get_user_by_username(username):
        flash('Username sudah dipakai. Gunakan username lain.', 'danger')
        return redirect(url_for('users_index'))

    if role == 'user' and not id_anggota:
        flash('Untuk role user, wajib pilih anggota.', 'danger')
        return redirect(url_for('users_index'))

    if id_anggota:
        anggota = baca_csv(FILE_ANGGOTA)
        if not any(a.get('id_anggota') == id_anggota for a in anggota):
            flash('Anggota tidak ditemukan.', 'danger')
            return redirect(url_for('users_index'))

    users = baca_csv(FILE_USERS)
    users.append({
        'id_user': str(uuid.uuid4()),
        'username': username,
        'password_hash': generate_password_hash(password),
        'role': role,
        'id_anggota': id_anggota if role == 'user' else '',
        'created_at': datetime.now().strftime('%Y-%m-%d')
    })
    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    flash('Akun berhasil dibuat.', 'success')
    return redirect(url_for('users_index'))


@app.route('/users/reset_password/<id_user>', methods=['POST'])
@admin_required
@csrf_protect
def users_reset_password(id_user):
    ensure_users_schema()
    new_password = request.form.get('new_password') or ''
    if not new_password:
        flash('Password baru wajib diisi.', 'danger')
        return redirect(url_for('users_index'))
    ok_pol, msg_pol = validate_password_policy(new_password)
    if not ok_pol:
        flash(msg_pol, 'danger')
        return redirect(url_for('users_index'))

    users = baca_csv(FILE_USERS)
    found = False
    for u in users:
        if u.get('id_user') == id_user:
            u['password_hash'] = generate_password_hash(new_password)
            found = True
            break
    if not found:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('users_index'))

    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    flash('Password berhasil direset.', 'success')
    return redirect(url_for('users_index'))


@app.route('/users/hapus/<id_user>', methods=['POST'])
@admin_required
@csrf_protect
def users_hapus(id_user):
    ensure_users_schema()
    current_id = session.get('id_user')
    if current_id and current_id == id_user:
        flash('Tidak bisa menghapus akun yang sedang login.', 'danger')
        return redirect(url_for('users_index'))

    users = baca_csv(FILE_USERS)
    before = len(users)
    users = [u for u in users if u.get('id_user') != id_user]
    if len(users) == before:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('users_index'))
    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    flash('User berhasil dihapus.', 'success')
    return redirect(url_for('users_index'))


# ──────────────────────────────────────────────
#  KONFIGURASI PINJAMAN
# ──────────────────────────────────────────────
JENIS_PINJAMAN = {
    'Jangka Panjang': {'tenor_maks': 24},
    'Jangka Menengah': {'tenor_maks': 12},
    'Solusi Cepat': {'tenor_maks': 3},
}

JENIS_SIMPANAN = ['Manasuka', 'Hari Raya', 'Pendidikan']


def bunga_untuk_jenis_pinjaman(jenis: str, tenor_bulan: int) -> float:
    """Bunga bulanan mengikuti tenor agar konsisten lintas produk."""
    return bunga_dari_tenor(tenor_bulan)


# ══════════════════════════════════════════════
#  ROUTE: DASHBOARD
# ══════════════════════════════════════════════
@app.route('/dashboard')
@login_required
def dashboard():
    ensure_simpanan_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    simpanan = baca_csv(FILE_SIMPANAN)
    pinjaman = baca_csv(FILE_PINJAMAN)

    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]
        simpanan = [s for s in simpanan if s.get('id_anggota') == id_anggota]
        pinjaman = [p for p in pinjaman if p.get('id_anggota') == id_anggota]

    total_anggota = len(anggota)
    total_simpanan = sum(float(s.get('total_simpanan') or 0) for s in simpanan)
    pv_aktif = [p for p in pinjaman if p.get('status') == 'Disetujui']
    total_pinjaman = sum(float(p.get('plafon') or 0) for p in pv_aktif)
    total_pinjaman_beredar = sum(float(p.get('sisa_pinjaman') or 0) for p in pv_aktif)

    # Saldo simpanan (skema per anggota; tampilan dashboard mengikuti jenis lama — alokasi ke Manasuka)
    simpanan_per_jenis = {j: 0.0 for j in JENIS_SIMPANAN}
    simpanan_per_jenis['Manasuka'] = total_simpanan

    # Ringkasan aktivitas: anggota terbaru
    anggota_sorted = sorted(anggota, key=lambda x: x.get('tgl_bergabung', ''), reverse=True)
    transaksi_terakhir = []
    for a in anggota_sorted[:6]:
        transaksi_terakhir.append({
            'tipe': 'Anggota',
            'tanggal': a.get('tgl_bergabung', '-'),
            'nama': a.get('nama', ''),
            'keterangan': f"Terdaftar — {a.get('no_anggota', '')}",
        })

    pengajuan_anggota = []
    jumlah_pengajuan_menunggu = 0
    if is_current_user_admin():
        pengajuan_anggota = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
        pengajuan_anggota.sort(key=lambda x: x.get('tanggal_pengajuan', ''), reverse=True)
        jumlah_pengajuan_menunggu = sum(1 for p in pengajuan_anggota if p.get('status') == 'Menunggu')
        pengajuan_anggota = pengajuan_anggota[:8]

    return render_template('dashboard.html',
                           total_anggota=total_anggota,
                           total_simpanan=total_simpanan,
                           total_pinjaman=total_pinjaman,
                           total_pinjaman_beredar=total_pinjaman_beredar,
                           simpanan_per_jenis=simpanan_per_jenis,
                           transaksi_terakhir=transaksi_terakhir,
                           pengajuan_anggota=pengajuan_anggota,
                           jumlah_pengajuan_menunggu=jumlah_pengajuan_menunggu)


@app.route('/pengajuan-anggota/konfirmasi/<id_pengajuan>', methods=['POST'])
@admin_required
@csrf_protect
def konfirmasi_pengajuan_anggota(id_pengajuan):
    ensure_pendaftaran_schema()
    pengajuan = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    found = False
    for p in pengajuan:
        if p.get('id_pengajuan') == id_pengajuan:
            # Auto-create anggota hanya sekali saat pengajuan pertama kali disetujui.
            if not p.get('id_anggota_dibuat'):
                anggota = baca_csv(FILE_ANGGOTA)
                id_anggota = str(uuid.uuid4())
                no_anggota = generate_no_anggota_berikutnya(anggota)
                ph = parse_rupiah_to_float(p.get('penghasilan_bersih', '0'))
                cl = parse_rupiah_to_float(p.get('cicilan_lain', '0'))
                anggota.append({
                    'id_anggota': id_anggota,
                    'no_anggota': no_anggota,
                    'nik': '',
                    'nama': p.get('nama', ''),
                    'alamat': p.get('alamat', ''),
                    'no_telp': p.get('no_telp', ''),
                    'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
                    'penghasilan_bersih': str(int(ph)),
                    'cicilan_lain': str(int(cl))
                })
                tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
                p['id_anggota_dibuat'] = id_anggota
                p['no_anggota_dibuat'] = no_anggota
            p['status'] = 'Disetujui'
            p['catatan_admin'] = (
                f"Dikonfirmasi {session.get('user', 'admin')} pada {datetime.now().strftime('%Y-%m-%d')} "
                f"(anggota: {p.get('no_anggota_dibuat', '-')})"
            )
            found = True
            break
    if not found:
        flash('Pengajuan anggota tidak ditemukan.', 'danger')
        return redirect(url_for('dashboard'))
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, pengajuan, PENDAFTARAN_FIELDNAMES)
    flash('Pengajuan anggota disetujui.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/pengajuan-anggota/tolak/<id_pengajuan>', methods=['POST'])
@admin_required
@csrf_protect
def tolak_pengajuan_anggota(id_pengajuan):
    ensure_pendaftaran_schema()
    pengajuan = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    found = False
    for p in pengajuan:
        if p.get('id_pengajuan') == id_pengajuan:
            p['status'] = 'Ditolak'
            p['catatan_admin'] = f"Ditolak {session.get('user', 'admin')} pada {datetime.now().strftime('%Y-%m-%d')}"
            found = True
            break
    if not found:
        flash('Pengajuan anggota tidak ditemukan.', 'danger')
        return redirect(url_for('dashboard'))
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, pengajuan, PENDAFTARAN_FIELDNAMES)
    flash('Pengajuan anggota ditolak.', 'warning')
    return redirect(url_for('dashboard'))


# ══════════════════════════════════════════════
#  ROUTE: ANGGOTA
# ══════════════════════════════════════════════
@app.route('/anggota')
@login_required
def halaman_anggota():
    ensure_anggota_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]
    else:
        for a in anggota:
            info = info_pinjaman_dsr_anggota(a)
            a['_dsr_persen'] = info['dsr_persen']
            a['_kap_cicilan'] = info['kapasitas_cicilan']
    return render_template('anggota.html', anggota=anggota)


@app.route('/anggota/tambah', methods=['POST'])
@admin_required
@csrf_protect
def tambah_anggota():
    ensure_anggota_schema()
    data = baca_csv(FILE_ANGGOTA)
    no_anggota = generate_no_anggota_berikutnya(data)
    id_anggota = str(uuid.uuid4())

    data.append({
        'id_anggota': id_anggota,
        'no_anggota': no_anggota,
        'nik': (request.form.get('nik') or '').strip(),
        'nama': request.form['nama'],
        'alamat': request.form['alamat'],
        'no_telp': request.form['no_telp'],
        'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
        'penghasilan_bersih': str(parse_rupiah_to_float(request.form.get('penghasilan_bersih', '0'))),
        'cicilan_lain': str(parse_rupiah_to_float(request.form.get('cicilan_lain', '0')))
    })
    tulis_csv(FILE_ANGGOTA, data, ANGGOTA_FIELDNAMES)
    return redirect('/anggota')


@app.route('/anggota/hapus/<id_anggota>', methods=['POST'])
@admin_required
@csrf_protect
def hapus_anggota(id_anggota):
    data = baca_csv(FILE_ANGGOTA)
    data = [a for a in data if a['id_anggota'] != id_anggota]
    tulis_csv(FILE_ANGGOTA, data, ANGGOTA_FIELDNAMES)
    return redirect('/anggota')


@app.route('/anggota/edit/<id_anggota>', methods=['GET', 'POST'])
@admin_required
def edit_anggota(id_anggota):
    ensure_anggota_schema()
    data = baca_csv(FILE_ANGGOTA)
    idx = next((i for i, a in enumerate(data) if a.get('id_anggota') == id_anggota), None)
    if idx is None:
        flash('Anggota tidak ditemukan.', 'danger')
        return redirect(url_for('halaman_anggota'))

    if request.method == 'POST':
        if not validate_csrf_token():
            flash('Permintaan ditolak: token keamanan (CSRF) tidak valid.', 'danger')
            return redirect(url_for('halaman_anggota'))
        data[idx]['nik'] = (request.form.get('nik') or '').strip()
        data[idx]['nama'] = (request.form.get('nama') or '').strip()
        data[idx]['alamat'] = (request.form.get('alamat') or '').strip()
        data[idx]['no_telp'] = (request.form.get('no_telp') or '').strip()
        data[idx]['penghasilan_bersih'] = str(int(parse_rupiah_to_float(request.form.get('penghasilan_bersih', '0'))))
        data[idx]['cicilan_lain'] = str(int(parse_rupiah_to_float(request.form.get('cicilan_lain', '0'))))
        tulis_csv(FILE_ANGGOTA, data, ANGGOTA_FIELDNAMES)
        flash('Data anggota berhasil diperbarui.', 'success')
        return redirect(url_for('halaman_anggota'))

    return render_template('anggota_edit.html', anggota=data[idx])


@app.route('/anggota/import-excel', methods=['POST'])
@admin_required
@csrf_protect
def import_anggota_excel():
    """Import anggota dari Excel dengan smart upsert berdasarkan NIK, fallback No Anggota."""
    ensure_anggota_schema()
    if load_workbook is None:
        flash('Fitur import Excel membutuhkan openpyxl.', 'danger')
        return redirect(url_for('halaman_anggota'))

    file = request.files.get('file_excel')
    if not file or not file.filename:
        flash('File Excel belum dipilih.', 'danger')
        return redirect(url_for('halaman_anggota'))
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('Format file tidak didukung. Gunakan .xlsx atau .xlsm', 'danger')
        return redirect(url_for('halaman_anggota'))

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        flash(f'Gagal membaca file Excel: {e}', 'danger')
        return redirect(url_for('halaman_anggota'))

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        flash('File Excel kosong.', 'danger')
        return redirect(url_for('halaman_anggota'))

    def norm(h):
        return str(h or '').strip().lower().replace(' ', '_')

    header_map = {norm(h): i for i, h in enumerate(headers_raw)}

    def get_val(row, *aliases):
        for a in aliases:
            idx = header_map.get(a)
            if idx is not None and idx < len(row):
                v = row[idx]
                return '' if v is None else str(v).strip()
        return ''

    anggota = baca_csv(FILE_ANGGOTA)
    by_nik = {(a.get('nik') or '').strip(): a for a in anggota if (a.get('nik') or '').strip()}
    by_no = {(a.get('no_anggota') or '').strip(): a for a in anggota if (a.get('no_anggota') or '').strip()}

    added = 0
    updated = 0
    max_no = 0
    for a in anggota:
        no = (a.get('no_anggota') or '').strip()
        if no.startswith('AGT-'):
            num = no.replace('AGT-', '')
            if num.isdigit():
                max_no = max(max_no, int(num))

    for row in rows_iter:
        nik = get_val(row, 'nik')
        no_anggota = get_val(row, 'no_anggota')
        nama = get_val(row, 'nama')
        alamat = get_val(row, 'alamat')
        no_telp = get_val(row, 'no_telp', 'no_telepon', 'telepon')
        penghasilan = get_val(row, 'penghasilan_bersih')
        cicilan = get_val(row, 'cicilan_lain')
        if not (nik or no_anggota or nama):
            continue

        target = None
        if nik and nik in by_nik:
            target = by_nik[nik]
        elif no_anggota and no_anggota in by_no:
            target = by_no[no_anggota]

        if target:
            target['nik'] = nik or target.get('nik', '')
            target['nama'] = nama or target.get('nama', '')
            target['alamat'] = alamat or target.get('alamat', '')
            target['no_telp'] = no_telp or target.get('no_telp', '')
            target['penghasilan_bersih'] = str(parse_rupiah_to_float(penghasilan or target.get('penghasilan_bersih', '0')))
            target['cicilan_lain'] = str(parse_rupiah_to_float(cicilan or target.get('cicilan_lain', '0')))
            updated += 1
            continue

        if not no_anggota:
            max_no += 1
            no_anggota = f"AGT-{max_no:04d}"

        new_row = {
            'id_anggota': str(uuid.uuid4()),
            'no_anggota': no_anggota,
            'nik': nik,
            'nama': nama,
            'alamat': alamat,
            'no_telp': no_telp,
            'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
            'penghasilan_bersih': str(parse_rupiah_to_float(penghasilan or '0')),
            'cicilan_lain': str(parse_rupiah_to_float(cicilan or '0')),
        }
        anggota.append(new_row)
        if nik:
            by_nik[nik] = new_row
        if no_anggota:
            by_no[no_anggota] = new_row
        added += 1

    tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
    flash(f'Import selesai: {added} anggota baru ditambahkan, {updated} data anggota diperbarui.', 'success')
    return redirect(url_for('halaman_anggota'))


def _norm_csv_header(h):
    return str(h or '').strip().lower().replace(' ', '_').replace('-', '_')


def _map_import_csv_headers(fieldnames):
    """Map normalized header -> canonical key."""
    if not fieldnames:
        return None
    norm = {_norm_csv_header(h): h for h in fieldnames}
    aliases = {
        'nama': ('nama', 'name'),
        'nik': ('nik',),
        'no_hp': ('no_hp', 'nohp', 'no_telp', 'telepon', 'telp', 'hp'),
        'simpanan': ('simpanan', 'tabungan'),
        'pinjaman': ('pinjaman', 'hutang'),
        'tenor_bulan': ('tenor_bulan', 'tenor', 'jangka_waktu', 'lama_bulan'),
    }
    out = {}
    for key, alist in aliases.items():
        for a in alist:
            if a in norm:
                out[key] = norm[a]
                break
    if 'nik' not in out:
        return None
    # Hindari file "satu kolom".
    if len(out.keys()) < 2:
        return None
    return out


def _parse_opt_float_cell(raw):
    s = (raw or '').strip()
    if not s:
        return None
    return parse_rupiah_to_float(s)


def excel_bytes_ke_baris_dict(raw: bytes) -> tuple:
    """Baca lembar aktif Excel: (daftar kunci header ter-normalisasi, list dict per baris)."""
    if load_workbook is None:
        raise ValueError('Import Excel membutuhkan openpyxl. Pasang: pip install openpyxl')
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [str(h or '').strip() for h in all_rows[0]]
    header_non_empty = [h for h in headers if h]
    if len(header_non_empty) < 2:
        raise ValueError(
            'Format Excel tidak valid: header harus terpisah di beberapa kolom, bukan satu kolom gabungan.'
        )
    header_norm_list = [_norm_csv_header(h) for h in headers if str(h or '').strip()]
    rows_src = []
    for rr in all_rows[1:]:
        row_map = {}
        for i, h in enumerate(headers):
            if not str(h or '').strip():
                continue
            key = _norm_csv_header(h)
            v = rr[i] if i < len(rr) else ''
            row_map[key] = '' if v is None else str(v).strip()
        isi_terisi = [v for v in row_map.values() if str(v or '').strip()]
        if len(isi_terisi) == 1 and any(sep in isi_terisi[0] for sep in (',', ';', '|', '\t')):
            raise ValueError('Format Excel tidak valid: data terdeteksi gabungan dalam satu kolom. Pisahkan per kolom.')
        rows_src.append(row_map)
    return header_norm_list, rows_src


def parse_anggota_csv_upload(file_storage):
    """Validasi & parse file Excel ringkasan anggota. Mengembalikan dict preview atau raise ValueError."""
    if not file_storage or not file_storage.filename:
        raise ValueError('File belum dipilih.')
    ext = os.path.splitext(file_storage.filename.lower())[1]
    if ext not in ('.xlsx', '.xlsm'):
        raise ValueError('Gunakan file Excel .xlsx atau .xlsm')
    file_storage.stream.seek(0)
    raw = file_storage.read()
    if len(raw) > 2 * 1024 * 1024:
        raise ValueError('Ukuran file maksimal 2 MB.')
    fieldnames, rows_src = excel_bytes_ke_baris_dict(raw)
    if not rows_src:
        raise ValueError('File kosong atau tidak ada baris data.')
    header_map = _map_import_csv_headers(fieldnames)
    if not header_map:
        raise ValueError(
            'Header tidak valid. Wajib ada kolom nik; disarankan: nama,nik,no hp,simpanan,pinjaman,tenor'
        )
    detected_columns = [k for k in ('nama', 'nik', 'no_hp', 'simpanan', 'pinjaman', 'tenor_bulan') if k in header_map]
    rows_out = []
    line_no = 1
    for row in rows_src:
        line_no += 1
        # Baris dari Excel memakai kunci header ter-normalisasi (nama, nik, tenor, ...)
        nama = (row.get('nama') or '').strip() if 'nama' in header_map else ''
        nik = (row.get('nik') or '').strip() if 'nik' in header_map else ''
        no_hp = (row.get('no_hp') or '').strip() if 'no_hp' in header_map else ''
        simp_raw = row.get('simpanan', '') if 'simpanan' in header_map else ''
        pin_raw = row.get('pinjaman', '') if 'pinjaman' in header_map else ''
        tenor_raw = ''
        if 'tenor_bulan' in header_map:
            tenor_raw = (row.get('tenor_bulan') or row.get('tenor') or '')
        err = []
        if not nik:
            err.append('NIK wajib diisi')
        simpanan = None
        pinjaman = None
        tenor_bulan = None
        try:
            if (simp_raw or '').strip():
                simpanan = _parse_opt_float_cell(simp_raw)
        except Exception:
            err.append('simpanan bukan angka valid')
        try:
            if (pin_raw or '').strip():
                pinjaman = _parse_opt_float_cell(pin_raw)
        except Exception:
            err.append('pinjaman bukan angka valid')
        try:
            if (tenor_raw or '').strip():
                tenor_bulan = int(float(str(tenor_raw).strip()))
                if tenor_bulan < 1 or tenor_bulan > 24:
                    err.append('tenor harus 1-24 bulan')
        except Exception:
            err.append('tenor bukan angka valid')
        rows_out.append({
            'line': line_no,
            'nama': nama,
            'nik': nik,
            'no_hp': no_hp,
            'simpanan': simpanan,
            'pinjaman': pinjaman,
            'tenor_bulan': tenor_bulan,
            'errors': err,
        })
    if not rows_out:
        raise ValueError('Tidak ada baris data (selain header).')
    if len(rows_out) > 500:
        raise ValueError('Maksimal 500 baris data per unggah.')
    return {'rows': rows_out, 'filename': file_storage.filename, 'detected_columns': detected_columns}


def _append_import_log(berhasil: int, gagal: int, mode: str, nama_file: str, catatan: str):
    ensure_import_log_schema()
    rows = baca_csv(FILE_IMPORT_LOG)
    rows.append({
        'waktu': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user': session.get('user') or '',
        'mode': mode,
        'nama_file': nama_file,
        'berhasil': str(berhasil),
        'gagal': str(gagal),
        'catatan': catatan[:2000],
    })
    tulis_csv(
        FILE_IMPORT_LOG,
        rows,
        ['waktu', 'user', 'mode', 'nama_file', 'berhasil', 'gagal', 'catatan'],
    )


def jalankan_import_csv_ringkasan(preview_rows: list, mode: str) -> tuple:
    """Legacy preview import (dinonaktifkan di UI)."""
    return 0, len(preview_rows or []), 'Import ringkasan tidak lagi digunakan.'


def _float_impor_csv(val) -> float:
    if val is None or str(val).strip() == '':
        return 0.0
    return float(str(val).strip().replace(',', '').replace(' ', ''))


def upsert_anggota_dari_baris_impor(anggota_list: list, id_a: str, nama: str, nik: str, alamat: str) -> None:
    for a in anggota_list:
        if a.get('id_anggota') == id_a:
            if nama:
                a['nama'] = nama
            if nik:
                a['nik'] = nik
            if alamat:
                a['alamat'] = alamat
            return
    no_baru = generate_no_anggota_berikutnya(anggota_list)
    anggota_list.append({
        'id_anggota': id_a,
        'no_anggota': no_baru,
        'nik': nik or '',
        'nama': nama or 'Anggota',
        'alamat': alamat or '',
        'no_telp': '',
        'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
        'penghasilan_bersih': '0',
        'cicilan_lain': '0',
    })


IMPORT_CSV_FIELDNAMES = [
    'id_anggota', 'nama', 'nik', 'alamat', 'simpanan', 'pinjaman', 'tenor',
]


@app.route('/import_csv', methods=['POST'])
@app.route('/import_excel', methods=['POST'])
@admin_required
@csrf_protect
def import_csv_unified():
    """Unggah satu file Excel: merge anggota, akumulasi simpanan & pinjaman (tenor max)."""
    redir = request.referrer or url_for('halaman_anggota')
    try:
        if 'file' not in request.files:
            flash('Tidak ada berkas yang diunggah.', 'danger')
            return redirect(redir)
        f = request.files['file']
        ext = os.path.splitext((f.filename or '').lower())[1]
        if ext not in ('.xlsx', '.xlsm'):
            flash('Unggah file Excel dengan ekstensi .xlsx atau .xlsm.', 'danger')
            return redirect(redir)
        raw = f.read()
        if len(raw) > 2 * 1024 * 1024:
            flash('Ukuran file maksimal 2 MB.', 'danger')
            return redirect(redir)
        fieldnames, rows_list = excel_bytes_ke_baris_dict(raw)
        if not fieldnames:
            flash('Excel tidak memiliki header yang valid.', 'danger')
            return redirect(redir)
        hdr_set = set(fieldnames)
        missing = [c for c in IMPORT_CSV_FIELDNAMES if c not in hdr_set]
        if missing:
            flash('Kolom wajib tidak lengkap: ' + ', '.join(missing), 'danger')
            return redirect(redir)
        if not rows_list:
            flash('File tidak berisi baris data (selain header).', 'danger')
            return redirect(redir)
        if len(rows_list) > 500:
            flash('Maksimal 500 baris data per unggah.', 'danger')
            return redirect(redir)

        ensure_anggota_schema()
        ensure_simpanan_schema()
        ensure_pinjaman_plafon_schema()
        anggota = baca_csv(FILE_ANGGOTA)
        simpanan = baca_csv(FILE_SIMPANAN)
        pinjaman = baca_csv(FILE_PINJAMAN)

        diproses = 0
        dilewati = 0
        for row in rows_list:
            id_a = (row.get('id_anggota') or '').strip()
            if not id_a:
                dilewati += 1
                continue
            try:
                v_simp = _float_impor_csv(row.get('simpanan'))
                v_pin = _float_impor_csv(row.get('pinjaman'))
                v_ten = int(float(str(row.get('tenor') or '0').strip()))
            except (TypeError, ValueError):
                dilewati += 1
                continue
            if v_simp < 0 or v_pin < 0 or v_ten < 0 or v_ten > 24:
                dilewati += 1
                continue

            nama = (row.get('nama') or '').strip()
            nik = (row.get('nik') or '').strip()
            alamat = (row.get('alamat') or '').strip()
            upsert_anggota_dari_baris_impor(anggota, id_a, nama, nik, alamat)
            if v_simp > 0:
                merge_akumulasi(simpanan, id_a, v_simp)
            if v_pin > 0 or v_ten > 0:
                merge_pinjaman_akumulasi(pinjaman, id_a, v_pin, v_ten, JENIS_IMPORT_CSV)
            diproses += 1

        simpanan[:] = _dedupe_rows_simpanan(simpanan)
        pinjaman[:] = _dedupe_rows_pinjaman(pinjaman)
        tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
        tulis_csv(FILE_SIMPANAN, simpanan, SIMPANAN_FIELDNAMES)
        amap_fin = {a.get('id_anggota'): a for a in anggota}
        for p in pinjaman:
            ag = amap_fin.get(p.get('id_anggota'))
            if ag:
                p['nama_anggota'] = ag.get('nama', '')
                p['no_anggota'] = ag.get('no_anggota', '')
        tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
        flash(
            f'Import selesai: {diproses} baris diproses, {dilewati} baris dilewati (tidak valid).',
            'success',
        )
    except Exception as ex:
        flash(f'Import gagal: {ex}', 'danger')
    return redirect(redir)


@app.route('/anggota/import-csv', methods=['GET'])
@admin_required
def halaman_import_csv_anggota():
    flash('Fitur import CSV ringkasan dinonaktifkan.', 'warning')
    return redirect(url_for('halaman_anggota'))


@app.route('/anggota/import-csv/sample')
@admin_required
def download_sample_csv_anggota():
    flash('Fitur import CSV ringkasan dinonaktifkan.', 'warning')
    return redirect(url_for('halaman_anggota'))


@app.route('/anggota/import-csv/preview', methods=['POST'])
@admin_required
@csrf_protect
def preview_import_csv_anggota():
    flash('Fitur import CSV ringkasan dinonaktifkan.', 'warning')
    return redirect(url_for('halaman_anggota'))


@app.route('/anggota/import-csv/preview', methods=['GET'])
@admin_required
def tampil_preview_import_csv():
    flash('Fitur import CSV ringkasan dinonaktifkan.', 'warning')
    return redirect(url_for('halaman_anggota'))


@app.route('/anggota/import-csv/execute', methods=['POST'])
@admin_required
@csrf_protect
def execute_import_csv_anggota():
    flash('Fitur import CSV ringkasan dinonaktifkan.', 'warning')
    return redirect(url_for('halaman_anggota'))


@app.route('/anggota/cari_nama', methods=['GET'])
@login_required
def cari_nama():
    """Autocomplete nama anggota berdasarkan No Anggota."""
    no_anggota = request.args.get('no_anggota', '')
    data = baca_csv(FILE_ANGGOTA)
    for a in data:
        if a['no_anggota'] == no_anggota:
            return jsonify({'nama': a['nama'], 'id_anggota': a['id_anggota']})
    return jsonify({'nama': '', 'id_anggota': ''})


@app.route('/anggota/cari_anggota', methods=['GET'])
@login_required
def cari_anggota_autocomplete():
    """Cari anggota berdasarkan nama (autocomplete)."""
    query = request.args.get('q', '').lower()
    data = baca_csv(FILE_ANGGOTA)
    hasil = [a for a in data if query in a['nama'].lower()]
    return jsonify(hasil)


# ══════════════════════════════════════════════
#  ROUTE: SIMPANAN
# ══════════════════════════════════════════════
@app.route('/simpanan')
@login_required
def halaman_simpanan():
    ensure_simpanan_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    anggota = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        simpanan = [s for s in simpanan if s.get('id_anggota') == id_anggota]
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]

    simpanan_tampil = enrich_simpanan_untuk_tampilan(simpanan, anggota)
    simpanan_tampil.reverse()

    saldo_anggota = {}
    for a in anggota:
        saldo_anggota[a['id_anggota']] = 0.0
    for s in simpanan:
        if s['id_anggota'] in saldo_anggota:
            saldo_anggota[s['id_anggota']] += float(s.get('total_simpanan') or 0)

    return render_template(
        'simpanan.html',
        simpanan=simpanan_tampil,
        anggota=anggota,
        saldo_anggota=saldo_anggota,
    )


@app.route('/simpanan/tambah', methods=['POST'])
@login_required
@csrf_protect
def tambah_simpanan():
    ensure_simpanan_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    if is_current_user_admin():
        id_anggota = request.form['id_anggota']
    else:
        id_anggota = get_current_user_id_anggota()
        if not id_anggota:
            abort(403)
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_data = next((a for a in anggota if a['id_anggota'] == id_anggota), None)

    if not anggota_data:
        return "Anggota tidak ditemukan", 400

    try:
        jumlah = float(request.form['jumlah'].replace(',', '').replace('.', ''))
    except ValueError:
        flash('Jumlah simpanan tidak valid.', 'danger')
        return redirect('/simpanan')
    if jumlah < 0:
        flash('Jumlah simpanan tidak boleh negatif.', 'danger')
        return redirect('/simpanan')

    merge_akumulasi(simpanan, id_anggota, jumlah)
    tulis_csv(FILE_SIMPANAN, simpanan, SIMPANAN_FIELDNAMES)
    flash('Simpanan berhasil ditambahkan ke saldo.', 'success')
    return redirect('/simpanan')


@app.route('/simpanan/konfirmasi/<id_simpanan>')
@admin_required
def konfirmasi_simpanan(id_simpanan):
    flash('Skema simpanan saldo tidak memerlukan konfirmasi baris.', 'info')
    return redirect('/simpanan')


@app.route('/simpanan/tolak/<id_simpanan>')
@admin_required
def tolak_simpanan(id_simpanan):
    flash('Skema simpanan saldo tidak memerlukan penolakan baris.', 'info')
    return redirect('/simpanan')


@app.route('/simpanan/hapus/<id_anggota>', methods=['POST'])
@admin_required
@csrf_protect
def hapus_simpanan(id_anggota):
    data = [s for s in baca_csv(FILE_SIMPANAN) if s.get('id_anggota') != id_anggota]
    tulis_csv(FILE_SIMPANAN, data, SIMPANAN_FIELDNAMES)
    flash('Saldo simpanan anggota dihapus dari daftar.', 'success')
    return redirect('/simpanan')


@app.route('/simpanan/saldo/<id_anggota>')
@login_required
def cek_saldo(id_anggota):
    """Cek total saldo simpanan anggota."""
    restrict_id_anggota_or_forbid(id_anggota)
    ensure_simpanan_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    total = 0.0
    for s in simpanan:
        if s.get('id_anggota') == id_anggota:
            total = float(s.get('total_simpanan') or 0)
            break
    return jsonify({'saldo': total})


# ══════════════════════════════════════════════
#  ROUTE: PINJAMAN
# ══════════════════════════════════════════════
@app.route('/pinjaman')
@login_required
def halaman_pinjaman():
    ensure_pinjaman_plafon_schema()
    pinjaman = baca_csv(FILE_PINJAMAN)
    anggota_full = baca_csv(FILE_ANGGOTA)
    cicilan_menunggu = []
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        pinjaman = [p for p in pinjaman if p.get('id_anggota') == id_anggota]
        anggota = [a for a in anggota_full if a.get('id_anggota') == id_anggota]
    else:
        anggota = anggota_full
        cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
        cicilan_menunggu = [c for c in cicilan if c.get('status') == 'Menunggu']
    pinjaman_tampil = enrich_pinjaman_untuk_tampilan(pinjaman, anggota_full)
    pinjaman_tampil.reverse()
    return render_template(
        'pinjaman.html',
        pinjaman=pinjaman_tampil,
        anggota=anggota,
        cicilan_menunggu=cicilan_menunggu,
        metode_bayar_choices=METODE_BAYAR_CHOICES,
        jenis_pinjaman_choices=list(JENIS_PINJAMAN.keys()),
    )


@app.route('/pinjaman/hitung', methods=['GET'])
def hitung_pinjaman():
    """Hitung estimasi pinjaman berdasarkan plafon + tenor fleksibel."""
    try:
        raw_plafon = request.args.get('plafon') or request.args.get('Pinjaman', 0)
        raw_tenor = request.args.get('tenor')
        id_anggota = request.args.get('id_anggota')
        jenis_pm = (request.args.get('jenis_pinjaman') or '').strip()
        plafon = float(raw_plafon)
        if not raw_tenor:
            return jsonify({'error': 'Tenor wajib diisi'})
        tenor = int(raw_tenor)
        if tenor <= 0:
            return jsonify({'error': 'Tenor harus lebih dari 0'}), 400
        if tenor > 24:
            return jsonify({'error': 'Tenor maksimal 24 bulan'}), 400
        bunga_persen = bunga_untuk_jenis_pinjaman(jenis_pm, tenor)
        bunga = bunga_persen / 100
        total_bayar = plafon + (plafon * bunga * tenor)
        cicilan = total_bayar / tenor

        kapasitas_cicilan = None
        plafon_maks = None
        dsr = None
        dsr_persen = None
        if id_anggota:
            ensure_anggota_schema()
            anggota = get_anggota_by_id(id_anggota)
            if anggota:
                penghasilan = float(anggota.get('penghasilan_bersih') or 0)
                cicilan_lain = float(anggota.get('cicilan_lain') or 0)
                dsr = dsr_otomatis_dari_penghasilan(penghasilan)
                dsr_persen = round(dsr * 100)
                kapasitas_cicilan = hitung_kapasitas_cicilan(penghasilan, cicilan_lain, dsr)
                plafon_maks = hitung_plafon_maks_anuitas(kapasitas_cicilan, bunga_persen, tenor)
        return jsonify({
            'bunga_persen': bunga_persen,
            'tenor': tenor,
            'total_bayar': round(total_bayar, 2),
            'cicilan_per_bulan': round(cicilan, 2),
            'kapasitas_cicilan': round(kapasitas_cicilan, 2) if kapasitas_cicilan is not None else None,
            'plafon_maks': round(plafon_maks, 2) if plafon_maks is not None else None,
            'dsr_persen': dsr_persen,
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/pinjaman/cek_tunggakan/<id_anggota>')
@login_required
def cek_tunggakan(id_anggota):
    """Cek apakah anggota masih punya tunggakan."""
    restrict_id_anggota_or_forbid(id_anggota)
    pinjaman = baca_csv(FILE_PINJAMAN)
    tunggakan = [
        p for p in pinjaman
        if p.get('id_anggota') == id_anggota
        and p.get('status') == 'Disetujui'
        and float(p.get('sisa_pinjaman') or p.get('total_pinjaman') or 0) > 0
    ]
    if tunggakan:
        return jsonify({'ada_tunggakan': True, 'pinjaman_aktif': len(tunggakan)})
    return jsonify({'ada_tunggakan': False, 'pinjaman_aktif': 0})


@app.route('/pinjaman/tambah', methods=['POST'])
@login_required
@csrf_protect
def tambah_pinjaman():
    pinjaman = baca_csv(FILE_PINJAMAN)
    if is_current_user_admin():
        id_anggota = request.form['id_anggota']
    else:
        id_anggota = get_current_user_id_anggota()
        if not id_anggota:
            abort(403)
    ensure_anggota_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_data = next((a for a in anggota if a['id_anggota'] == id_anggota), None)

    if not anggota_data:
        return "Anggota tidak ditemukan", 400

    jenis = (request.form.get('jenis_pinjaman') or 'Jangka Menengah').strip()
    plafon = float(request.form['plafon'].replace(',', '').replace('.', ''))
    tenor = int(request.form['tenor_bulan'])
    if tenor < 1 or tenor > 24:
        flash('Tenor pinjaman harus antara 1 sampai 24 bulan.', 'danger')
        return redirect('/pinjaman')

    try:
        bunga_persen = bunga_untuk_jenis_pinjaman(jenis, tenor)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect('/pinjaman')
    bunga = bunga_persen / 100
    total_bayar = plafon + (plafon * bunga * tenor)
    cicilan = total_bayar / tenor

    penghasilan = float(anggota_data.get('penghasilan_bersih') or 0)
    cicilan_lain = float(anggota_data.get('cicilan_lain') or 0)
    dsr = dsr_otomatis_dari_penghasilan(penghasilan)
    kapasitas_cicilan = hitung_kapasitas_cicilan(penghasilan, cicilan_lain, dsr)
    plafon_maks = hitung_plafon_maks_anuitas(kapasitas_cicilan, bunga_persen, tenor)
    if kapasitas_cicilan <= 0:
        flash('Pengajuan ditolak: kapasitas cicilan anggota tidak mencukupi.', 'danger')
        return redirect('/pinjaman')
    if cicilan > kapasitas_cicilan:
        flash(
            f"Pengajuan ditolak: cicilan bulanan Rp {cicilan:,.0f} melebihi kapasitas Rp {kapasitas_cicilan:,.0f}.",
            'danger'
        )
        return redirect('/pinjaman')
    if plafon > plafon_maks:
        flash(
            f"Pengajuan melebihi plafon maksimal anggota. Maksimal Rp {plafon_maks:,.0f} "
            f"(kapasitas cicilan Rp {kapasitas_cicilan:,.0f}/bln).",
            'danger'
        )
        return redirect('/pinjaman')

    tgl = datetime.now().strftime('%Y-%m-%d')
    pid = str(uuid.uuid4())
    pinjaman.append({
        'id_pinjaman': pid,
        'id_anggota': id_anggota,
        'nama_anggota': anggota_data.get('nama', ''),
        'no_anggota': anggota_data.get('no_anggota', ''),
        'jenis_pinjaman': jenis,
        'plafon': str(round(plafon, 2)),
        'tenor_bulan': str(tenor),
        'bunga_persen': str(bunga_persen),
        'total_bayar': str(round(total_bayar, 2)),
        'cicilan_per_bulan': str(round(cicilan, 2)),
        'sisa_pinjaman': '0',
        'tanggal_pengajuan': tgl,
        'status': 'Menunggu',
        'tanggal_lunas': '',
    })
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    flash('Pengajuan pinjaman tercatat dan menunggu konfirmasi admin.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/konfirmasi/<id_pinjaman>', methods=['POST'])
@admin_required
@csrf_protect
def konfirmasi_pinjaman(id_pinjaman):
    pinjaman = baca_csv(FILE_PINJAMAN)
    found = False
    for p in pinjaman:
        if p.get('id_pinjaman') == id_pinjaman and p.get('status') == 'Menunggu':
            p['status'] = 'Disetujui'
            tb = float(p.get('total_bayar') or 0)
            p['sisa_pinjaman'] = str(round(tb, 2))
            found = True
            break
    if not found:
        flash('Pengajuan tidak ditemukan atau sudah diproses.', 'warning')
        return redirect('/pinjaman')
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    flash('Pinjaman disetujui.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/tolak/<id_pinjaman>', methods=['POST'])
@admin_required
@csrf_protect
def tolak_pinjaman(id_pinjaman):
    pinjaman = baca_csv(FILE_PINJAMAN)
    found = False
    for p in pinjaman:
        if p.get('id_pinjaman') == id_pinjaman and p.get('status') == 'Menunggu':
            p['status'] = 'Ditolak'
            found = True
            break
    if not found:
        flash('Pengajuan tidak ditemukan atau sudah diproses.', 'warning')
        return redirect('/pinjaman')
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    flash('Pengajuan pinjaman ditolak.', 'warning')
    return redirect('/pinjaman')


def _proses_angsur_pinjaman(id_pinjaman: str):
    """Kurangi sisa kewajiban satu kali cicilan."""
    pinjaman = baca_csv(FILE_PINJAMAN)
    for p in pinjaman:
        if p.get('id_pinjaman') != id_pinjaman:
            continue
        if p.get('status') != 'Disetujui':
            return
        sisa = float(p.get('sisa_pinjaman') or 0)
        cic = float(p.get('cicilan_per_bulan') or 0)
        if cic <= 0 or sisa <= 0:
            return
        baru = max(sisa - cic, 0)
        p['sisa_pinjaman'] = str(round(baru, 2))
        if baru <= 0:
            p['status'] = 'Lunas'
            p['tanggal_lunas'] = datetime.now().strftime('%Y-%m-%d')
        break
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)


def _pinjaman_row_dengan_nama(id_pinjaman: str):
    pinjaman = baca_csv(FILE_PINJAMAN)
    target = next((p for p in pinjaman if p.get('id_pinjaman') == id_pinjaman), None)
    if not target:
        return None, None
    ag = get_anggota_by_id(target.get('id_anggota', '')) or {}
    cic = float(target.get('cicilan_per_bulan') or 0) or cicilan_per_bulan_saldo(target)
    return target, {
        'no_anggota': target.get('no_anggota') or ag.get('no_anggota', ''),
        'nama_anggota': target.get('nama_anggota') or ag.get('nama', ''),
        'cicilan_per_bulan': str(round(cic, 2)),
    }


@app.route('/pinjaman/angsur/<id_pinjaman>', methods=['POST'])
@admin_required
@csrf_protect
def angsur_pinjaman(id_pinjaman):
    """Bayar cicilan langsung oleh admin; mencatat metode Admin di riwayat cicilan."""
    ensure_pinjaman_cicilan_schema()
    ket = (request.form.get('keterangan') or '').strip() or 'Pembayaran cicilan melalui admin (kas koperasi)'
    target, meta = _pinjaman_row_dengan_nama(id_pinjaman)
    if not target:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    if target.get('status') != 'Disetujui' or float(target.get('sisa_pinjaman') or 0) <= 0:
        flash('Pinjaman ini tidak memiliki cicilan yang dapat dibayar.', 'warning')
        return redirect('/pinjaman')

    id_anggota = target.get('id_anggota', '')
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    cicilan.append({
        'id_cicilan': str(uuid.uuid4()),
        'id_pinjaman': id_pinjaman,
        'id_anggota': id_anggota,
        'no_anggota': meta['no_anggota'],
        'nama_anggota': meta['nama_anggota'],
        'jumlah': meta['cicilan_per_bulan'],
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Disetujui',
        'tanggal_konfirmasi': datetime.now().strftime('%Y-%m-%d'),
        'dikonfirmasi_oleh': session.get('user') or '',
        'diajukan_oleh': session.get('user') or '',
        'keterangan': ket,
        'metode_pembayaran': 'Admin',
        'detail_pembayaran': '',
    })
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    _proses_angsur_pinjaman(id_pinjaman)
    flash('Cicilan dicatat (bayar melalui admin). Saldo pinjaman diperbarui.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/ajukan-cicilan/<id_pinjaman>', methods=['POST'])
@login_required
@csrf_protect
def ajukan_cicilan(id_pinjaman):
    """Pengajuan bayar cicilan oleh user; perlu dikonfirmasi admin sebelum mengurangi sisa pinjaman."""
    ensure_pinjaman_cicilan_schema()
    target, meta = _pinjaman_row_dengan_nama(id_pinjaman)
    if not target:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')

    id_anggota = target.get('id_anggota', '')
    if not is_current_user_admin():
        id_anggota_user = get_current_user_id_anggota()
        if not id_anggota_user or id_anggota != id_anggota_user:
            abort(403)

    if target.get('status') != 'Disetujui' or float(target.get('sisa_pinjaman') or 0) <= 0:
        flash('Pinjaman ini tidak memiliki cicilan yang perlu dibayar.', 'warning')
        return redirect('/pinjaman')

    metode = (request.form.get('metode_pembayaran') or 'Lainnya').strip()
    if metode not in METODE_BAYAR_CHOICES:
        metode = 'Lainnya'
    detail = (request.form.get('detail_pembayaran') or '').strip()
    ket_parts = [f'Pengajuan pembayaran via {metode}']
    if detail:
        ket_parts.append(detail)
    keterangan = ' — '.join(ket_parts)

    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    cicilan.append({
        'id_cicilan': str(uuid.uuid4()),
        'id_pinjaman': id_pinjaman,
        'id_anggota': id_anggota,
        'no_anggota': meta['no_anggota'],
        'nama_anggota': meta['nama_anggota'],
        'jumlah': meta['cicilan_per_bulan'],
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Menunggu',
        'tanggal_konfirmasi': '',
        'dikonfirmasi_oleh': '',
        'diajukan_oleh': session.get('user') or '',
        'keterangan': keterangan,
        'metode_pembayaran': metode,
        'detail_pembayaran': detail,
    })
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    flash('Pengajuan bayar cicilan berhasil dikirim. Menunggu konfirmasi admin.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/cicilan/konfirmasi/<id_cicilan>', methods=['POST'])
@admin_required
@csrf_protect
def konfirmasi_cicilan(id_cicilan):
    """Admin menyetujui pengajuan cicilan dan langsung mengurangi sisa pinjaman."""
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    target = None
    for c in cicilan:
        if c.get('id_cicilan') == id_cicilan:
            target = c
            break
    if not target:
        flash('Data pengajuan cicilan tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    if target.get('status') != 'Menunggu':
        flash('Pengajuan cicilan ini sudah diproses sebelumnya.', 'warning')
        return redirect('/pinjaman')

    target['status'] = 'Disetujui'
    target['tanggal_konfirmasi'] = datetime.now().strftime('%Y-%m-%d')
    target['dikonfirmasi_oleh'] = session.get('user') or ''
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)

    _proses_angsur_pinjaman(target['id_pinjaman'])
    flash('Pengajuan bayar cicilan disetujui dan sisa pinjaman telah diperbarui.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/cicilan/tolak/<id_cicilan>', methods=['POST'])
@admin_required
@csrf_protect
def tolak_cicilan(id_cicilan):
    """Admin menolak pengajuan cicilan (tanpa mengubah sisa pinjaman)."""
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    found = False
    for c in cicilan:
        if c.get('id_cicilan') == id_cicilan:
            c['status'] = 'Ditolak'
            c['tanggal_konfirmasi'] = datetime.now().strftime('%Y-%m-%d')
            c['dikonfirmasi_oleh'] = session.get('user') or ''
            found = True
            break
    if not found:
        flash('Data pengajuan cicilan tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    flash('Pengajuan bayar cicilan ditolak.', 'warning')
    return redirect('/pinjaman')


@app.route('/pinjaman/hapus/<id_pinjaman>', methods=['POST'])
@admin_required
@csrf_protect
def hapus_pinjaman(id_pinjaman):
    data = [p for p in baca_csv(FILE_PINJAMAN) if p.get('id_pinjaman') != id_pinjaman]
    tulis_csv(FILE_PINJAMAN, data, PINJAMAN_FIELDNAMES)
    flash('Data pinjaman dihapus.', 'success')
    return redirect('/pinjaman')


# ══════════════════════════════════════════════
#  ROUTE: LAPORAN & EXPORT
# ══════════════════════════════════════════════

def upsert_anggota_from_riwayat(pinjaman_rows: list, simpanan_rows: list) -> dict:
    """Sinkronkan anggota dari baris saldo pinjaman/simpanan (berbasis id_anggota)."""
    ensure_anggota_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    by_id = {a.get('id_anggota'): a for a in anggota}

    added = 0

    def ensure_id(id_a: str):
        nonlocal added
        id_a = (id_a or '').strip()
        if not id_a:
            return
        if id_a in by_id:
            return
        no_baru = generate_no_anggota_berikutnya(anggota)
        anggota.append({
            'id_anggota': id_a,
            'no_anggota': no_baru,
            'nik': '',
            'nama': '',
            'alamat': '',
            'no_telp': '',
            'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
            'penghasilan_bersih': '0',
            'cicilan_lain': '0',
        })
        by_id[id_a] = anggota[-1]
        added += 1

    for p in pinjaman_rows:
        ensure_id(p.get('id_anggota'))
    for s in simpanan_rows:
        ensure_id(s.get('id_anggota'))

    tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
    return {'added': added, 'updated': 0}


def hitung_shu_pinjaman() -> dict:
    """SHU perkiraan: proporsi dari nominal cicilan yang sudah disetujui (saldo pokok)."""
    ensure_pinjaman_cicilan_schema()
    cicilan_rows = [c for c in baca_csv(FILE_PINJAMAN_CICILAN) if c.get('status') == 'Disetujui']
    anggota_map = {a.get('id_anggota'): a for a in baca_csv(FILE_ANGGOTA)}

    bunga_per_anggota = {}
    total_bunga_dibayar = 0.0
    for c in cicilan_rows:
        id_anggota = (c.get('id_anggota') or '').strip()
        j = float(c.get('jumlah') or 0)
        if j <= 0:
            continue
        total_bunga_dibayar += j
        ag = anggota_map.get(id_anggota, {})
        if id_anggota not in bunga_per_anggota:
            bunga_per_anggota[id_anggota] = {
                'id_anggota': id_anggota,
                'no_anggota': ag.get('no_anggota', ''),
                'nama_anggota': ag.get('nama', ''),
                'bunga_dibayar': 0.0,
            }
        bunga_per_anggota[id_anggota]['bunga_dibayar'] += j

    shu_kotor = total_bunga_dibayar
    shu_anggota_65 = shu_kotor * 0.65

    alokasi = []
    for row in bunga_per_anggota.values():
        prop = (row['bunga_dibayar'] / shu_kotor) if shu_kotor > 0 else 0.0
        shu_member = shu_anggota_65 * prop
        alokasi.append({
            'no_anggota': row['no_anggota'],
            'nama_anggota': row['nama_anggota'],
            'bunga_dibayar': row['bunga_dibayar'],
            'shu_anggota': shu_member,
        })

    alokasi.sort(key=lambda x: x['shu_anggota'], reverse=True)
    return {
        'shu_kotor': shu_kotor,
        'shu_anggota_65': shu_anggota_65,
        'shu_lain': shu_kotor - shu_anggota_65,
        'alokasi': alokasi
    }


def generate_excel_laporan_terpadu():
    """Buat workbook Excel berisi 2 tabel: pinjaman + simpanan, plus tabel SHU pinjaman."""
    ensure_pinjaman_plafon_schema()
    ensure_simpanan_schema()
    ensure_pinjaman_cicilan_schema()

    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            raise RuntimeError("Fitur export membutuhkan openpyxl.")

    anggota_full = baca_csv(FILE_ANGGOTA)
    pinjaman_t = enrich_pinjaman_untuk_tampilan(baca_csv(FILE_PINJAMAN), anggota_full)
    pinjaman_t.reverse()

    wb = wb_cls()
    ws = wb.active
    ws.title = 'Laporan Terpadu'

    # -------------------
    # Tabel: PINJAMAN
    # -------------------
    ws.append(['DATA PINJAMAN'])
    ws.append(['No', 'Jenis', 'No Anggota', 'Nama', 'Plafon', 'Tenor', 'Bunga %', 'Cicilan/bln', 'Sisa', 'Status'])
    for i, p in enumerate(pinjaman_t, 1):
        ws.append([
            i,
            p.get('jenis_pinjaman', ''),
            p.get('no_anggota', ''),
            p.get('nama_anggota', ''),
            float(p.get('plafon') or p.get('total_pinjaman') or 0),
            int(float(p.get('tenor_bulan') or p.get('tenor') or 0)),
            float(p.get('bunga_persen') or 0),
            float(p.get('cicilan_per_bulan') or 0),
            float(p.get('sisa_pinjaman') or 0),
            p.get('status', ''),
        ])

    # -------------------
    # SHU (65% untuk anggota)
    # -------------------
    shu = hitung_shu_pinjaman()
    ws.append([])
    ws.append(['SHU PINJAMAN'])
    ws.append(['SHU Kotor (bunga dibayar)', float(shu['shu_kotor'])])
    ws.append(['SHU Anggota (65%)', float(shu['shu_anggota_65'])])
    ws.append(['SHU Lainnya (35%)', float(shu['shu_lain'])])
    ws.append([])

    ws.append(['Alokasi SHU Anggota 65% (proporsional bunga dibayar)'])
    ws.append(['No', 'No Anggota', 'Nama', 'Bunga dibayar', 'SHU anggota'])
    for i, a in enumerate(shu['alokasi'], 1):
        ws.append([i, a['no_anggota'], a['nama_anggota'], float(a['bunga_dibayar']), float(a['shu_anggota'])])

    # -------------------
    # Tabel: SIMPANAN
    # -------------------
    simpanan_t = enrich_simpanan_untuk_tampilan(baca_csv(FILE_SIMPANAN), anggota_full)
    simpanan_t.reverse()
    ws.append([])
    ws.append(['DATA SIMPANAN (saldo per anggota)'])
    ws.append(['No', 'No Anggota', 'Nama', 'Total Simpanan'])
    for i, s in enumerate(simpanan_t, 1):
        ws.append([
            i,
            s.get('no_anggota', ''),
            s.get('nama_anggota', ''),
            float(s.get('total_simpanan') or 0),
        ])

    return wb


@app.route('/laporan')
@login_required
def halaman_laporan():
    ensure_simpanan_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    pinjaman = baca_csv(FILE_PINJAMAN)
    anggota = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        simpanan = [s for s in simpanan if s.get('id_anggota') == id_anggota]
        pinjaman = [p for p in pinjaman if p.get('id_anggota') == id_anggota]
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]
    return render_template('laporan.html', simpanan=simpanan, pinjaman=pinjaman, anggota=anggota)


@app.route('/export/laporan-terpadu')
@admin_required
def export_laporan_terpadu():
    """Upsert anggota dari data pinjaman/simpanan lalu siapkan download Excel 1 file."""
    ensure_simpanan_schema()
    ensure_pinjaman_plafon_schema()

    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    simpanan_rows = baca_csv(FILE_SIMPANAN)
    hasil = upsert_anggota_from_riwayat(pinjaman_rows, simpanan_rows)

    return render_template(
        'export_laporan_terpadu.html',
        added=hasil['added'],
        updated=hasil['updated'],
        download_url=url_for('export_laporan_terpadu_unduh'),
    )


@app.route('/export/laporan-terpadu/unduh')
@admin_required
def export_laporan_terpadu_unduh():
    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            return f"Fitur export membutuhkan openpyxl. Jalankan: \"{sys.executable}\" -m pip install openpyxl", 500

    wb = generate_excel_laporan_terpadu()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='laporan_terpadu_pinjam_simpan.xlsx',
        as_attachment=True
    )


@app.route('/export/simpanan')
@admin_required
def export_simpanan():
    ensure_simpanan_schema()
    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            return f"Fitur export membutuhkan openpyxl. Jalankan: \"{sys.executable}\" -m pip install openpyxl", 500
    wb = wb_cls()
    ws = wb.active
    ws.title = "Data Simpanan"
    ws.append(['No', 'No Anggota', 'Nama', 'Total Simpanan'])
    simpanan = enrich_simpanan_untuk_tampilan(baca_csv(FILE_SIMPANAN), baca_csv(FILE_ANGGOTA))
    for i, s in enumerate(simpanan, 1):
        ws.append([i, s['no_anggota'], s['nama_anggota'], float(s.get('total_simpanan') or 0)])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='laporan_simpanan.xlsx', as_attachment=True)


@app.route('/export/pinjaman')
@admin_required
def export_pinjaman():
    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            return f"Fitur export membutuhkan openpyxl. Jalankan: \"{sys.executable}\" -m pip install openpyxl", 500
    wb = wb_cls()
    ws = wb.active
    ws.title = "Data Pinjaman"
    ws.append(['No', 'Jenis', 'No Anggota', 'Nama', 'Plafon', 'Tenor', 'Cicilan/Bulan', 'Sisa', 'Status'])
    pinjaman = enrich_pinjaman_untuk_tampilan(baca_csv(FILE_PINJAMAN), baca_csv(FILE_ANGGOTA))
    for i, p in enumerate(pinjaman, 1):
        ws.append([
            i,
            p.get('jenis_pinjaman', ''),
            p['no_anggota'],
            p['nama_anggota'],
            float(p.get('plafon') or 0),
            int(float(p.get('tenor_bulan') or 0)),
            float(p.get('cicilan_per_bulan') or 0),
            float(p.get('sisa_pinjaman') or 0),
            p.get('status', ''),
        ])

    # Tambahkan SHU Pinjaman
    shu = hitung_shu_pinjaman()
    ws.append([])
    ws.append(['SHU PINJAMAN'])
    ws.append(['SHU Kotor (bunga dibayar)', float(shu['shu_kotor'])])
    ws.append(['SHU Anggota (65%)', float(shu['shu_anggota_65'])])
    ws.append(['SHU Lainnya (35%)', float(shu['shu_lain'])])
    ws.append([])
    ws.append(['Alokasi SHU Anggota 65% (proporsional bunga dibayar)'])
    ws.append(['No', 'No Anggota', 'Nama', 'Bunga dibayar', 'SHU anggota'])
    for i, a in enumerate(shu['alokasi'], 1):
        ws.append([i, a['no_anggota'], a['nama_anggota'], float(a['bunga_dibayar']), float(a['shu_anggota'])])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='laporan_pinjaman.xlsx', as_attachment=True)


@app.route('/laporan/anggota/<id_anggota>')
@login_required
def laporan_anggota(id_anggota):
    """Detail laporan per anggota."""
    restrict_id_anggota_or_forbid(id_anggota)
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_data = next((a for a in anggota if a['id_anggota'] == id_anggota), None)
    if not anggota_data:
        return "Anggota tidak ditemukan", 404

    ensure_simpanan_schema()
    anggota_all = baca_csv(FILE_ANGGOTA)
    simpanan_raw = [s for s in baca_csv(FILE_SIMPANAN) if s['id_anggota'] == id_anggota]
    pinjaman_raw = [p for p in baca_csv(FILE_PINJAMAN) if p['id_anggota'] == id_anggota]
    simpanan = enrich_simpanan_untuk_tampilan(simpanan_raw, anggota_all)
    pinjaman = enrich_pinjaman_untuk_tampilan(pinjaman_raw, anggota_all)
    total_simpanan = sum(float(s.get('total_simpanan') or 0) for s in simpanan_raw)
    total_pinjaman = sum(
        float(p.get('sisa_pinjaman') or 0)
        for p in pinjaman_raw
        if p.get('status') == 'Disetujui'
    )

    return render_template('cetak_struk.html',
                           anggota=anggota_data,
                           simpanan=simpanan,
                           pinjaman=pinjaman,
                           total_simpanan=total_simpanan,
                           total_pinjaman=total_pinjaman)


@app.route('/favicon.ico')
def favicon():
    return '', 204


# ══════════════════════════════════════════════
#  JALANKAN APLIKASI
# ══════════════════════════════════════════════
if __name__ == '__main__':
    print("=" * 50)
    print("  APLIKASI KOPERASI BERBASIS WEB")
    print("  Teknologi: Python Flask + CSV Database")
    print("  Beranda: http://localhost:5000/")
    print("  Login : http://localhost:5000/login")
    print("=" * 50)
    app.run(debug=True, host='0.0.0.0', port=5000)