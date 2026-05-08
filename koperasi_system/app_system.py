import os
import csv
import calendar
import re
import json
import shutil
import uuid
import smtplib
import hmac
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, session, flash, abort
import io
import sys
import zipfile
from email.mime.text import MIMEText
from secrets import token_hex
try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:
    Workbook = None
    load_workbook = None
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import bindparam, text
from sqlalchemy.exc import SQLAlchemyError
from koperasi_system.db import db_session, ping_database
from koperasi_system.settings import (
    BASE_DIR,
    DATABASE_URL,
    FILE_ANGGOTA,
    FILE_SIMPANAN,
    FILE_SIMPANAN_TRANSAKSI,
    FILE_SIMPANAN_PENGAJUAN,
    FILE_IURAN_SOSIAL,
    FILE_PINJAMAN,
    FILE_PINJAMAN_CICILAN,
    FILE_USERS,
    FILE_PENDAFTARAN_ANGGOTA,
    FILE_IMPORT_LOG,
    FILE_BERITA,
    FILE_SHU_TAHUNAN,
    FILE_SHU_ALOKASI,
    BACKUP_DIR,
    DSR_DEFAULT,
    PROVISI_RATE_LONG_TENOR,
    PROVISI_MIN_TENOR_BULAN,
    JENIS_SIMPANAN_IMPORT,
    DEFAULT_TENOR_IMPORT_PINJAMAN,
    METODE_BAYAR_CHOICES,
    KOPERASI_REKENING_BANK,
    ADMIN_NOTIFICATION_EMAIL,
    CICILAN_FIELDNAMES,
    SIMPANAN_FIELDNAMES,
    SIMPANAN_TRANSAKSI_FIELDNAMES,
    SIMPANAN_PENGAJUAN_FIELDNAMES,
    IURAN_SOSIAL_FIELDNAMES,
    SHU_TAHUNAN_FIELDNAMES,
    SHU_ALOKASI_FIELDNAMES,
    PINJAMAN_FIELDNAMES,
    JENIS_IMPORT_CSV,
    IMPORT_PREVIEW_DIR,
    ANGGOTA_FIELDNAMES,
    PENDAFTARAN_FIELDNAMES,
    JENIS_PINJAMAN,
    JENIS_PINJAMAN_CHOICES,
    JENIS_SIMPANAN,
)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'templates'),
    static_folder=os.path.join(BASE_DIR, 'static'),
)
app.secret_key = os.getenv('KOPERASI_SECRET_KEY') or token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.getenv('SESSION_COOKIE_SECURE', '0') == '1',
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=int(os.getenv('USER_SESSION_TIMEOUT_MINUTES', '120'))),
)

LOGIN_MAX_ATTEMPTS = int(os.getenv('LOGIN_MAX_ATTEMPTS', '5'))
LOGIN_LOCK_MINUTES = int(os.getenv('LOGIN_LOCK_MINUTES', '15'))
ADMIN_SESSION_TIMEOUT_MINUTES = int(os.getenv('ADMIN_SESSION_TIMEOUT_MINUTES', '30'))
USER_SESSION_TIMEOUT_MINUTES = int(os.getenv('USER_SESSION_TIMEOUT_MINUTES', '120'))
MAX_BUKTI_TRANSFER_BYTES = int(os.getenv('MAX_BUKTI_TRANSFER_BYTES', str(2 * 1024 * 1024)))

_FAILED_LOGIN_STATE = {}
_LAST_AUTO_BACKUP_DATE = None


def _safe_next_url(raw_next: str, default_url: str) -> str:
    candidate = (raw_next or '').strip()
    if candidate.startswith('/') and not candidate.startswith('//'):
        return candidate
    return default_url


def _role_home_url(role: str) -> str:
    if (role or '').strip().lower() in STAFF_ROLES:
        return url_for('admin_portal.dashboard')
    return url_for('public_portal.dashboard')


def _get_or_create_csrf_token() -> str:
    token = (session.get('_csrf_token') or '').strip()
    if not token:
        token = token_hex(32)
        session['_csrf_token'] = token
    return token


def csrf_protect(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        form_token = (request.form.get('csrf_token') or '').strip()
        session_token = (session.get('_csrf_token') or '').strip()
        if not form_token or not session_token or not hmac.compare_digest(form_token, session_token):
            abort(400)
        return view_func(*args, **kwargs)

    return wrapper


def csrf_protect_if_post(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if request.method == 'POST':
            form_token = (request.form.get('csrf_token') or '').strip()
            session_token = (session.get('_csrf_token') or '').strip()
            if not form_token or not session_token or not hmac.compare_digest(form_token, session_token):
                abort(400)
        return view_func(*args, **kwargs)

    return wrapper


def _login_state_key(username: str) -> str:
    ip_addr = (request.remote_addr or '-').strip()
    return f"{username.lower()}|{ip_addr}"


def _remaining_login_lock_seconds(username: str) -> int:
    state = _FAILED_LOGIN_STATE.get(_login_state_key(username))
    if not state:
        return 0
    locked_until = state.get('locked_until')
    if not locked_until:
        return 0
    now = datetime.utcnow()
    if now >= locked_until:
        _FAILED_LOGIN_STATE.pop(_login_state_key(username), None)
        return 0
    return int((locked_until - now).total_seconds())


def _register_failed_login(username: str):
    key = _login_state_key(username)
    state = _FAILED_LOGIN_STATE.get(key) or {'count': 0, 'locked_until': None}
    state['count'] = int(state.get('count') or 0) + 1
    if state['count'] >= LOGIN_MAX_ATTEMPTS:
        state['locked_until'] = datetime.utcnow() + timedelta(minutes=LOGIN_LOCK_MINUTES)
        state['count'] = 0
    _FAILED_LOGIN_STATE[key] = state


def _clear_failed_login(username: str):
    _FAILED_LOGIN_STATE.pop(_login_state_key(username), None)


def _is_supported_image_signature(ext: str, header: bytes) -> bool:
    ext = (ext or '').lower()
    if ext == '.png':
        return header.startswith(b'\x89PNG\r\n\x1a\n')
    if ext in ('.jpg', '.jpeg'):
        return header.startswith(b'\xff\xd8\xff')
    if ext == '.webp':
        return len(header) >= 12 and header[:4] == b'RIFF' and header[8:12] == b'WEBP'
    return False


def _read_berita_items() -> list[dict]:
    if not os.path.exists(FILE_BERITA):
        return []
    try:
        with open(FILE_BERITA, 'r', encoding='utf-8') as fh:
            raw_items = json.load(fh)
    except Exception:
        return []

    if not isinstance(raw_items, list):
        return []

    items = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        items.append({
            'id': str(item.get('id') or str(uuid.uuid4())),
            'judul': (item.get('judul') or '').strip(),
            'kategori': (item.get('kategori') or 'Pengumuman').strip() or 'Pengumuman',
            'isi': (item.get('isi') or '').strip(),
            'tanggal': (item.get('tanggal') or datetime.now().strftime('%Y-%m-%d')).strip(),
            'status': (item.get('status') or 'Aktif').strip() or 'Aktif',
            'foto': (item.get('foto') or '').strip(),
        })
    items.sort(key=lambda x: x.get('tanggal', ''), reverse=True)
    return items


# --- Backup management UI & actions (super_admin only via permission 'backup.manage')
@app.route('/admin/backups')
def admin_backups():
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    if not has_permission('backup.manage'):
        abort(403)
    entries = []
    try:
        if os.path.exists(BACKUP_DIR):
            for name in sorted(os.listdir(BACKUP_DIR), reverse=True):
                path = os.path.join(BACKUP_DIR, name)
                if not os.path.isdir(path):
                    continue
                manifest = {}
                mpath = os.path.join(path, 'manifest.json')
                if os.path.exists(mpath):
                    try:
                        with open(mpath, 'r', encoding='utf-8') as fh:
                            manifest = json.load(fh) or {}
                    except Exception:
                        manifest = {}
                entries.append({'stamp': name, 'manifest': manifest})
    except Exception:
        entries = []

    log = _load_backup_log()
    return render_template('backups.html', backups=entries, backup_log=log, csrf_token=_get_or_create_csrf_token())


@app.route('/admin/backups/download/<stamp>')
def admin_backups_download(stamp: str):
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    if not has_permission('backup.manage'):
        abort(403)
    target = os.path.join(BACKUP_DIR, stamp)
    if not os.path.isdir(target):
        abort(404)
    # Create zip in memory
    bio = io.BytesIO()
    try:
        with zipfile.ZipFile(bio, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, _, files in os.walk(target):
                for fn in files:
                    full = os.path.join(root, fn)
                    arcname = os.path.relpath(full, target)
                    zf.write(full, arcname)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=f'backup-{stamp}.zip')
    except Exception:
        abort(500)


@app.route('/admin/backups/apply/<stamp>', methods=['POST'])
@csrf_protect
def admin_backups_apply(stamp: str):
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    if not has_permission('backup.manage'):
        abort(403)
    target = os.path.join(BACKUP_DIR, stamp)
    if not os.path.isdir(target):
        abort(404)

    # Map backup basenames to current file paths
    mapping = {
        os.path.basename(FILE_ANGGOTA): FILE_ANGGOTA,
        os.path.basename(FILE_SIMPANAN): FILE_SIMPANAN,
        os.path.basename(FILE_SIMPANAN_TRANSAKSI): FILE_SIMPANAN_TRANSAKSI,
        os.path.basename(FILE_SIMPANAN_PENGAJUAN): FILE_SIMPANAN_PENGAJUAN,
        os.path.basename(FILE_IURAN_SOSIAL): FILE_IURAN_SOSIAL,
        os.path.basename(FILE_PINJAMAN): FILE_PINJAMAN,
        os.path.basename(FILE_PINJAMAN_CICILAN): FILE_PINJAMAN_CICILAN,
        os.path.basename(FILE_USERS): FILE_USERS,
        os.path.basename(FILE_PENDAFTARAN_ANGGOTA): FILE_PENDAFTARAN_ANGGOTA,
        os.path.basename(FILE_IMPORT_LOG): FILE_IMPORT_LOG,
        os.path.basename(FILE_BERITA): FILE_BERITA,
    }

    applied = []
    try:
        for fname in os.listdir(target):
            if fname == 'manifest.json' or fname == 'backup_log.json':
                continue
            src = os.path.join(target, fname)
            dest = mapping.get(fname)
            if dest:
                try:
                    os.makedirs(os.path.dirname(dest), exist_ok=True)
                    shutil.copy2(src, dest)
                    applied.append(fname)
                except Exception:
                    pass
        # record manual apply in backup log
        entries = _load_backup_log()
        entries.append({'type': 'manual', 'stamp': stamp, 'applied_by': session.get('user'), 'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'files': applied})
        _save_backup_log(entries)
        flash('Backup applied successfully (files overwritten).', 'success')
    except Exception as e:
        flash(f'Failed applying backup: {e}', 'danger')

    return redirect(url_for('admin_backups'))


# --- Admin: Roles Management (Tahap 2 feature)
@app.route('/admin/roles')
def admin_roles():
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    if not has_permission('roles.manage'):
        abort(403)
    
    roles_list = []
    if _is_db_mode_enabled():
        roles_list = get_all_roles_from_db()
    else:
        # Fallback ke file-based ROLE_LABELS jika DB mode OFF
        for role_name, label in ROLE_LABELS.items():
            roles_list.append({
                'id_role': role_name,
                'role_name': role_name,
                'deskripsi': f'Legacy role: {label}',
                'created_at': None
            })
    
    return render_template(
        'admin_roles.html',
        roles=roles_list,
        _is_db_mode=_is_db_mode_enabled(),
        csrf_token=_get_or_create_csrf_token()
    )


@app.route('/admin/roles/<role_name>/permissions')
def admin_role_permissions(role_name: str):
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    if not has_permission('roles.manage'):
        abort(403)
    
    if not _is_db_mode_enabled():
        # Fallback ke file-based ROLE_PERMISSIONS
        permissions = sorted(ROLE_PERMISSIONS.get(role_name, set()))
        role_info = {'role_name': role_name, 'deskripsi': ROLE_LABELS.get(role_name, 'Unknown')}
    else:
        role_info = get_role_by_name_from_db(role_name)
        if not role_info:
            abort(404)
        perms = get_permissions_for_role_from_db(role_name)
        permissions = [p['permission_name'] for p in perms]
    
    return render_template(
        'admin_role_permissions.html',
        role_info=role_info,
        permissions=permissions,
        csrf_token=_get_or_create_csrf_token()
    )


# --- Admin: Permissions Audit (Tahap 2 feature, read-only)
@app.route('/admin/permissions')
def admin_permissions():
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    if not has_permission('audit.view'):
        abort(403)
    
    all_perms = []
    if _is_db_mode_enabled():
        all_perms = get_all_permissions_from_db()
    else:
        # Fallback: collect semua permissions dari ROLE_PERMISSIONS
        all_perm_set = set()
        for role_perms in ROLE_PERMISSIONS.values():
            all_perm_set.update(role_perms)
        all_perms = [{'id_permission': p, 'permission_name': p} for p in sorted(all_perm_set)]
    
    return render_template(
        'admin_permissions.html',
        permissions=all_perms,
        _is_db_mode=_is_db_mode_enabled()
    )


def _write_berita_items(items: list[dict]) -> None:
    cleaned_items = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        cleaned_items.append({
            'id': str(item.get('id') or str(uuid.uuid4())),
            'judul': (item.get('judul') or '').strip(),
            'kategori': (item.get('kategori') or 'Pengumuman').strip() or 'Pengumuman',
            'isi': (item.get('isi') or '').strip(),
            'tanggal': (item.get('tanggal') or datetime.now().strftime('%Y-%m-%d')).strip(),
            'status': (item.get('status') or 'Aktif').strip() or 'Aktif',
            'foto': (item.get('foto') or '').strip(),
        })
    cleaned_items.sort(key=lambda x: x.get('tanggal', ''), reverse=True)
    os.makedirs(os.path.dirname(FILE_BERITA), exist_ok=True)
    with open(FILE_BERITA, 'w', encoding='utf-8') as fh:
        json.dump(cleaned_items, fh, ensure_ascii=False, indent=2)

UPLOAD_BUKTI_DIR = os.path.join(app.static_folder, 'uploads', 'bukti_transfer')
ALLOWED_BUKTI_EXT = {'.png', '.jpg', '.jpeg', '.webp'}
UPLOAD_BERITA_DIR = os.path.join(app.static_folder, 'uploads', 'berita')
ALLOWED_BERITA_EXT = {'.png', '.jpg', '.jpeg', '.webp'}
MAX_BERITA_IMAGE_BYTES = int(os.getenv('MAX_BERITA_IMAGE_BYTES', str(2 * 1024 * 1024)))

PAYMENT_STATUS_DRAFT = 'Draft'
PAYMENT_STATUS_WAITING_PAYMENT = 'Menunggu Pembayaran'
PAYMENT_STATUS_WAITING_VERIFICATION = 'Menunggu Verifikasi'
PAYMENT_STATUS_SUCCESS = 'Berhasil'
PAYMENT_STATUS_FAILED = 'Gagal'
PAYMENT_STATUS_EXPIRED = 'Expired'

DB_BACKED_FILES = {
    FILE_ANGGOTA: {
        'table': 'anggota',
        'columns': ANGGOTA_FIELDNAMES,
        'pk': 'id_anggota',
    },
    FILE_USERS: {
        'table': 'users',
        'columns': ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'],
        'pk': 'id_user',
    },
    FILE_SIMPANAN: {
        'table': 'simpanan',
        'columns': SIMPANAN_FIELDNAMES,
        'pk': 'id_anggota',
    },
    FILE_SIMPANAN_TRANSAKSI: {
        'table': 'simpanan_transaksi',
        'columns': SIMPANAN_TRANSAKSI_FIELDNAMES,
        'pk': 'id_transaksi',
    },
    FILE_SIMPANAN_PENGAJUAN: {
        'table': 'simpanan_pengajuan',
        'columns': SIMPANAN_PENGAJUAN_FIELDNAMES,
        'pk': 'id_pengajuan',
    },
    FILE_IURAN_SOSIAL: {
        'table': 'iuran_sosial',
        'columns': IURAN_SOSIAL_FIELDNAMES,
        'pk': 'id_iuran',
    },
    FILE_PINJAMAN: {
        'table': 'pinjaman',
        'columns': PINJAMAN_FIELDNAMES,
        'pk': 'id_pinjaman',
    },
    FILE_PINJAMAN_CICILAN: {
        'table': 'pinjaman_cicilan',
        'columns': CICILAN_FIELDNAMES,
        'pk': 'id_cicilan',
    },
    FILE_PENDAFTARAN_ANGGOTA: {
        'table': 'pendaftaran_anggota',
        'columns': PENDAFTARAN_FIELDNAMES,
        'pk': 'id_pengajuan',
    },
}
_DB_READY_CACHE = None


def _is_db_mode_enabled() -> bool:
    global _DB_READY_CACHE
    if not DATABASE_URL:
        return False
    if _DB_READY_CACHE is None:
        try:
            _DB_READY_CACHE = bool(ping_database())
        except Exception:
            _DB_READY_CACHE = False
    return bool(_DB_READY_CACHE)


def _db_select_rows_for_file(filepath: str):
    cfg = DB_BACKED_FILES.get(filepath)
    if not cfg:
        return []
    table = cfg['table']
    columns = cfg['columns']
    query = text(f"SELECT {', '.join(columns)} FROM {table}")
    try:
        with db_session() as conn:
            rows = conn.execute(query).mappings().all()
        out = []
        for row in rows:
            out.append({col: '' if row.get(col) is None else str(row.get(col)).strip() for col in columns})
        return out
    except SQLAlchemyError as e:
        print(f"Error membaca DB table {table}: {e}")
        return []


def _db_write_rows_for_file(filepath: str, data, fieldnames):
    cfg = DB_BACKED_FILES.get(filepath)
    if not cfg:
        return False
    if cfg.get('read_only'):
        return False
    table = cfg['table']
    pk = cfg.get('pk')
    columns = [c for c in cfg['columns'] if c in fieldnames]
    if not columns:
        return True
    placeholders = ', '.join(f":{c}" for c in columns)
    insert_sql = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})"
    update_columns = [c for c in columns if c != pk]
    if pk and pk in columns:
        if update_columns:
            set_clause = ', '.join(f"{c}=EXCLUDED.{c}" for c in update_columns)
            insert_sql += f" ON CONFLICT ({pk}) DO UPDATE SET {set_clause}"
        else:
            insert_sql += f" ON CONFLICT ({pk}) DO NOTHING"
    insert_stmt = text(insert_sql)
    try:
        with db_session() as conn:
            seen_pk_values = []
            for row_dict in data:
                payload = {}
                for c in columns:
                    value = row_dict.get(c, '')
                    if isinstance(value, str):
                        value = value.strip()
                    if c == 'id_anggota' and value == '':
                        value = None
                    if value == '':
                        value = None
                    payload[c] = value
                if pk and pk in payload and payload.get(pk) is None:
                    continue
                if pk and pk in payload:
                    seen_pk_values.append(payload[pk])
                conn.execute(insert_stmt, payload)

            if pk and pk in columns:
                if seen_pk_values:
                    delete_stmt = text(f"DELETE FROM {table} WHERE {pk} NOT IN :ids").bindparams(bindparam('ids', expanding=True))
                    conn.execute(delete_stmt, {'ids': seen_pk_values})
                else:
                    conn.execute(text(f"DELETE FROM {table}"))
        return True
    except SQLAlchemyError as e:
        print(f"Error menulis DB table {table}: {e}")
        return False

def _save_bukti_transfer(file_storage, id_pinjaman: str) -> str:
    """Simpan file bukti transfer (return path relatif dari /static) atau '' jika gagal."""
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return ''
    filename = secure_filename(file_storage.filename or '')
    _, ext = os.path.splitext(filename.lower())
    if ext not in ALLOWED_BUKTI_EXT:
        return ''
    try:
        file_storage.stream.seek(0, os.SEEK_END)
        file_size = file_storage.stream.tell()
        file_storage.stream.seek(0)
    except Exception:
        file_size = 0
    if file_size <= 0 or file_size > MAX_BUKTI_TRANSFER_BYTES:
        return ''
    header = file_storage.stream.read(16)
    file_storage.stream.seek(0)
    if not _is_supported_image_signature(ext, header):
        return ''
    os.makedirs(UPLOAD_BUKTI_DIR, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d%H%M%S')
    safe_id = re.sub(r'[^a-zA-Z0-9_-]+', '', (id_pinjaman or 'pinjaman'))[:40] or 'pinjaman'
    out_name = f"bukti_{safe_id}_{stamp}_{uuid.uuid4().hex[:8]}{ext}"
    out_path = os.path.join(UPLOAD_BUKTI_DIR, out_name)
    file_storage.save(out_path)
    return f"uploads/bukti_transfer/{out_name}"


def _save_berita_image(file_storage, berita_id: str) -> str:
    """Simpan foto berita dan kembalikan path relatif dari /static; '' jika tidak valid."""
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return ''
    filename = secure_filename(file_storage.filename or '')
    _, ext = os.path.splitext(filename.lower())
    if ext not in ALLOWED_BERITA_EXT:
        return ''
    try:
        file_storage.stream.seek(0, os.SEEK_END)
        file_size = file_storage.stream.tell()
        file_storage.stream.seek(0)
    except Exception:
        file_size = 0
    if file_size <= 0 or file_size > MAX_BERITA_IMAGE_BYTES:
        return ''
    header = file_storage.stream.read(16)
    file_storage.stream.seek(0)
    if not _is_supported_image_signature(ext, header):
        return ''

    os.makedirs(UPLOAD_BERITA_DIR, exist_ok=True)
    safe_id = re.sub(r'[^a-zA-Z0-9_-]+', '', (berita_id or 'berita'))[:40] or 'berita'
    out_name = f"berita_{safe_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    out_path = os.path.join(UPLOAD_BERITA_DIR, out_name)
    file_storage.save(out_path)
    return f"uploads/berita/{out_name}"


def _delete_static_file_relpath(rel_path: str) -> None:
    rel = (rel_path or '').strip().replace('\\', '/')
    if not rel:
        return
    abs_path = os.path.abspath(os.path.join(app.static_folder, rel))
    static_abs = os.path.abspath(app.static_folder)
    if not abs_path.startswith(static_abs):
        return
    try:
        if os.path.exists(abs_path):
            os.remove(abs_path)
    except OSError:
        pass

def _format_currency_idr(value: float) -> str:
    return f"Rp {float(value or 0):,.0f}"

def _generate_va_number(no_anggota: str, id_pinjaman: str) -> str:
    """VA pseudo unik per transaksi untuk memudahkan rekonsiliasi."""
    anggota_digits = re.sub(r'\D', '', no_anggota or '')[-6:] or '000001'
    pinjaman_digits = re.sub(r'\D', '', id_pinjaman or '')[-6:] or uuid.uuid4().hex[:6]
    ts = datetime.now().strftime('%d%H%M%S')
    return f"88{anggota_digits}{pinjaman_digits}{ts}"[:24]

def _mark_expired_cicilan(rows: list) -> bool:
    """Tandai cicilan menunggu yang melewati batas waktu sebagai Expired."""
    changed = False
    now = datetime.now()
    for c in rows:
        if (c.get('status') or '').strip() != 'Menunggu':
            continue
        trx_status = (c.get('status_transaksi') or '').strip()
        if trx_status not in (PAYMENT_STATUS_WAITING_PAYMENT, PAYMENT_STATUS_WAITING_VERIFICATION, ''):
            continue
        expires_raw = (c.get('expires_at') or '').strip()
        if not expires_raw:
            continue
        try:
            expires_at = datetime.strptime(expires_raw, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
        if now > expires_at:
            c['status'] = 'Ditolak'
            c['status_transaksi'] = PAYMENT_STATUS_EXPIRED
            c['tanggal_konfirmasi'] = now.strftime('%Y-%m-%d')
            c['dikonfirmasi_oleh'] = 'system-expired'
            changed = True
    return changed

def _send_email_notification(subject: str, body: str, to_email: str = '') -> bool:
    """Kirim notifikasi email via SMTP env. Return True jika sukses."""
    recipient = (to_email or ADMIN_NOTIFICATION_EMAIL or '').strip()
    if not recipient:
        return False
    smtp_host = os.getenv('SMTP_HOST', '').strip()
    smtp_user = os.getenv('SMTP_USER', '').strip()
    smtp_pass = os.getenv('SMTP_PASS', '').strip()
    smtp_port = int(os.getenv('SMTP_PORT', '587') or 587)
    if not smtp_host or not smtp_user or not smtp_pass:
        return False
    msg = MIMEText(body, _charset='utf-8')
    msg['Subject'] = subject
    msg['From'] = os.getenv('SMTP_FROM', ADMIN_NOTIFICATION_EMAIL or smtp_user)
    msg['To'] = recipient
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(msg['From'], [recipient], msg.as_string())
        return True
    except Exception:
        return False

def _resolve_user_email_for_cicilan(cicilan_row: dict) -> str:
    """Ambil email user pengaju dari username akun jika format email."""
    username = (cicilan_row.get('diajukan_oleh') or '').strip()
    if '@' in username and '.' in username:
        return username
    users = baca_csv(FILE_USERS)
    for u in users:
        if (u.get('username') or '').strip() != username:
            continue
        email = (u.get('email') or '').strip()
        if '@' in email and '.' in email:
            return email
        break
    return ''

def _build_qris_payload_dana(no_dana: str, amount: float, id_pinjaman: str) -> str:
    """Payload teks untuk QR pembayaran ke DANA admin."""
    dana = re.sub(r'\D', '', no_dana or '')
    amt = int(max(float(amount or 0), 0))
    return f"DANA:{dana}|AMOUNT:{amt}|REF:{(id_pinjaman or '')[:16]}"

# ──────────────────────────────────────────────
#  HELPER: Baca & Tulis Excel/CSV
# ──────────────────────────────────────────────
def baca_csv(filepath):
    """Baca file Excel (.xlsx) atau CSV (.csv) dan kembalikan list of dict."""
    if _is_db_mode_enabled() and filepath in DB_BACKED_FILES:
        return _db_select_rows_for_file(filepath)

    if not os.path.exists(filepath):
        return []
    
    file_ext = os.path.splitext(filepath)[1].lower()
    
    # Jika file Excel
    if file_ext == '.xlsx':
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            
            headers = [str(h or '').strip() for h in rows[0]]
            result = []
            for row in rows[1:]:
                row_dict = {}
                for i, header in enumerate(headers):
                    if header:
                        value = row[i] if i < len(row) else None
                        row_dict[header] = '' if value is None else str(value).strip()
                # Hanya tambah jika ada minimal satu field yang terisi
                if any(row_dict.values()):
                    result.append(row_dict)
            return result
        except Exception as e:
            print(f"Error membaca {filepath}: {e}")
            return []
    
    # Fallback untuk CSV (backward compatibility)
    elif file_ext == '.csv':
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                return list(reader) if reader else []
        except Exception as e:
            print(f"Error membaca CSV {filepath}: {e}")
            return []
    
    return []


def tulis_csv(filepath, data, fieldnames):
    """Tulis list of dict ke file Excel (.xlsx)."""
    if _is_db_mode_enabled() and filepath in DB_BACKED_FILES:
        if _db_write_rows_for_file(filepath, data, fieldnames):
            return

    try:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Tulis header
        ws.append(fieldnames)
        
        # Tulis data dengan tipe yang tepat
        for row_dict in data:
            row_values = []
            for field in fieldnames:
                value = row_dict.get(field, '')
                # Coba convert ke angka jika memungkinkan
                if value and str(value).replace(',', '').replace('.', '').isdigit():
                    try:
                        if ',' in str(value):
                            row_values.append(float(str(value).replace(',', '')))
                        else:
                            row_values.append(float(value) if '.' in str(value) else int(value))
                    except (ValueError, TypeError):
                        row_values.append(value)
                else:
                    row_values.append(value)
            ws.append(row_values)
        
        # Format header (bold)
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    except Exception as e:
        print(f"Error menulis {filepath}: {e}")


def migrate_csv_to_excel():
    """Konversi file CSV lama ke format Excel (jika ada)."""
    DATA_DIR = os.path.dirname(FILE_ANGGOTA)
    csv_files = {
        'anggota.csv': (FILE_ANGGOTA, ANGGOTA_FIELDNAMES),
        'simpanan.csv': (FILE_SIMPANAN, SIMPANAN_FIELDNAMES),
        'simpanan_transaksi.csv': (FILE_SIMPANAN_TRANSAKSI, SIMPANAN_TRANSAKSI_FIELDNAMES),
        'iuran_sosial.csv': (FILE_IURAN_SOSIAL, IURAN_SOSIAL_FIELDNAMES),
        'pinjaman.csv': (FILE_PINJAMAN, PINJAMAN_FIELDNAMES),
        'pinjaman_cicilan.csv': (FILE_PINJAMAN_CICILAN, CICILAN_FIELDNAMES),
        'users.csv': (FILE_USERS, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at']),
        'pendaftaran_anggota.csv': (FILE_PENDAFTARAN_ANGGOTA, PENDAFTARAN_FIELDNAMES),
        'import_log.csv': (FILE_IMPORT_LOG, ['waktu', 'user', 'mode', 'nama_file', 'berhasil', 'gagal', 'catatan']),
    }
    
    for csv_filename, (excel_path, fieldnames) in csv_files.items():
        csv_path = os.path.join(DATA_DIR, csv_filename)
        if os.path.exists(csv_path) and not os.path.exists(excel_path):
            try:
                # Baca CSV lama
                with open(csv_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader) if reader else []
                
                # Tulis ke Excel baru
                tulis_csv(excel_path, rows, fieldnames)
                
                # Hapus file CSV lama
                os.remove(csv_path)
                print(f"Migrated {csv_filename} -> {csv_filename.replace('.csv', '.xlsx')}")
            except Exception as e:
                print(f"Error migrating {csv_filename}: {e}")


def init_csv():
    """Inisialisasi file Excel jika belum ada."""
    db_mode = _is_db_mode_enabled()
    if not os.path.exists(FILE_ANGGOTA):
        if not db_mode:
            tulis_csv(FILE_ANGGOTA, [], ANGGOTA_FIELDNAMES)
    if not os.path.exists(FILE_SIMPANAN):
        tulis_csv(FILE_SIMPANAN, [], SIMPANAN_FIELDNAMES)
    if not os.path.exists(FILE_SIMPANAN_TRANSAKSI):
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, [], SIMPANAN_TRANSAKSI_FIELDNAMES)
    if not os.path.exists(FILE_IURAN_SOSIAL):
        tulis_csv(FILE_IURAN_SOSIAL, [], IURAN_SOSIAL_FIELDNAMES)
    if not os.path.exists(FILE_PINJAMAN):
        tulis_csv(FILE_PINJAMAN, [], PINJAMAN_FIELDNAMES)
    if not os.path.exists(FILE_PINJAMAN_CICILAN):
        tulis_csv(FILE_PINJAMAN_CICILAN, [], CICILAN_FIELDNAMES)
    if db_mode:
        if not baca_csv(FILE_USERS):
            default_users, _ = _seed_default_user_rows([])
            tulis_csv(FILE_USERS, default_users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
        else:
            users = baca_csv(FILE_USERS)
            users, added = _seed_default_user_rows(users)
            if added:
                tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    elif not os.path.exists(FILE_USERS):
        default_users, _ = _seed_default_user_rows([])
        tulis_csv(FILE_USERS, default_users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    else:
        users = baca_csv(FILE_USERS)
        users, added = _seed_default_user_rows(users)
        if added:
            tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    # Mode DB: pendaftaran_anggota dikelola di PostgreSQL, tidak perlu file Excel harian.
    if (not db_mode) and (not os.path.exists(FILE_PENDAFTARAN_ANGGOTA)):
        tulis_csv(FILE_PENDAFTARAN_ANGGOTA, [], PENDAFTARAN_FIELDNAMES)


def kategori_pinjaman_dari_tenor(tenor_bulan: int) -> str:
    """Kelompokkan pinjaman impor menjadi Jangka Pendek atau Jangka Panjang."""
    tenor = max(int(tenor_bulan or 0), 0)
    if tenor >= PROVISI_MIN_TENOR_BULAN:
        return 'Jangka Panjang'
    return 'Jangka Pendek'


def bootstrap_storage_files():
    """Jalankan bootstrap file setelah semua fungsi global terdefinisi."""
    # Mode file (legacy): lakukan migrasi CSV->Excel dan bootstrap file.
    # Mode DB: operasi harian tidak bergantung Excel, jadi jangan jalankan otomatis.
    if not _is_db_mode_enabled():
        migrate_csv_to_excel()
        init_csv()
        _ensure_role_linked_demo_data()


def migrate_simpanan_ke_saldo():
    """Migrasi skema lama (transaksi) ke saldo per id_anggota."""
    rows = baca_csv(FILE_SIMPANAN)
    if not rows:
        return
    if 'total_simpanan' in rows[0] and 'id_anggota' in rows[0]:
        merged = _dedupe_rows_simpanan(rows)
        tulis_csv(FILE_SIMPANAN, merged, SIMPANAN_FIELDNAMES)
        return
    agg = {}
    for s in rows:
        id_a = (s.get('id_anggota') or '').strip()
        if not id_a:
            continue
        if (s.get('status') or 'Disetujui') != 'Disetujui':
            continue
        try:
            agg[id_a] = agg.get(id_a, 0.0) + float(s.get('jumlah') or 0)
        except (TypeError, ValueError):
            continue
    out = [{'id_anggota': k, 'total_simpanan': str(round(v, 2))} for k, v in sorted(agg.items())]
    tulis_csv(FILE_SIMPANAN, out, SIMPANAN_FIELDNAMES)


def migrate_pinjaman_ke_saldo():
    """Migrasi berbagai skema lama ke format multi-baris + jenis pinjaman."""
    rows = baca_csv(FILE_PINJAMAN)
    if not rows:
        return
    sample = rows[0]

    if 'id_pinjaman' in sample:
        normalized = []
        for r in rows:
            row = {k: (r.get(k) if r.get(k) is not None else '') for k in PINJAMAN_FIELDNAMES}
            if not row.get('id_pinjaman'):
                row['id_pinjaman'] = str(uuid.uuid4())
            if not row.get('tenor_awal'):
                row['tenor_awal'] = row.get('tenor_bulan') or ''
            jenis_raw = (row.get('jenis_pinjaman') or '').strip()
            if not jenis_raw or jenis_raw == JENIS_IMPORT_CSV:
                try:
                    tenor_norm = int(float(row.get('tenor_bulan') or 0))
                except (TypeError, ValueError):
                    tenor_norm = 0
                row['jenis_pinjaman'] = kategori_pinjaman_dari_tenor(tenor_norm)
            normalized.append(row)
        tulis_csv(FILE_PINJAMAN, normalized, PINJAMAN_FIELDNAMES)
        return

    # Skema saldo 3 kolom (id_anggota, total_pinjaman, tenor)
    if 'total_pinjaman' in sample and 'tenor' in sample and 'plafon' not in sample:
        amap = {a.get('id_anggota'): a for a in baca_csv(FILE_ANGGOTA)}
        out = []
        for r in rows:
            id_a = (r.get('id_anggota') or '').strip()
            if not id_a:
                continue
            try:
                plaf = float(r.get('total_pinjaman') or 0)
                ten = int(float(r.get('tenor') or 0))
            except (TypeError, ValueError):
                continue
            ag = amap.get(id_a) or {}
            bunga_p = bunga_dari_tenor(ten) if ten > 0 else 0.0
            total_bayar, _ = hitung_total_bayar_flat(plaf, bunga_p, ten, kategori_pinjaman_dari_tenor(ten))
            # Hitung cicilan bulanan: cicilan = (p / b) + (p × j%)
            cic = hitung_cicilan_bulanan(plaf, bunga_p, ten)
            out.append({
                'id_pinjaman': str(uuid.uuid4()),
                'id_anggota': id_a,
                'nama_anggota': ag.get('nama_lengkap', ''),
                'no_anggota': ag.get('no_anggota', ''),
                'jenis_pinjaman': kategori_pinjaman_dari_tenor(ten),
                'plafon': str(round(plaf, 2)),
                'tenor_awal': str(ten),
                'tenor_bulan': str(ten),
                'bunga_persen': str(bunga_p),
                'total_bayar': str(round(total_bayar, 2)),
                'cicilan_per_bulan': str(round(cic, 2)),
                'sisa_pinjaman': str(round(plaf, 2)),
                'tanggal_pengajuan': '-',
                'status': 'Disetujui',
                'tanggal_lunas': '',
            })
        tulis_csv(FILE_PINJAMAN, out, PINJAMAN_FIELDNAMES)
        return

    # Skema transaksi lama (plafon, tenor_bulan, ...)
    agg_tot = {}
    agg_tenor = {}
    for p in rows:
        id_a = (p.get('id_anggota') or '').strip()
        if not id_a:
            continue
        st = p.get('status', '')
        if st not in ('Disetujui', 'Lunas', 'Menunggu'):
            continue
        try:
            plaf = float(p.get('plafon') or 0)
            t = int(float(p.get('tenor_bulan') or 0))
        except (TypeError, ValueError):
            continue
        agg_tot[id_a] = agg_tot.get(id_a, 0.0) + plaf
        agg_tenor[id_a] = max(agg_tenor.get(id_a, 0), t)
    amap = {a.get('id_anggota'): a for a in baca_csv(FILE_ANGGOTA)}
    out = []
    for id_a in sorted(agg_tot.keys()):
        ag = amap.get(id_a) or {}
        plaf = agg_tot[id_a]
        ten = max(agg_tenor.get(id_a, 0), 0)
        bunga_p = bunga_dari_tenor(ten) if ten > 0 else 0.0
        total_bayar, _ = hitung_total_bayar_flat(plaf, bunga_p, ten, kategori_pinjaman_dari_tenor(ten))
        # Hitung cicilan bulanan: cicilan = (p / b) + (p × j%)
        cic = hitung_cicilan_bulanan(plaf, bunga_p, ten)
        out.append({
            'id_pinjaman': str(uuid.uuid4()),
            'id_anggota': id_a,
            'nama_anggota': ag.get('nama_lengkap', ''),
            'no_anggota': ag.get('no_anggota', ''),
            'jenis_pinjaman': kategori_pinjaman_dari_tenor(ten),
            'plafon': str(round(plaf, 2)),
            'tenor_awal': str(ten),
            'tenor_bulan': str(ten),
            'bunga_persen': str(bunga_p),
            'total_bayar': str(round(total_bayar, 2)),
            'cicilan_per_bulan': str(round(cic, 2)),
            'sisa_pinjaman': str(round(plaf, 2)),
            'tanggal_pengajuan': '-',
            'status': 'Disetujui',
            'tanggal_lunas': '',
        })
    tulis_csv(FILE_PINJAMAN, out, PINJAMAN_FIELDNAMES)


def _dedupe_rows_simpanan(rows):
    seen = {}
    for r in rows:
        id_a = (r.get('id_anggota') or '').strip()
        if not id_a:
            continue
        try:
            v = float(r.get('total_simpanan') or 0)
        except (TypeError, ValueError):
            v = 0.0
        seen[id_a] = seen.get(id_a, 0.0) + v
    return [{'id_anggota': k, 'total_simpanan': str(round(v, 2))} for k, v in sorted(seen.items())]


def _dedupe_rows_pinjaman(rows):
    """Satukan baris dengan (id_anggota, jenis_pinjaman) sama — untuk impor CSV."""
    if not rows:
        return []
    if 'jenis_pinjaman' not in rows[0]:
        return rows
    merged = {}
    order = []
    for r in rows:
        id_a = (r.get('id_anggota') or '').strip()
        tenor_raw = r.get('tenor_bulan') or r.get('tenor') or 0
        try:
            tenor_norm = int(float(tenor_raw))
        except (TypeError, ValueError):
            tenor_norm = 0
        jn = (r.get('jenis_pinjaman') or '').strip()
        if not jn or jn == JENIS_IMPORT_CSV:
            jn = kategori_pinjaman_dari_tenor(tenor_norm)
        if not id_a:
            continue
        key = (id_a, jn)
        if key not in merged:
            merged[key] = dict(r)
            order.append(key)
            continue
        m = merged[key]
        try:
            plaf = float(m.get('plafon') or 0) + float(r.get('plafon') or 0)
            ten = max(int(float(m.get('tenor_bulan') or 0)), int(float(r.get('tenor_bulan') or 0)))
        except (TypeError, ValueError):
            continue
        bp = float(m.get('bunga_persen') or 0)
        tb, _ = hitung_total_bayar_flat(plaf, bp, ten, jn)
        cic = hitung_cicilan_bulanan(plaf, bp, ten)
        m['plafon'] = str(round(plaf, 2))
        m['tenor_awal'] = str(ten)
        m['tenor_bulan'] = str(ten)
        m['total_bayar'] = str(round(tb, 2))
        m['cicilan_per_bulan'] = str(round(cic, 2))
        if m.get('status') == 'Disetujui':
            m['sisa_pinjaman'] = str(round(plaf, 2))
    return [merged[k] for k in order]


def merge_akumulasi(simpanan_rows: list, id_anggota: str, tambah: float) -> None:
    """Akumulasi simpanan per id_anggota (satu baris per anggota)."""
    tambah = max(float(tambah or 0), 0.0)
    for r in simpanan_rows:
        if r.get('id_anggota') == id_anggota:
            cur = float(r.get('total_simpanan') or 0)
            r['total_simpanan'] = str(round(cur + tambah, 2))
            return
    simpanan_rows.append({'id_anggota': id_anggota, 'total_simpanan': str(round(tambah, 2))})


def catat_simpanan_pokok_awal(
    anggota_row: dict,
    simpanan_rows: list,
    simpanan_transaksi_rows: list,
    jumlah_pokok: float,
    diajukan_oleh: str = '',
    keterangan: str = 'Simpanan Pokok anggota baru',
) -> None:
    """Catat Simpanan Pokok sekali saja untuk validasi anggota baru."""
    try:
        jumlah = max(float(jumlah_pokok or 0), 0.0)
    except (TypeError, ValueError):
        jumlah = 0.0
    if jumlah <= 0:
        anggota_row['simpanan_pokok'] = '0'
        return

    id_anggota = (anggota_row.get('id_anggota') or '').strip()
    if not id_anggota:
        return

    anggota_row['simpanan_pokok'] = str(round(jumlah, 2))
    sudah_ada = next(
        (
            t for t in simpanan_transaksi_rows
            if t.get('id_anggota') == id_anggota
            and normalize_jenis_simpanan(t.get('jenis_simpanan')) == 'Simpanan Pokok'
        ),
        None,
    )
    if sudah_ada:
        return

    merge_akumulasi(simpanan_rows, id_anggota, jumlah)
    simpanan_transaksi_rows.append({
        'id_transaksi': str(uuid.uuid4()),
        'id_anggota': id_anggota,
        'no_anggota': anggota_row.get('no_anggota', ''),
        'nama_anggota': anggota_row.get('nama_lengkap', ''),
        'tanggal': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'jenis_simpanan': 'Simpanan Pokok',
        'jumlah': str(round(jumlah, 2)),
        'keterangan': keterangan,
        'diajukan_oleh': diajukan_oleh or session.get('user') or '',
    })


def merge_pinjaman_akumulasi(
    pinjaman_rows: list,
    id_anggota: str,
    tambah_pinjaman: float,
    tenor_baru: int,
    jenis: str = None,
) -> None:
    """Akumulasi pinjaman per (anggota, jenis). Jangka pendek & panjang tidak digabung."""
    jenis_key = (jenis or kategori_pinjaman_dari_tenor(tenor_baru)).strip()
    tambah_pinjaman = max(float(tambah_pinjaman or 0), 0.0)
    tenor_baru = max(int(tenor_baru or 0), 0)
    for r in pinjaman_rows:
        if r.get('id_anggota') != id_anggota:
            continue
        if (r.get('jenis_pinjaman') or '').strip() != jenis_key:
            continue
        plaf = float(r.get('plafon') or 0) + tambah_pinjaman
        ten = max(int(float(r.get('tenor_bulan') or 0)), tenor_baru)
        bp = float(r.get('bunga_persen') or 0)
        tb, _ = hitung_total_bayar_flat(plaf, bp, ten, jenis_key)
        cic = hitung_cicilan_bulanan(plaf, bp, ten)
        r['plafon'] = str(round(plaf, 2))
        r['tenor_bulan'] = str(ten)
        r['total_bayar'] = str(round(tb, 2))
        r['cicilan_per_bulan'] = str(round(cic, 2))
        if r.get('status') == 'Disetujui':
            r['sisa_pinjaman'] = str(round(plaf, 2))
        return
    try:
        bp = float(bunga_untuk_jenis_pinjaman(jenis_key, tenor_baru)) if tenor_baru > 0 else 0.0
    except Exception:
        bp = bunga_dari_tenor(tenor_baru) if tenor_baru > 0 else 0.0
    plaf = tambah_pinjaman
    tb, _ = hitung_total_bayar_flat(plaf, bp, tenor_baru, jenis_key)
    cic = hitung_cicilan_bulanan(plaf, bp, tenor_baru)
    pinjaman_rows.append({
        'id_pinjaman': str(uuid.uuid4()),
        'id_anggota': id_anggota,
        'nama_anggota': '',
        'no_anggota': '',
        'jenis_pinjaman': jenis_key,
        'plafon': str(round(plaf, 2)),
        'tenor_awal': str(tenor_baru),
        'tenor_bulan': str(tenor_baru),
        'bunga_persen': str(bp),
        'total_bayar': str(round(tb, 2)),
        'cicilan_per_bulan': str(round(cic, 2)),
        'sisa_pinjaman': str(round(plaf, 2)),
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Disetujui',
        'tanggal_lunas': '',
    })


def merge_pinjaman_sama_anggota(pinjaman_rows: list, id_pinjaman_baru: str, id_anggota: str) -> None:
    """Merge pinjaman yang baru disetujui dengan pinjaman existing dari anggota sama.
    Gabungkan plafon dengan pinjaman baru, lalu pakai tenor dan bunga dari pinjaman terbaru."""
    pinjaman_baru = None
    pinjaman_existing = None
    idx_existing = -1

    # Cari pinjaman baru yang baru dikonfirmasi
    for p in pinjaman_rows:
        if p.get('id_pinjaman') == id_pinjaman_baru:
            pinjaman_baru = p
            break

    if not pinjaman_baru or pinjaman_baru.get('status') != 'Disetujui':
        return

    # Cari pinjaman approved existing dari anggota sama dengan jenis sama
    jenis_baru = (pinjaman_baru.get('jenis_pinjaman') or '').strip()
    for i, p in enumerate(pinjaman_rows):
        jenis_existing = (p.get('jenis_pinjaman') or '').strip()
        if (p.get('id_anggota') == id_anggota and 
            p.get('status') == 'Disetujui' and 
            p.get('id_pinjaman') != id_pinjaman_baru and
            jenis_existing == jenis_baru):
            pinjaman_existing = p
            idx_existing = i
            break

    # Jika ada pinjaman existing, merge keduanya
    if pinjaman_existing and idx_existing >= 0:
        try:
            plaf_baru = float(pinjaman_baru.get('plafon') or 0)
            plaf_existing = float(pinjaman_existing.get('plafon') or 0)
            cic_baru = float(pinjaman_baru.get('cicilan_per_bulan') or 0)
            cic_existing = float(pinjaman_existing.get('cicilan_per_bulan') or 0)
            ten_baru = int(float(pinjaman_baru.get('tenor_bulan') or 0))
            bp = float(pinjaman_baru.get('bunga_persen') or 0)

            # Gabung plafon
            plaf_merged = plaf_baru + plaf_existing

            # Gunakan tenor terbaru dari pinjaman baru.
            ten_merged = ten_baru

            # Total kewajiban mengikuti tenor terbaru, bukan tenor gabungan.
            tb_merged = hitung_total_bayar_tanpa_provisi(plaf_merged, bp, ten_merged, jenis_baru)

            # Hitung cicilan merged
            cic_merged = hitung_cicilan_bulanan(plaf_merged, bp, ten_merged)

            # Update pinjaman existing dengan data merged
            pinjaman_existing['plafon'] = str(round(plaf_merged, 2))
            pinjaman_existing['tenor_bulan'] = str(ten_merged)
            pinjaman_existing['total_bayar'] = str(round(tb_merged, 2))
            pinjaman_existing['cicilan_per_bulan'] = str(round(cic_merged, 2))
            sisa_existing = saldo_pinjaman_aktual(pinjaman_existing)
            pinjaman_existing['sisa_pinjaman'] = str(round(sisa_existing + plaf_baru, 2))

            # Hapus pinjaman baru (karena sudah digabung ke existing)
            pinjaman_rows.pop(pinjaman_rows.index(pinjaman_baru))
        except (TypeError, ValueError):
            pass


def get_pinjaman_saldo(pinjaman_rows: list, id_anggota: str):
    for p in pinjaman_rows:
        if p.get('id_anggota') == id_anggota:
            return p
    return None


def cicilan_per_bulan_saldo(p: dict) -> float:
    if p.get('cicilan_per_bulan') not in (None, ''):
        try:
            return float(p['cicilan_per_bulan'])
        except (TypeError, ValueError):
            pass
    tot = saldo_pinjaman_aktual(p)
    ten = int(float(p.get('tenor_bulan') or p.get('tenor') or 0))
    if ten <= 0 or tot <= 0:
        return 0.0
    return tot / ten


def tenor_sisa_pinjaman_aktual(p: dict) -> int:
    """Jumlah tenor tersisa yang tersimpan di data pinjaman."""
    try:
        tenor = int(float(p.get('tenor_bulan') or p.get('tenor_awal') or p.get('tenor') or 0))
    except (TypeError, ValueError):
        tenor = 0
    return max(tenor, 0)


def enrich_simpanan_untuk_tampilan(simpanan_rows: list, anggota_rows: list) -> list:
    amap = {a.get('id_anggota'): a for a in anggota_rows}
    out = []
    for s in simpanan_rows:
        a = amap.get(s.get('id_anggota'), {})
        out.append({
            **s,
            'no_anggota': a.get('no_anggota', ''),
            'nama_anggota': a.get('nama', ''),
            'jumlah': s.get('total_simpanan', '0'),
            'tanggal': '-',
            'status': 'Disetujui',
            'jenis_simpanan': 'Saldo',
            'keterangan': 'Total saldo per anggota',
            'id_simpanan': s.get('id_anggota', ''),
        })
    return out


def enrich_pinjaman_untuk_tampilan(pinjaman_rows: list, anggota_rows: list) -> list:
    amap = {a.get('id_anggota'): a for a in anggota_rows}
    riwayat_jenis_map = {}

    for p in pinjaman_rows:
        id_a = p.get('id_anggota', '')
        tenor_raw_hist = p.get('tenor_bulan', p.get('tenor', '0'))
        try:
            tenor_norm_hist = int(float(tenor_raw_hist or 0))
        except (TypeError, ValueError):
            tenor_norm_hist = 0
        jenis_hist = (p.get('jenis_pinjaman') or '').strip()
        if not jenis_hist or jenis_hist == JENIS_IMPORT_CSV:
            jenis_hist = kategori_pinjaman_dari_tenor(tenor_norm_hist)
        if not id_a or not jenis_hist:
            continue
        riwayat_jenis_map.setdefault(id_a, [])
        if jenis_hist not in riwayat_jenis_map[id_a]:
            riwayat_jenis_map[id_a].append(jenis_hist)

    out = []
    for p in pinjaman_rows:
        a = amap.get(p.get('id_anggota'), {})
        tenor_raw = p.get('tenor_bulan', p.get('tenor', '0'))
        try:
            tenor_norm = int(float(tenor_raw or 0))
        except (TypeError, ValueError):
            tenor_norm = 0
        jenis_raw = (p.get('jenis_pinjaman') or '').strip()
        if not jenis_raw or jenis_raw == JENIS_IMPORT_CSV:
            jenis_raw = kategori_pinjaman_dari_tenor(tenor_norm)
        if p.get('plafon') is not None or p.get('id_pinjaman'):
            plaf = float(p.get('plafon') or 0)
            sisa = saldo_pinjaman_aktual(p)
            cic = float(p.get('cicilan_per_bulan') or 0) or cicilan_per_bulan_saldo(p)
            st = p.get('status') or 'Menunggu'
            jatuh_tempo = tanggal_jatuh_tempo_pinjaman(p)
            tenor_tampil = tenor_sisa_pinjaman_aktual(p)
            tenor_provisi = int(float(p.get('tenor_awal') or p.get('tenor_bulan') or 0))
            provisi_nominal = provisi_nominal_pinjaman(jenis_raw, plaf, tenor_provisi)
            out.append({
                **p,
                'no_anggota': p.get('no_anggota') or a.get('no_anggota', ''),
                'nama_anggota': p.get('nama_anggota') or a.get('nama_lengkap', ''),
                'plafon': str(plaf),
                'tenor_bulan': str(max(int(tenor_tampil), 0)),
                'cicilan_per_bulan': str(round(cic, 2)),
                'sisa_pinjaman': str(round(sisa, 2)),
                'status': st,
                'jatuh_tempo': jatuh_tempo.strftime('%Y-%m-%d') if jatuh_tempo else '',
                'telat_bayar': pinjaman_telat_bayar_aktual(p),
                'tanggal_pengajuan': p.get('tanggal_pengajuan') or '-',
                'id_pinjaman': p.get('id_pinjaman', ''),
                'id_anggota': p.get('id_anggota', ''),
                'jenis_pinjaman': jenis_raw,
                'riwayat_jenis_pinjaman': ', '.join(riwayat_jenis_map.get(p.get('id_anggota', ''), [jenis_raw])),
                'jenis_simpanan': (p.get('jenis_simpanan') or 'Manasuka').strip() or 'Manasuka',
                'provisi_nominal': str(round(provisi_nominal, 2)),
            })
            continue
        tot = float(p.get('total_pinjaman') or 0)
        ten = tenor_norm
        cic = cicilan_per_bulan_saldo(p)
        st = 'Disetujui' if tot > 0 else 'Lunas'
        out.append({
            **p,
            'no_anggota': a.get('no_anggota', ''),
            'nama_anggota': a.get('nama', ''),
            'plafon': p.get('total_pinjaman', '0'),
            'tenor_bulan': p.get('tenor', '0'),
            'cicilan_per_bulan': str(round(cic, 2)),
            'sisa_pinjaman': str(round(tot, 2)),
            'status': st,
            'jatuh_tempo': '',
            'telat_bayar': False,
            'tanggal_pengajuan': '-',
            'jenis_pinjaman': jenis_raw or 'Saldo',
            'riwayat_jenis_pinjaman': ', '.join(riwayat_jenis_map.get(p.get('id_anggota', ''), [jenis_raw or 'Saldo'])),
            'jenis_simpanan': (p.get('jenis_simpanan') or 'Manasuka').strip() or 'Manasuka',
            'provisi_nominal': '0',
            'id_pinjaman': p.get('id_anggota', ''),
        })
    return out


def ensure_anggota_schema():
    """Pastikan data anggota punya kolom lengkap termasuk data bank."""
    rows = baca_csv(FILE_ANGGOTA)
    if not rows:
        return
    required = {
        'nik', 'penghasilan_bersih', 'cicilan_lain', 'simpanan_pokok',
        'no_rekening', 'nama_bank',
    }
    if required.issubset(set(rows[0].keys())):
        return
    for r in rows:
        r['nik'] = r.get('nik', '')
        r['no_rekening'] = r.get('no_rekening', '')
        r['nama_bank'] = r.get('nama_bank', '')
        r['penghasilan_bersih'] = r.get('penghasilan_bersih', '0')
        r['cicilan_lain'] = r.get('cicilan_lain', '0')
        r['simpanan_pokok'] = r.get('simpanan_pokok', '0')
    tulis_csv(FILE_ANGGOTA, rows, ANGGOTA_FIELDNAMES)


def ensure_pinjaman_plafon_schema():
    """Kompatibilitas: migrasi skema lama ke saldo per id_anggota."""
    migrate_pinjaman_ke_saldo()


def ensure_pinjaman_cicilan_schema():
    """Pastikan file pinjaman_cicilan.csv memiliki fieldnames terbaru."""
    if not os.path.exists(FILE_PINJAMAN_CICILAN):
        tulis_csv(FILE_PINJAMAN_CICILAN, [], CICILAN_FIELDNAMES)
        return
    rows = baca_csv(FILE_PINJAMAN_CICILAN)
    normalized = []
    for r in rows:
        item = {k: (r.get(k) or '') for k in CICILAN_FIELDNAMES}
        if not item.get('status_transaksi'):
            st = (item.get('status') or '').strip()
            if st == 'Disetujui':
                item['status_transaksi'] = PAYMENT_STATUS_SUCCESS
            elif st == 'Ditolak':
                item['status_transaksi'] = PAYMENT_STATUS_FAILED
            elif st == 'Menunggu':
                item['status_transaksi'] = PAYMENT_STATUS_WAITING_VERIFICATION
            else:
                item['status_transaksi'] = PAYMENT_STATUS_DRAFT
        normalized.append(item)
    tulis_csv(FILE_PINJAMAN_CICILAN, normalized, CICILAN_FIELDNAMES)


def ensure_import_log_schema():
    # Mode DB: schema dikelola via SQL migrasi (db/schema.sql), tidak membuat file Excel.
    if _is_db_mode_enabled():
        return
    if not os.path.exists(FILE_IMPORT_LOG):
        tulis_csv(
            FILE_IMPORT_LOG,
            [],
            ['waktu', 'user', 'mode', 'nama_file', 'berhasil', 'gagal', 'catatan'],
        )


def ensure_import_preview_dir():
    os.makedirs(IMPORT_PREVIEW_DIR, exist_ok=True)


def migrate_sisa_pinjaman_aktif_ke_total():
    """Migrasi data lama: pinjaman aktif memakai sisa pokok (plafon)."""
    rows = baca_csv(FILE_PINJAMAN)
    if not rows:
        return
    changed = False
    for r in rows:
        status = (r.get('status') or '').strip()
        if status not in ('Menunggu', 'Disetujui'):
            continue
        try:
            plaf = float(r.get('plafon') or r.get('total_pinjaman') or 0)
        except (TypeError, ValueError):
            plaf = 0.0
        if plaf > 0:
            try:
                sisa_now = float(r.get('sisa_pinjaman') or 0)
            except (TypeError, ValueError):
                sisa_now = 0.0
            if sisa_now > plaf:
                r['sisa_pinjaman'] = str(round(plaf, 2))
                changed = True
                continue
        try:
            sisa = float(r.get('sisa_pinjaman') or 0)
        except (TypeError, ValueError):
            sisa = 0.0
        if sisa > 0:
            continue
        if plaf > 0:
            r['sisa_pinjaman'] = str(round(plaf, 2))
            changed = True
    if changed:
        tulis_csv(FILE_PINJAMAN, rows, PINJAMAN_FIELDNAMES)


def ensure_pendaftaran_schema():
    """Migrasi kolom pengajuan anggota (penghasilan, cicilan, id anggota dibuat, dll.)."""
    # Mode DB: schema dikelola via SQL migrasi (db/schema.sql), tidak mengubah file Excel.
    if _is_db_mode_enabled():
        return
    if not os.path.exists(FILE_PENDAFTARAN_ANGGOTA):
        return
    rows = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    if rows:
        sample = rows[0]
        if all(k in sample for k in PENDAFTARAN_FIELDNAMES):
            return
        for r in rows:
            for k in PENDAFTARAN_FIELDNAMES:
                if k not in r:
                    r[k] = '0' if k in ('penghasilan_bersih', 'cicilan_lain', 'simpanan_pokok') else ''
        tulis_csv(FILE_PENDAFTARAN_ANGGOTA, rows, PENDAFTARAN_FIELDNAMES)
        return
    # File kosong: tulis ulang dengan header schema terbaru.
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, [], PENDAFTARAN_FIELDNAMES)


ensure_pendaftaran_schema()

PASSWORD_POLICY_MSG = (
    'Password minimal 8 karakter dan wajib memuat huruf besar (A-Z), huruf kecil (a-z), '
    'angka (0-9), dan simbol khusus (contoh: !@#$).'
)


def validate_password_policy(password: str):
    """Minimal 8 karakter, kombinasi huruf besar/kecil, angka, dan simbol khusus."""
    if not password or len(password) < 8:
        return False, PASSWORD_POLICY_MSG
    if not re.search(r'[A-Z]', password):
        return False, PASSWORD_POLICY_MSG
    if not re.search(r'[a-z]', password):
        return False, PASSWORD_POLICY_MSG
    if not re.search(r'\d', password):
        return False, PASSWORD_POLICY_MSG
    simbol = set('!@#$%^&*()_+-=[]{}|;:,.<>?/`~')
    if not any(c in password for c in simbol):
        return False, PASSWORD_POLICY_MSG
    return True, ''


def get_anggota_by_id(id_anggota: str):
    for a in baca_csv(FILE_ANGGOTA):
        if a.get('id_anggota') == id_anggota:
            return a
    return None


def generate_no_anggota_berikutnya(data_anggota):
    """Generate nomor AGT-XXXX berbasis nomor tertinggi yang sudah ada."""
    max_no = 0
    for a in data_anggota:
        no = (a.get('no_anggota') or '').strip()
        if no.startswith('AGT-'):
            bagian = no.replace('AGT-', '')
            if bagian.isdigit():
                max_no = max(max_no, int(bagian))
    return f"AGT-{(max_no + 1):04d}"


def ensure_simpanan_schema():
    """Kompatibilitas: migrasi skema lama ke saldo per id_anggota."""
    migrate_simpanan_ke_saldo()


def ensure_simpanan_transaksi_schema():
    """Pastikan file transaksi simpanan tersedia dengan skema terbaru."""
    if not os.path.exists(FILE_SIMPANAN_TRANSAKSI):
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, [], SIMPANAN_TRANSAKSI_FIELDNAMES)
        return
    rows = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    normalized = []
    for r in rows:
        normalized.append({k: (r.get(k) or '') for k in SIMPANAN_TRANSAKSI_FIELDNAMES})
    tulis_csv(FILE_SIMPANAN_TRANSAKSI, normalized, SIMPANAN_TRANSAKSI_FIELDNAMES)


def ensure_simpanan_pengajuan_schema():
    """Pastikan file pengajuan simpanan tersedia dengan skema terbaru."""
    if not os.path.exists(FILE_SIMPANAN_PENGAJUAN):
        tulis_csv(FILE_SIMPANAN_PENGAJUAN, [], SIMPANAN_PENGAJUAN_FIELDNAMES)
        return
    rows = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    normalized = []
    for r in rows:
        normalized.append({k: (r.get(k) or '') for k in SIMPANAN_PENGAJUAN_FIELDNAMES})
    tulis_csv(FILE_SIMPANAN_PENGAJUAN, normalized, SIMPANAN_PENGAJUAN_FIELDNAMES)


def ensure_iuran_sosial_schema():
    """Pastikan file iuran sosial tersedia dengan skema terbaru."""
    if not os.path.exists(FILE_IURAN_SOSIAL):
        tulis_csv(FILE_IURAN_SOSIAL, [], IURAN_SOSIAL_FIELDNAMES)
        return
    rows = baca_csv(FILE_IURAN_SOSIAL)
    normalized = []
    for r in rows:
        normalized.append({k: (r.get(k) or '') for k in IURAN_SOSIAL_FIELDNAMES})
    tulis_csv(FILE_IURAN_SOSIAL, normalized, IURAN_SOSIAL_FIELDNAMES)


def normalize_jenis_simpanan(value: str) -> str:
    """Normalisasi label simpanan lama ke nama produk baru."""
    jenis = (value or '').strip()
    alias = {
        'Hari Koperasi': 'Simpanan Hari Koperasi',
        'Simpan Hari Koperasi': 'Simpanan Hari Koperasi',
        'Pensiun': 'Simpanan Pensiun',
        'Simpan Pensiun': 'Simpanan Pensiun',
        'Hari Raya': 'Simpanan Hari Raya',
        'Manasuka': 'Simpanan Manasuka',
        'Pendidikan': 'Simpanan Pendidikan',
    }
    return alias.get(jenis, jenis)


def migrate_iuran_sosial_legacy():
    """Pindahkan catatan Iuran Sosial lama dari file transaksi simpanan ke file iuran khusus."""
    ensure_simpanan_transaksi_schema()
    ensure_iuran_sosial_schema()
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    iuran_sosial = baca_csv(FILE_IURAN_SOSIAL)
    existing = {
        (
            (r.get('id_anggota') or '').strip(),
            (r.get('tanggal') or '').strip(),
            str(r.get('jumlah') or '').strip(),
        )
        for r in iuran_sosial
    }

    kept_transaksi = []
    moved = 0
    for t in simpanan_transaksi:
        jenis = normalize_jenis_simpanan(t.get('jenis_simpanan'))
        if jenis != 'Iuran Sosial':
            kept_transaksi.append(t)
            continue
        key = (
            (t.get('id_anggota') or '').strip(),
            (t.get('tanggal') or '').strip(),
            str(t.get('jumlah') or '').strip(),
        )
        if key not in existing:
            iuran_sosial.append({
                'id_iuran': str(t.get('id_transaksi') or uuid.uuid4()),
                'id_anggota': t.get('id_anggota', ''),
                'no_anggota': t.get('no_anggota', ''),
                'nama_anggota': t.get('nama_anggota', ''),
                'tanggal': t.get('tanggal', ''),
                'jumlah': t.get('jumlah', '0'),
                'keterangan': t.get('keterangan') or 'Catatan iuran sosial (tidak menambah saldo simpanan)',
                'diajukan_oleh': t.get('diajukan_oleh', ''),
            })
            existing.add(key)
        moved += 1

    if moved:
        tulis_csv(FILE_IURAN_SOSIAL, iuran_sosial, IURAN_SOSIAL_FIELDNAMES)
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, kept_transaksi, SIMPANAN_TRANSAKSI_FIELDNAMES)


def migrate_simpanan_transaksi_legacy_labels():
    """Normalisasi label jenis simpanan lama agar konsisten dengan aturan terbaru."""
    ensure_simpanan_transaksi_schema()
    rows = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    if not rows:
        return

    changed = False
    normalized = []
    for r in rows:
        row = {k: (r.get(k) or '') for k in SIMPANAN_TRANSAKSI_FIELDNAMES}
        jenis_raw = row.get('jenis_simpanan', '')
        jenis_norm = normalize_jenis_simpanan(jenis_raw)
        if jenis_norm != jenis_raw:
            row['jenis_simpanan'] = jenis_norm
            changed = True
        normalized.append(row)

    if changed:
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, normalized, SIMPANAN_TRANSAKSI_FIELDNAMES)


def bunga_dari_tenor(tenor_bulan: int) -> float:
    """Fallback bunga per bulan jika jenis pinjaman tidak ditentukan."""
    if tenor_bulan <= 0:
        return 0.0
    if 1 <= tenor_bulan <= 2:
        return 2.0
    if 3 <= tenor_bulan <= 12:
        return 1.5
    if 13 <= tenor_bulan <= 24:
        return 0.80
    if 36 <= tenor_bulan <= 120:
        return 0.75
    raise ValueError('Tenor valid: 1-2, 3-12, 13-24, atau 36-120 bulan.')


def provisi_persen_dari_tenor(tenor_bulan: int) -> float:
    """Provisi one-time 2% dari pokok jika tenor > 12 bulan."""
    return PROVISI_RATE_LONG_TENOR * 100 if tenor_bulan >= PROVISI_MIN_TENOR_BULAN else 0.0


def provisi_persen_dari_pinjaman(jenis_pinjaman: str, tenor_bulan: int) -> float:
    """Provisi per jenis pinjaman.

    - Solusi Cepat: selalu 2%
    - Modal Usaha: selalu 2%
    - Jangka Panjang: 2% jika tenor >= 13 bulan
    - Jenis lain: 0%
    """
    jenis = (jenis_pinjaman or '').strip().lower()
    if jenis == 'solusi cepat':
        return PROVISI_RATE_LONG_TENOR * 100
    if jenis == 'modal usaha':
        return PROVISI_RATE_LONG_TENOR * 100
    if jenis == 'jangka panjang':
        return provisi_persen_dari_tenor(tenor_bulan)
    return 0.0


def provisi_nominal_pinjaman(jenis_pinjaman: str, plafon: float, tenor_bulan: int) -> float:
    """Hitung provisi sesuai jenis pinjaman dan tenor."""
    plafon = max(float(plafon or 0), 0.0)
    return plafon * (provisi_persen_dari_pinjaman(jenis_pinjaman, tenor_bulan) / 100.0)


def hitung_total_bayar_flat(plafon: float, bunga_persen_bulanan: float, tenor_bulan: int, jenis_pinjaman: str = '') -> tuple:
    """Hitung total bayar skema flat bulanan + provisi tenor panjang."""
    if tenor_bulan <= 0:
        return max(float(plafon or 0), 0.0), 0.0
    plafon = max(float(plafon or 0), 0.0)
    # Bunga/jasa hanya dihitung pada bulan 1 s.d. n-1; bulan terakhir bebas bunga.
    bulan_berbunga = max(tenor_bulan - 1, 0)
    bunga_nominal = plafon * (bunga_persen_bulanan / 100.0) * bulan_berbunga
    provisi_nominal = plafon * (provisi_persen_dari_pinjaman(jenis_pinjaman, tenor_bulan) / 100.0)
    return plafon + bunga_nominal + provisi_nominal, provisi_nominal


def hitung_total_bayar_tanpa_provisi(plafon: float, bunga_persen_bulanan: float, tenor_bulan: int, jenis_pinjaman: str = '') -> float:
    """Hitung total bayar skema flat bulanan TANPA provisi (untuk pengajuan awal)."""
    total_bayar, _ = hitung_total_bayar_flat(plafon, bunga_persen_bulanan, tenor_bulan, jenis_pinjaman)
    return total_bayar


def apply_provisi_setelah_konfirmasi(plafon: float, tenor_bulan: int, jenis_pinjaman: str = '') -> float:
    """Hitung provisi yang diterapkan saat admin mengkonfirmasi pinjaman."""
    provisi_nominal = plafon * (provisi_persen_dari_pinjaman(jenis_pinjaman, tenor_bulan) / 100.0)
    return provisi_nominal


def hitung_cicilan_bulanan(plafon: float, bunga_persen_bulanan: float, tenor_bulan: int) -> float:
    """
    Hitung cicilan bulanan per formula: cicilan = (p / b) + (p × j%)
    
    Keterangan:
    - p = plafon (pokok pinjaman)
    - b = tenor (bulan)
    - j = bunga_persen_bulanan (% per bulan)
    
    Contoh: p=15.000.000, b=10, j=1.5%
    cicilan = (15.000.000 / 10) + (15.000.000 × 1.5%) = 1.500.000 + 225.000 = 1.725.000
    """
    if tenor_bulan <= 0:
        return 0.0
    plafon = max(float(plafon or 0), 0.0)
    bunga_persen_bulanan = max(float(bunga_persen_bulanan or 0), 0.0)
    
    # cicilan = (p / b) + (p × j%)
    cicilan_pokok = plafon / tenor_bulan
    total_bunga = plafon * (bunga_persen_bulanan / 100.0)
    cicilan = cicilan_pokok + total_bunga
    
    return cicilan


def nominal_cicilan_aktual(pinjaman: dict) -> float:
    """Nominal yang benar-benar dibayarkan pada transaksi cicilan berjalan."""
    sisa = saldo_pinjaman_aktual(pinjaman)
    jenis = (pinjaman.get('jenis_pinjaman') or '').strip()
    # Solusi Cepat dibayar sekali melalui admin (bulan ke-1 atau ke-2).
    if jenis == 'Solusi Cepat':
        return max(sisa, 0.0)
    cic = float(pinjaman.get('cicilan_per_bulan') or 0) or cicilan_per_bulan_saldo(pinjaman)
    if sisa <= 0 or cic <= 0:
        return 0.0
    return min(sisa, cic)


def saldo_pinjaman_aktual(pinjaman: dict) -> float:
    """Saldo pinjaman yang dipakai UI dan proses bayar.

    Jika data lama masih menyimpan sisa 0, fallback ke plafon (pokok pinjaman).
    """
    try:
        sisa = max(float(pinjaman.get('sisa_pinjaman') or 0), 0.0)
    except (TypeError, ValueError):
        sisa = 0.0
    if sisa > 0:
        return sisa
    status = (pinjaman.get('status') or '').strip()
    if status in ('Lunas', 'Ditolak'):
        return 0.0
    try:
        total = max(float(pinjaman.get('plafon') or pinjaman.get('total_pinjaman') or 0), 0.0)
    except (TypeError, ValueError):
        total = 0.0
    return total


def _parse_tanggal_iso(value: str):
    value = (value or '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _tambah_bulan_ke_tanggal(tanggal_awal, jumlah_bulan: int):
    if not tanggal_awal:
        return None
    jumlah_bulan = max(int(jumlah_bulan or 0), 0)
    if jumlah_bulan <= 0:
        return tanggal_awal
    total_bulan = tanggal_awal.month - 1 + jumlah_bulan
    year = tanggal_awal.year + (total_bulan // 12)
    month = (total_bulan % 12) + 1
    day = min(tanggal_awal.day, calendar.monthrange(year, month)[1])
    return tanggal_awal.replace(year=year, month=month, day=day)


def _geser_bulan_tanggal(tanggal_awal, delta_bulan: int):
    """Geser tanggal maju/mundur sejumlah bulan dengan penyesuaian akhir bulan."""
    if not tanggal_awal:
        return None
    total_bulan = (tanggal_awal.year * 12 + (tanggal_awal.month - 1)) + int(delta_bulan or 0)
    year = total_bulan // 12
    month = (total_bulan % 12) + 1
    day = min(tanggal_awal.day, calendar.monthrange(year, month)[1])
    return tanggal_awal.replace(year=year, month=month, day=day)


def tanggal_jatuh_tempo_pinjaman(pinjaman: dict):
    tanggal_pengajuan = _parse_tanggal_iso(pinjaman.get('tanggal_pengajuan'))
    tenor = int(float(pinjaman.get('tenor_awal') or pinjaman.get('tenor_bulan') or pinjaman.get('tenor') or 0))
    return _tambah_bulan_ke_tanggal(tanggal_pengajuan, tenor)


def pinjaman_telat_bayar_aktual(pinjaman: dict) -> bool:
    if (pinjaman.get('status') or '').strip() != 'Disetujui':
        return False
    if saldo_pinjaman_aktual(pinjaman) <= 0:
        return False
    jatuh_tempo = tanggal_jatuh_tempo_pinjaman(pinjaman)
    if not jatuh_tempo:
        return False
    return datetime.now().date() > jatuh_tempo


def ada_pinjaman_solusi_cepat_aktif(pinjaman_rows: list, id_anggota: str) -> bool:
    """Cek apakah anggota masih punya pinjaman Solusi Cepat yang belum lunas."""
    for p in pinjaman_rows:
        if p.get('id_anggota') != id_anggota:
            continue
        if (p.get('jenis_pinjaman') or '').strip() != 'Solusi Cepat':
            continue
        status = (p.get('status') or '').strip()
        if status in ('Menunggu', 'Disetujui'):
            return True
        if saldo_pinjaman_aktual(p) > 0 and status != 'Ditolak':
            return True
    return False


migrate_simpanan_ke_saldo()
migrate_pinjaman_ke_saldo()
migrate_sisa_pinjaman_aktif_ke_total()
ensure_pinjaman_cicilan_schema()
ensure_anggota_schema()
ensure_import_log_schema()
ensure_import_preview_dir()
ensure_simpanan_transaksi_schema()
ensure_simpanan_pengajuan_schema()
ensure_iuran_sosial_schema()
migrate_simpanan_transaksi_legacy_labels()
migrate_iuran_sosial_legacy()


def _db_insert_import_log(berhasil: int, gagal: int, mode: str, nama_file: str, catatan: str) -> None:
    """Append 1 baris import_log ke PostgreSQL (tanpa bergantung file Excel)."""
    stmt = text(
        """
        INSERT INTO import_log (waktu, "user", mode, nama_file, berhasil, gagal, catatan)
        VALUES (:waktu, :user, :mode, :nama_file, :berhasil, :gagal, :catatan)
        """
    )
    payload = {
        "waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": (session.get("user") or ""),
        "mode": (mode or "").strip(),
        "nama_file": (nama_file or "").strip(),
        "berhasil": int(berhasil or 0),
        "gagal": int(gagal or 0),
        "catatan": (catatan or "")[:2000],
    }
    with db_session() as conn:
        conn.execute(stmt, payload)


def _db_select_import_log_latest(limit: int = 50) -> list:
    """Ambil import_log terbaru (DB mode)."""
    stmt = text(
        """
        SELECT waktu, "user", mode, nama_file, berhasil, gagal, catatan
        FROM import_log
        ORDER BY waktu DESC
        LIMIT :limit
        """
    )
    try:
        with db_session() as conn:
            rows = conn.execute(stmt, {"limit": int(limit or 50)}).mappings().all()
        out = []
        for r in rows:
            out.append({
                "waktu": "" if r.get("waktu") is None else str(r.get("waktu")),
                "user": "" if r.get("user") is None else str(r.get("user")),
                "mode": "" if r.get("mode") is None else str(r.get("mode")),
                "nama_file": "" if r.get("nama_file") is None else str(r.get("nama_file")),
                "berhasil": "0" if r.get("berhasil") is None else str(r.get("berhasil")),
                "gagal": "0" if r.get("gagal") is None else str(r.get("gagal")),
                "catatan": "" if r.get("catatan") is None else str(r.get("catatan")),
            })
        return out
    except SQLAlchemyError as e:
        print(f"Error membaca import_log dari DB: {e}")
        return []


def hitung_kapasitas_cicilan(penghasilan_bersih: float, cicilan_lain: float, dsr: float) -> float:
    return max((penghasilan_bersih * dsr) - cicilan_lain, 0.0)


def dsr_otomatis_dari_penghasilan(penghasilan_bersih: float) -> float:
    """DSR 30–40% mengikuti penghasilan bersih (otomatis, tanpa pilihan manual)."""
    if penghasilan_bersih <= 0:
        return DSR_DEFAULT
    if penghasilan_bersih <= 5_000_000:
        return 0.40
    if penghasilan_bersih <= 15_000_000:
        return 0.35
    return 0.30


def hitung_plafon_maks_anuitas(cicilan_bulanan: float, bunga_persen_bulanan: float, tenor_bulan: int, jenis_pinjaman: str = '') -> float:
    n = int(tenor_bulan or 0)
    if n <= 0:
        raise ValueError('Tenor harus lebih dari 0.')
    cicilan_bulanan = max(float(cicilan_bulanan or 0), 0.0)
    i = max(float(bunga_persen_bulanan or 0), 0.0) / 100.0
    provisi_rate = provisi_persen_dari_pinjaman(jenis_pinjaman, n) / 100.0
    faktor = (1 + (i * n) + provisi_rate)
    if faktor <= 0:
        return 0.0
    return (cicilan_bulanan * n) / faktor


def info_pinjaman_dsr_anggota(anggota_row: dict) -> dict:
    """Ringkasan DSR otomatis dan kapasitas cicilan untuk satu baris anggota (tampilan admin)."""
    ph = max(float(anggota_row.get('penghasilan_bersih') or 0), 0.0)
    cl = max(float(anggota_row.get('cicilan_lain') or 0), 0.0)
    dsr = dsr_otomatis_dari_penghasilan(ph)
    kap = hitung_kapasitas_cicilan(ph, cl, dsr)
    return {
        'dsr_persen': int(round(dsr * 100)),
        'kapasitas_cicilan': kap,
    }


def parse_rupiah_to_float(value):
    return float(str(value or '0').replace(',', '').replace('.', ''))


def normalize_nik(value: str) -> str:
    """Normalisasi NIK: hanya digit, tanpa spasi/simbol."""
    return re.sub(r'\D', '', str(value or '').strip())


def is_valid_nik(value: str) -> bool:
    """Valid jika NIK tepat 16 digit angka."""
    return bool(re.fullmatch(r'\d{16}', str(value or '')))


def get_user_by_username(username: str):
    users = baca_csv(FILE_USERS)
    for u in users:
        if u.get('username', '').lower() == (username or '').lower():
            return u
    return None


def get_user_by_id(id_user: str):
    users = baca_csv(FILE_USERS)
    for u in users:
        if u.get('id_user') == id_user:
            return u
    return None


def get_current_user_id_anggota():
    return session.get('id_anggota') or ''


def is_current_user_admin():
    return (session.get('role') or '').strip().lower() in ADMIN_PANEL_ROLES


SIMPANAN_RULES = {
    'Simpanan Pokok': {
        'fixed': 500_000,
        'min': 500_000,
        'max': 500_000,
        'one_time': True,
        'monthly': False,
        'akumulasi': True,
    },
    'Simpanan Wajib': {
        'fixed': 250_000,
        'min': 250_000,
        'max': 250_000,
        'one_time': False,
        'monthly': True,
        'akumulasi': True,
    },
    'Simpanan Hari Koperasi': {
        'fixed': 20_000,
        'min': 20_000,
        'max': 20_000,
        'one_time': False,
        'monthly': True,
        'akumulasi': True,
    },
    'Iuran Sosial': {
        'fixed': 10_000,
        'min': 10_000,
        'max': 10_000,
        'one_time': False,
        'monthly': True,
        'akumulasi': False,
    },
    'Simpanan Pensiun': {
        'fixed': None,
        'min': 50_000,
        'max': None,
        'one_time': False,
        'monthly': False,
        'akumulasi': True,
    },
    'Simpanan Hari Raya': {
        'fixed': None,
        'min': 50_000,
        'max': None,
        'one_time': False,
        'monthly': False,
        'akumulasi': True,
    },
    'Simpanan Manasuka': {
        'fixed': None,
        'min': 50_000,
        'max': None,
        'one_time': False,
        'monthly': False,
        'akumulasi': True,
    },
    'Simpanan Pendidikan': {
        'fixed': None,
        'min': 50_000,
        'max': None,
        'one_time': False,
        'monthly': False,
        'akumulasi': True,
    },
}
SIMPANAN_JENIS_ALIAS = {
    'Simpan Pokok': 'Simpanan Pokok',
    'Simpan Wajib': 'Simpanan Wajib',
    'Hari Koperasi': 'Simpanan Hari Koperasi',
    'Simpan Hari Koperasi': 'Simpanan Hari Koperasi',
    'Pensiun': 'Simpanan Pensiun',
    'Simpan Pensiun': 'Simpanan Pensiun',
    'Hari Raya': 'Simpanan Hari Raya',
    'Manasuka': 'Simpanan Manasuka',
    'Pendidikan': 'Simpanan Pendidikan',
}
SIMPANAN_CHOICES = list(SIMPANAN_RULES.keys())
SIMPANAN_AKUMULASI_CHOICES = [k for k, v in SIMPANAN_RULES.items() if v.get('akumulasi')]
SIMPANAN_DEFAULT_AKUMULASI = 'Simpanan Wajib'


def normalize_jenis_simpanan(value: str) -> str:
    jenis = (value or '').strip()
    if jenis in SIMPANAN_RULES:
        return jenis
    return SIMPANAN_JENIS_ALIAS.get(jenis, jenis)


def _saldo_per_jenis_akumulasi(simpanan_transaksi: list) -> dict:
    saldo_per_jenis_map = {}
    for t in simpanan_transaksi:
        id_a = (t.get('id_anggota') or '').strip()
        if not id_a:
            continue
        jenis = normalize_jenis_simpanan(t.get('jenis_simpanan'))
        if jenis not in SIMPANAN_AKUMULASI_CHOICES:
            continue
        try:
            nominal = float(t.get('jumlah') or 0)
        except (TypeError, ValueError):
            nominal = 0.0
        if id_a not in saldo_per_jenis_map:
            saldo_per_jenis_map[id_a] = {k: 0.0 for k in SIMPANAN_AKUMULASI_CHOICES}
        saldo_per_jenis_map[id_a][jenis] += nominal
    return saldo_per_jenis_map


def _jenis_sudah_tercatat_anggota(simpanan_transaksi: list, id_anggota: str, jenis_simpanan: str) -> bool:
    for t in simpanan_transaksi:
        if t.get('id_anggota') != id_anggota:
            continue
        if normalize_jenis_simpanan(t.get('jenis_simpanan')) == jenis_simpanan:
            return True
    return False


def _jenis_sudah_dibayar_bulan_ini(
    simpanan_transaksi: list,
    id_anggota: str,
    jenis_simpanan: str,
    month_key: str,
    jenis_default: str = '',
) -> bool:
    for t in simpanan_transaksi:
        if t.get('id_anggota') != id_anggota:
            continue
        jenis_tercatat = t.get('jenis_simpanan') or jenis_default
        if normalize_jenis_simpanan(jenis_tercatat) != jenis_simpanan:
            continue
        tanggal = (t.get('tanggal') or '').strip()
        if tanggal[:7] == month_key:
            return True
    return False


def _jenis_simpanan_menunggu_pengajuan(
    pengajuan_simpanan: list,
    id_anggota: str,
    jenis_simpanan: str,
    month_key: str = '',
    jenis_default: str = '',
) -> bool:
    for p in pengajuan_simpanan:
        if p.get('id_anggota') != id_anggota:
            continue
        if (p.get('status') or '').strip() != 'Menunggu':
            continue
        jenis_tercatat = p.get('jenis_simpanan') or jenis_default
        if normalize_jenis_simpanan(jenis_tercatat) != jenis_simpanan:
            continue
        if month_key:
            tanggal = (p.get('tanggal_pengajuan') or '').strip()
            if tanggal[:7] != month_key:
                continue
        return True
    return False


def _riwayat_jenis_pinjaman_per_anggota(pinjaman_rows: list) -> dict:
    """Kumpulkan daftar jenis pinjaman per anggota untuk kebutuhan export."""
    out = {}
    for p in pinjaman_rows:
        id_a = (p.get('id_anggota') or '').strip()
        if not id_a:
            continue
        jenis_raw = (p.get('jenis_pinjaman') or '').strip()
        if not jenis_raw:
            try:
                tenor_norm = int(float(p.get('tenor_bulan') or p.get('tenor') or 0))
            except (TypeError, ValueError):
                tenor_norm = 0
            jenis_raw = kategori_pinjaman_dari_tenor(tenor_norm) if tenor_norm > 0 else ''
        if not jenis_raw:
            continue
        bucket = out.setdefault(id_a, [])
        if jenis_raw not in bucket:
            bucket.append(jenis_raw)
    return {k: ', '.join(v) for k, v in out.items()}


def _riwayat_jenis_simpanan_per_anggota(simpanan_transaksi_rows: list, iuran_sosial_rows: list) -> dict:
    """Kumpulkan daftar jenis simpanan/iuran per anggota untuk kebutuhan export."""
    out = {}
    for t in simpanan_transaksi_rows:
        id_a = (t.get('id_anggota') or '').strip()
        if not id_a:
            continue
        jenis = normalize_jenis_simpanan(t.get('jenis_simpanan'))
        if not jenis:
            continue
        bucket = out.setdefault(id_a, [])
        if jenis not in bucket:
            bucket.append(jenis)

    for i in iuran_sosial_rows:
        id_a = (i.get('id_anggota') or '').strip()
        if not id_a:
            continue
        bucket = out.setdefault(id_a, [])
        if 'Iuran Sosial' not in bucket:
            bucket.append('Iuran Sosial')

    return {k: ', '.join(v) for k, v in out.items()}


def ensure_users_schema():
    """Pastikan users.csv punya kolom id_anggota (upgrade schema lama)."""
    users = baca_csv(FILE_USERS)
    if not users:
        return
    if 'id_anggota' in users[0]:
        return
    for u in users:
        u['id_anggota'] = ''
    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])


ROLE_LABELS = {
    'super_admin': 'Super Admin',
    'admin_koperasi': 'Admin Koperasi',
    'bendahara': 'Bendahara',
    'ketua_pengurus': 'Ketua Pengurus',
    'anggota': 'Anggota',
    'auditor': 'Auditor',
    'admin': 'Admin',
    'user': 'User',
}

STAFF_ROLES = {'admin', 'super_admin', 'admin_koperasi', 'bendahara', 'ketua_pengurus', 'auditor'}
ADMIN_PANEL_ROLES = {'admin', 'super_admin', 'admin_koperasi', 'bendahara', 'ketua_pengurus', 'auditor'}
MEMBER_ROLES = {'user', 'anggota'}
ROLE_OPTIONS = [
    ('super_admin', 'Super Admin'),
    ('admin_koperasi', 'Admin Koperasi'),
    ('bendahara', 'Bendahara'),
    ('ketua_pengurus', 'Ketua Pengurus'),
    ('anggota', 'Anggota'),
    ('auditor', 'Auditor'),
    ('admin', 'Admin (legacy)'),
    ('user', 'User (legacy)'),
]

DEFAULT_USER_SEEDS = [
    {'username': 'superadmin', 'password': 'Admin@123', 'role': 'super_admin', 'id_anggota': ''},
    {'username': 'adminkoperasi', 'password': 'Admin@123', 'role': 'admin_koperasi', 'id_anggota': ''},
    {'username': 'bendahara', 'password': 'Bendahara@123', 'role': 'bendahara', 'id_anggota': ''},
    {'username': 'ketuapengurus', 'password': 'Ketua@123', 'role': 'ketua_pengurus', 'id_anggota': ''},
    {'username': 'anggota_demo', 'password': 'Anggota@123', 'role': 'anggota', 'id_anggota': ''},
    {'username': 'auditor', 'password': 'Auditor@123', 'role': 'auditor', 'id_anggota': ''},
]


def _seed_default_user_rows(users: list[dict]) -> tuple[list[dict], int]:
    existing_usernames = {str(u.get('username') or '').strip().lower() for u in users if isinstance(u, dict)}
    added = 0
    for seed in DEFAULT_USER_SEEDS:
        username = seed['username'].strip().lower()
        if username in existing_usernames:
            continue
        users.append({
            'id_user': str(uuid.uuid4()),
            'username': seed['username'],
            'password_hash': generate_password_hash(seed['password']),
            'role': seed['role'],
            'id_anggota': seed['id_anggota'],
            'created_at': datetime.now().strftime('%Y-%m-%d'),
        })
        existing_usernames.add(username)
        added += 1
    return users, added


def _ensure_role_linked_demo_data() -> None:
    """Buat data demo lintas modul agar role bisa diuji end-to-end."""
    today = datetime.now().strftime('%Y-%m-%d')
    anggota_rows = baca_csv(FILE_ANGGOTA)
    users = baca_csv(FILE_USERS)
    simpanan_rows = baca_csv(FILE_SIMPANAN)
    simpanan_tx_rows = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    simpanan_pengajuan_rows = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    cicilan_rows = baca_csv(FILE_PINJAMAN_CICILAN)

    demo_member_id = 'AGT-DEMO-001'
    demo_member_no = 'KOP-0001'
    demo_member_name = 'Anggota Demo'

    if not any((r.get('id_anggota') or '').strip() == demo_member_id for r in anggota_rows):
        anggota_rows.append({
            'id_anggota': demo_member_id,
            'no_anggota': demo_member_no,
            'nik': '3273000000000001',
            'nama': demo_member_name,
            'alamat': 'Bandung',
            'no_telp': '081200000001',
            'tgl_bergabung': today,
            'no_rekening': '000111222333',
            'nama_bank': 'BCA',
            'penghasilan_bersih': '6500000',
            'cicilan_lain': '500000',
            'simpanan_pokok': '250000',
        })
        tulis_csv(FILE_ANGGOTA, anggota_rows, ANGGOTA_FIELDNAMES)

    changed_users = False
    for user in users:
        if (user.get('username') or '').strip().lower() == 'anggota_demo':
            if (user.get('id_anggota') or '').strip() != demo_member_id:
                user['id_anggota'] = demo_member_id
                changed_users = True
            if not (user.get('role') or '').strip():
                user['role'] = 'anggota'
                changed_users = True
            break
    if changed_users:
        tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])

    if not any((r.get('id_anggota') or '').strip() == demo_member_id for r in simpanan_rows):
        simpanan_rows.append({'id_anggota': demo_member_id, 'total_simpanan': '1750000'})
        tulis_csv(FILE_SIMPANAN, simpanan_rows, SIMPANAN_FIELDNAMES)

    if not any((r.get('id_anggota') or '').strip() == demo_member_id for r in simpanan_tx_rows):
        simpanan_tx_rows.append({
            'id_transaksi': str(uuid.uuid4()),
            'id_anggota': demo_member_id,
            'no_anggota': demo_member_no,
            'nama_anggota': demo_member_name,
            'tanggal': today,
            'jenis_simpanan': 'Manasuka',
            'jumlah': '500000',
            'keterangan': 'Setoran awal demo',
            'diajukan_oleh': 'admin_koperasi',
        })
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, simpanan_tx_rows, SIMPANAN_TRANSAKSI_FIELDNAMES)

    if not any((r.get('id_anggota') or '').strip() == demo_member_id for r in simpanan_pengajuan_rows):
        simpanan_pengajuan_rows.append({
            'id_pengajuan': str(uuid.uuid4()),
            'id_anggota': demo_member_id,
            'no_anggota': demo_member_no,
            'nama_anggota': demo_member_name,
            'tanggal_pengajuan': today,
            'jenis_simpanan': 'Manasuka',
            'jumlah': '200000',
            'keterangan': 'Pengajuan setoran dari anggota',
            'status': 'Menunggu',
            'tanggal_konfirmasi': '',
            'dikonfirmasi_oleh': '',
            'diajukan_oleh': 'anggota_demo',
        })
        tulis_csv(FILE_SIMPANAN_PENGAJUAN, simpanan_pengajuan_rows, SIMPANAN_PENGAJUAN_FIELDNAMES)

    demo_loan = next((r for r in pinjaman_rows if (r.get('id_anggota') or '').strip() == demo_member_id), None)
    if demo_loan is None:
        loan_id = str(uuid.uuid4())
        pinjaman_rows.append({
            'id_pinjaman': loan_id,
            'id_anggota': demo_member_id,
            'nama_anggota': demo_member_name,
            'no_anggota': demo_member_no,
            'jenis_pinjaman': 'Jangka Pendek',
            'jenis_simpanan': 'Manasuka',
            'plafon': '3000000',
            'tenor_awal': '12',
            'tenor_bulan': '12',
            'bunga_persen': '1.5',
            'total_bayar': '3540000',
            'cicilan_per_bulan': '295000',
            'sisa_pinjaman': '3540000',
            'tanggal_pengajuan': today,
            'status': 'Menunggu Persetujuan',
            'tanggal_lunas': '',
        })
        tulis_csv(FILE_PINJAMAN, pinjaman_rows, PINJAMAN_FIELDNAMES)
    else:
        loan_id = demo_loan.get('id_pinjaman') or ''

    if loan_id and not any((r.get('id_pinjaman') or '').strip() == loan_id for r in cicilan_rows):
        cicilan_rows.append({
            'id_cicilan': str(uuid.uuid4()),
            'id_pinjaman': loan_id,
            'id_anggota': demo_member_id,
            'no_anggota': demo_member_no,
            'nama_anggota': demo_member_name,
            'jumlah': '295000',
            'tanggal_pengajuan': today,
            'status': 'Menunggu',
            'tanggal_konfirmasi': '',
            'dikonfirmasi_oleh': '',
            'diajukan_oleh': 'anggota_demo',
            'keterangan': 'Upload bukti cicilan pertama',
            'metode_pembayaran': 'Transfer Bank',
            'detail_pembayaran': '',
            'status_transaksi': 'pending',
            'va_number': '',
            'idempotency_key': '',
            'periode_tagihan': f'{datetime.now().year}-{datetime.now().month:02d}',
            'expires_at': '',
        })
        tulis_csv(FILE_PINJAMAN_CICILAN, cicilan_rows, CICILAN_FIELDNAMES)


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        if (session.get('role') or '').strip().lower() not in ADMIN_PANEL_ROLES:
            abort(403)
        return view_func(*args, **kwargs)
    return wrapper


ROLE_PERMISSIONS = {
    'admin': {
        'members.view',
        'members.manage',
        'savings.deposit.request',
        'savings.deposit.input',
        'savings.deposit.validate',
        'savings.withdraw.request',
        'savings.withdraw.validate',
        'loan.documents.review',
        'loan.disbursement.input',
        'installments.manage',
        'reports.export',
        'reports.strategic.view',
        'shu.view',
        'excel.import',
        'news.manage',
    },
    'super_admin': {
        'members.view',
        'members.manage',
        'members.approve',
        'savings.deposit.request',
        'savings.deposit.input',
        'savings.deposit.validate',
        'savings.withdraw.request',
        'savings.withdraw.validate',
        'loan.documents.review',
        'loan.eligibility.analyze',
        'loans.approve',
        'loan.disbursement.input',
        'installments.manage',
        'cash.manage',
        'shu.manage',
        'shu.validate',
        'shu.view',
        'reports.export',
        'reports.strategic.view',
        'backup.manage',
        'excel.import',
        'news.manage',
        'users.manage',
        'roles.manage',
        'audit.view',
        'system.manage',
    },
    'admin_koperasi': {
        'members.view',
        'members.manage',
        'savings.deposit.request',
        'savings.deposit.input',
        'savings.deposit.validate',
        'savings.withdraw.request',
        'savings.withdraw.validate',
        'loan.documents.review',
        'loan.disbursement.input',
        'installments.manage',
        'reports.export',
        'reports.strategic.view',
        'shu.view',
        'excel.import',
        'news.manage',
    },
    'bendahara': {
        'members.view',
        'members.manage.limited',
        'savings.deposit.request',
        'savings.deposit.input',
        'savings.deposit.validate',
        'savings.withdraw.request',
        'savings.withdraw.validate',
        'loan.eligibility.analyze',
        'loan.disbursement.input',
        'installments.manage',
        'cash.manage',
        'shu.manage',
        'shu.view',
        'reports.export',
        'reports.strategic.view',
        'excel.import',
    },
    'ketua_pengurus': {
        'members.view',
        'loans.approve',
        'reports.export',
        'reports.strategic.view',
        'shu.validate',
        'shu.view',
    },
    'anggota': {
        'members.self.view',
        'members.self.edit.limited',
        'savings.deposit.request',
        'savings.withdraw.request',
        'loan.request',
        'installments.proof.upload',
        'shu.self.view',
        'reports.self.view',
    },
    'auditor': {
        'members.view',
        'shu.view',
        'audit.view',
        'reports.export',
        'reports.strategic.view',
    },
    'user': {
        'members.self.view',
        'members.self.edit.limited',
        'savings.deposit.request',
        'savings.withdraw.request',
        'loan.request',
        'installments.proof.upload',
        'shu.self.view',
        'reports.self.view',
    },
}

# Cache untuk role/permission dari DB (Tahap 2: DB-backed RBAC)
_ROLE_PERMISSIONS_CACHE = {}
_ROLE_PERMISSIONS_CACHE_TIMESTAMP = 0


def _get_role_permissions_from_db(role_name: str) -> set:
    """Baca permission untuk role dari DB role_permissions table (Tahap 2 feature)."""
    global _ROLE_PERMISSIONS_CACHE, _ROLE_PERMISSIONS_CACHE_TIMESTAMP
    import time
    
    # Cache timeout: 5 menit
    now = time.time()
    if now - _ROLE_PERMISSIONS_CACHE_TIMESTAMP > 300:
        _ROLE_PERMISSIONS_CACHE.clear()
        _ROLE_PERMISSIONS_CACHE_TIMESTAMP = now
    
    if role_name in _ROLE_PERMISSIONS_CACHE:
        return _ROLE_PERMISSIONS_CACHE[role_name]
    
    perms = set()
    if not _is_db_mode_enabled():
        return perms
    
    try:
        with db_session() as conn:
            result = conn.execute(
                text("""
                    SELECT p.permission_name
                    FROM role_permissions rp
                    JOIN roles r ON rp.id_role = r.id_role
                    JOIN permissions p ON rp.id_permission = p.id_permission
                    WHERE LOWER(r.role_name) = LOWER(:role)
                """),
                {'role': role_name}
            )
            perms = {row[0] for row in result}
    except Exception as e:
        print(f"Error membaca role_permissions dari DB untuk role '{role_name}': {e}")
    
    _ROLE_PERMISSIONS_CACHE[role_name] = perms
    return perms


def _current_role() -> str:
    return (session.get('role') or '').strip().lower()


def has_permission(permission: str, role: str | None = None) -> bool:
    """
    Cek apakah role memiliki permission.
    
    Strategi:
    1. Jika DB mode ON, baca dari DB role_permissions (Tahap 2)
    2. Jika DB mode OFF atau permission tidak ditemukan di DB, fallback ke ROLE_PERMISSIONS file-based
    3. Ini memastikan backward compatibility dan smooth transition ke DB-backed RBAC
    """
    role_name = (role or _current_role()).strip().lower()
    
    # Jika DB mode aktif, coba baca dari DB
    if _is_db_mode_enabled():
        db_perms = _get_role_permissions_from_db(role_name)
        if db_perms:
            return permission in db_perms
    
    # Fallback ke file-based ROLE_PERMISSIONS (legacy mode)
    return permission in ROLE_PERMISSIONS.get(role_name, set())


def _invalidate_role_permissions_cache():
    """Clear cache role_permissions dari DB (dipanggil saat ada perubahan role/permission)."""
    global _ROLE_PERMISSIONS_CACHE, _ROLE_PERMISSIONS_CACHE_TIMESTAMP
    _ROLE_PERMISSIONS_CACHE.clear()
    _ROLE_PERMISSIONS_CACHE_TIMESTAMP = 0


def get_all_roles_from_db() -> list:
    """Baca semua roles dari DB (Tahap 2 feature)."""
    if not _is_db_mode_enabled():
        return []
    try:
        with db_session() as conn:
            result = conn.execute(text("""
                SELECT id_role, role_name, deskripsi, created_at 
                FROM roles 
                ORDER BY role_name
            """))
            return [
                {
                    'id_role': row[0],
                    'role_name': row[1],
                    'deskripsi': row[2],
                    'created_at': row[3],
                }
                for row in result
            ]
    except Exception as e:
        print(f"Error membaca roles dari DB: {e}")
        return []


def get_role_by_name_from_db(role_name: str) -> dict:
    """Baca role by name dari DB."""
    if not _is_db_mode_enabled():
        return {}
    try:
        with db_session() as conn:
            result = conn.execute(
                text("""
                    SELECT id_role, role_name, deskripsi, created_at 
                    FROM roles 
                    WHERE LOWER(role_name) = LOWER(:role)
                """),
                {'role': role_name}
            )
            row = result.fetchone()
            if row:
                return {
                    'id_role': row[0],
                    'role_name': row[1],
                    'deskripsi': row[2],
                    'created_at': row[3],
                }
    except Exception as e:
        print(f"Error membaca role '{role_name}' dari DB: {e}")
    return {}


def get_permissions_for_role_from_db(role_name: str) -> list:
    """Baca list permissions untuk role dari DB (detailed, bukan set)."""
    if not _is_db_mode_enabled():
        return []
    try:
        with db_session() as conn:
            result = conn.execute(
                text("""
                    SELECT p.id_permission, p.permission_name
                    FROM role_permissions rp
                    JOIN roles r ON rp.id_role = r.id_role
                    JOIN permissions p ON rp.id_permission = p.id_permission
                    WHERE LOWER(r.role_name) = LOWER(:role)
                    ORDER BY p.permission_name
                """),
                {'role': role_name}
            )
            return [{'id_permission': row[0], 'permission_name': row[1]} for row in result]
    except Exception as e:
        print(f"Error membaca permissions untuk role '{role_name}': {e}")
        return []


def get_all_permissions_from_db() -> list:
    """Baca semua permissions dari DB."""
    if not _is_db_mode_enabled():
        return []
    try:
        with db_session() as conn:
            result = conn.execute(text("""
                SELECT id_permission, permission_name
                FROM permissions
                ORDER BY permission_name
            """))
            return [{'id_permission': row[0], 'permission_name': row[1]} for row in result]
    except Exception as e:
        print(f"Error membaca permissions dari DB: {e}")
        return []


def permission_required(*permissions: str):
    required = [perm.strip() for perm in permissions if perm and perm.strip()]

    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapper(*args, **kwargs):
            role_name = _current_role()
            if required and not any(has_permission(permission, role_name) for permission in required):
                abort(403)
            return view_func(*args, **kwargs)

        return wrapper

    return decorator


def restrict_id_anggota_or_forbid(id_anggota: str):
    """User biasa hanya boleh akses data id_anggota dirinya. Admin boleh semua."""
    if is_current_user_admin():
        return
    current = get_current_user_id_anggota()
    if not current or current != id_anggota:
        abort(403)


@app.before_request
def enforce_session_timeout():
    if not session.get('user'):
        return None

    endpoint = request.endpoint or ''
    if endpoint.startswith('static'):
        return None

    now_ts = datetime.utcnow().timestamp()
    last_ts = float(session.get('_last_activity_ts') or 0)
    role = (session.get('role') or 'user').strip().lower()
    timeout_minutes = ADMIN_SESSION_TIMEOUT_MINUTES if role in STAFF_ROLES else USER_SESSION_TIMEOUT_MINUTES

    if last_ts and (now_ts - last_ts) > (timeout_minutes * 60):
        session.clear()
        flash('Sesi berakhir karena tidak ada aktivitas. Silakan login ulang.', 'warning')
        return redirect(url_for('login', next=request.path))

    session['_last_activity_ts'] = now_ts
    session.permanent = True
    _maybe_run_daily_backup()
    return None


def _backup_legacy_data() -> None:
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    target_dir = os.path.join(BACKUP_DIR, stamp)
    os.makedirs(target_dir, exist_ok=True)
    source_files = [
        FILE_ANGGOTA,
        FILE_SIMPANAN,
        FILE_SIMPANAN_TRANSAKSI,
        FILE_SIMPANAN_PENGAJUAN,
        FILE_IURAN_SOSIAL,
        FILE_PINJAMAN,
        FILE_PINJAMAN_CICILAN,
        FILE_USERS,
        FILE_PENDAFTARAN_ANGGOTA,
        FILE_IMPORT_LOG,
        FILE_BERITA,
    ]
    copied = []
    for source in source_files:
        if not os.path.exists(source):
            continue
        destination = os.path.join(target_dir, os.path.basename(source))
        shutil.copy2(source, destination)
        copied.append(os.path.basename(source))
    with open(os.path.join(target_dir, 'manifest.json'), 'w', encoding='utf-8') as fh:
        json.dump({'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'files': copied}, fh, ensure_ascii=False, indent=2)
    return stamp


def _maybe_run_daily_backup() -> None:
    global _LAST_AUTO_BACKUP_DATE
    today = datetime.now().date()
    if _LAST_AUTO_BACKUP_DATE == today:
        return
    try:
        stamp = _backup_legacy_data()
        _LAST_AUTO_BACKUP_DATE = today
        try:
            _record_auto_backup_if_weekly(stamp)
        except Exception:
            # non-fatal: logging of auto-backup should not break request
            pass
    except Exception as exc:
        print(f'Error auto backup data: {exc}')


def _backup_log_path() -> str:
    os.makedirs(BACKUP_DIR, exist_ok=True)
    return os.path.join(BACKUP_DIR, 'backup_log.json')


def _load_backup_log() -> list:
    path = _backup_log_path()
    if not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            data = json.load(fh)
            if isinstance(data, list):
                return data
    except Exception:
        pass
    return []


def _save_backup_log(entries: list) -> None:
    path = _backup_log_path()
    try:
        with open(path, 'w', encoding='utf-8') as fh:
            json.dump(entries, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _record_auto_backup_if_weekly(stamp: str) -> None:
    """Record an automatic backup log only once per 7 days; remove previous auto entries."""
    entries = _load_backup_log()
    now = datetime.now()
    last_auto = None
    for e in reversed(entries):
        if e.get('type') == 'auto':
            last_auto = e
            break
    do_append = False
    if not last_auto:
        do_append = True
    else:
        try:
            last_dt = datetime.strptime(last_auto.get('created_at', '') or '', '%Y-%m-%d %H:%M:%S')
            if (now - last_dt).days >= 7:
                do_append = True
        except Exception:
            do_append = True

    if do_append:
        # remove previous auto entries
        entries = [e for e in entries if e.get('type') != 'auto']
        entries.append({'type': 'auto', 'stamp': stamp, 'created_at': now.strftime('%Y-%m-%d %H:%M:%S')})
        _save_backup_log(entries)


@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, private'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'no-referrer'
    response.headers['X-Robots-Tag'] = 'noindex, nofollow, noarchive, nosnippet'
    return response


def _build_admin_pengajuan_notifications(limit: int = 8):
    """Kumpulkan notifikasi pengajuan pinjaman & simpanan yang masih menunggu untuk admin."""
    if not is_current_user_admin():
        return [], 0

    items = []

    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    for p in pinjaman_rows:
        if (p.get('status') or '').strip() != 'Menunggu':
            continue
        try:
            nominal = float(p.get('plafon') or 0)
        except (TypeError, ValueError):
            nominal = 0.0
        items.append({
            'jenis': 'Pinjaman',
            'nama': (p.get('nama_anggota') or '-').strip() or '-',
            'tanggal': (p.get('tanggal_pengajuan') or '-').strip() or '-',
            'nominal': f"Rp {nominal:,.0f}",
            'url': '/pinjaman',
        })

    simpanan_rows = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    for s in simpanan_rows:
        if (s.get('status') or '').strip() != 'Menunggu':
            continue
        try:
            nominal = float(s.get('jumlah') or 0)
        except (TypeError, ValueError):
            nominal = 0.0
        jenis_simpanan = (s.get('jenis_simpanan') or 'Simpanan').strip() or 'Simpanan'
        items.append({
            'jenis': f"Simpanan ({jenis_simpanan})",
            'nama': (s.get('nama_anggota') or '-').strip() or '-',
            'tanggal': (s.get('tanggal_pengajuan') or '-').strip() or '-',
            'nominal': f"Rp {nominal:,.0f}",
            'url': '/simpanan',
        })

    items.sort(key=lambda x: x.get('tanggal', ''), reverse=True)
    total = len(items)
    return items[:limit], total


@app.context_processor
def inject_globals():
    notif_pengajuan_items, notif_pengajuan_count = _build_admin_pengajuan_notifications()
    current_role = (session.get('role') or '').strip().lower()
    
    # Ambil current permissions (dari DB jika aktif, otherwise fallback ke file-based)
    if _is_db_mode_enabled() and current_role:
        db_perms = _get_role_permissions_from_db(current_role)
        current_perms = list(db_perms) if db_perms else list(ROLE_PERMISSIONS.get(current_role, set()))
    else:
        current_perms = list(ROLE_PERMISSIONS.get(current_role, set()))
    
    return {
        'now': datetime.now().strftime('%d %B %Y'),
        'current_user': session.get('user'),
        'current_role': current_role,
        'is_admin': is_current_user_admin(),
        'current_id_anggota': get_current_user_id_anggota(),
        'csrf_token': _get_or_create_csrf_token(),
        'notif_pengajuan_items': notif_pengajuan_items,
        'notif_pengajuan_count': notif_pengajuan_count,
        'current_permissions': sorted(current_perms),
        'role_labels': ROLE_LABELS,
    }


# ══════════════════════════════════════════════
#  ROUTE: LANDING & AUTH
# ══════════════════════════════════════════════
@app.route('/')
def landing():
    if session.get('user'):
        return redirect(url_for('dashboard'))
    ensure_simpanan_schema()
    ensure_pinjaman_plafon_schema()
    berita_items = [item for item in _read_berita_items() if (item.get('status') or 'Aktif') == 'Aktif']
    anggota = baca_csv(FILE_ANGGOTA)
    simpanan = baca_csv(FILE_SIMPANAN)
    pinjaman = baca_csv(FILE_PINJAMAN)
    n_anggota = len(anggota)
    total_simpanan_disetujui = sum(float(s.get('total_simpanan') or 0) for s in simpanan)
    total_pinjaman_disalurkan = sum(
        float(p.get('plafon') or p.get('total_pinjaman') or 0)
        for p in pinjaman
        if p.get('status') in ('Disetujui', 'Lunas')
    )
    # Estimasi ilustratif (bukan pembukuan resmi SHU): proporsi dari total simpanan disetujui
    shu_estimasi_ilustratif = int(total_simpanan_disetujui * 0.08)
    return render_template(
        'landing.html',
        n_anggota=n_anggota,
        total_pinjaman_disalurkan=total_pinjaman_disalurkan,
        shu_estimasi_ilustratif=shu_estimasi_ilustratif,
        total_simpanan_disetujui=total_simpanan_disetujui,
        berita_items=berita_items,
    )


@app.route('/pendaftaran-anggota', methods=['POST'])
@csrf_protect
def pengajuan_anggota_baru():
    """Pengajuan anggota baru dari landing page."""
    ensure_pendaftaran_schema()
    nama_lengkap = (request.form.get('nama_lengkap') or '').strip()
    email = (request.form.get('email') or '').strip()
    alamat = (request.form.get('alamat') or '').strip()
    no_hp = (request.form.get('no_hp') or '').strip()
    kategori_anggota = (request.form.get('kategori_anggota') or '').strip()
    penghasilan = parse_rupiah_to_float(request.form.get('penghasilan_bersih', '0'))
    cicilan_lain = parse_rupiah_to_float(request.form.get('cicilan_lain', '0'))
    simpanan_pokok = parse_rupiah_to_float(request.form.get('simpanan_pokok', '500000'))
    simpanan_pokok_fix = float(SIMPANAN_RULES['Simpanan Pokok']['fixed'])

    if not nama_lengkap or not alamat:
        flash('Nama dan alamat wajib diisi untuk pengajuan anggota.', 'danger')
        return redirect(url_for('landing'))
    if penghasilan <= 0:
        flash('Penghasilan bersih per bulan wajib diisi (lebih dari Rp 0) untuk pengajuan anggota.', 'danger')
        return redirect(url_for('landing'))
    if simpanan_pokok != simpanan_pokok_fix:
        flash(f'Simpanan pokok wajib Rp {int(simpanan_pokok_fix):,}.'.replace(',', '.'), 'danger')
        return redirect(url_for('landing'))

    data = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    data.append({
        'id_pengajuan': str(uuid.uuid4()),
        'nama_lengkap': nama_lengkap,
        'email': email,
        'alamat': alamat,
        'no_hp': no_hp,
        'kategori_anggota': kategori_anggota,
        'penghasilan_bersih': str(int(penghasilan)),
        'cicilan_lain': str(int(cicilan_lain)),
        'simpanan_pokok': str(int(simpanan_pokok)),
        'status': 'Menunggu',
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'catatan_admin': '',
        'id_anggota_dibuat': '',
        'no_anggota_dibuat': ''
    })
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, data, PENDAFTARAN_FIELDNAMES)
    flash('Pengajuan anggota berhasil dikirim. Mohon tunggu konfirmasi admin.', 'success')
    return redirect(url_for('landing'))


@app.route('/robots.txt')
def robots_txt():
    return 'User-agent: *\nDisallow: /\n', 200, {'Content-Type': 'text/plain; charset=utf-8'}


@app.route('/login', methods=['GET', 'POST'])
@csrf_protect_if_post
def login():
    ensure_users_schema()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        next_url = request.form.get('next') or ''

        remaining_lock = _remaining_login_lock_seconds(username)
        if remaining_lock > 0:
            wait_min = max((remaining_lock + 59) // 60, 1)
            flash(f'Terlalu banyak percobaan login. Coba lagi dalam {wait_min} menit.', 'danger')
            return render_template('login.html', next=next_url), 429

        user = get_user_by_username(username)
        if not user or not check_password_hash(user.get('password_hash', ''), password):
            _register_failed_login(username)
            flash('Username atau password salah.', 'danger')
            return render_template('login.html', next=next_url), 401

        _clear_failed_login(username)
        session.clear()

        session['user'] = user.get('username')
        session['role'] = (user.get('role') or '').strip().lower()
        session['id_user'] = user.get('id_user')
        session['id_anggota'] = user.get('id_anggota', '')
        session['_csrf_token'] = token_hex(32)
        session['_last_activity_ts'] = datetime.utcnow().timestamp()
        session.permanent = True
        if session.get('role') in MEMBER_ROLES and not session.get('id_anggota'):
            session.clear()
            flash('Akun ini belum dihubungkan ke anggota. Hubungi admin.', 'danger')
            return render_template('login.html', next=next_url), 403
        flash(f"Selamat datang, {session['user']}!", 'success')
        role_home = _role_home_url(session.get('role') or 'user')
        safe_next = _safe_next_url(next_url, role_home)
        if not next_url or safe_next == url_for('dashboard'):
            safe_next = role_home
        return redirect(safe_next)

    if session.get('user'):
        return redirect(_role_home_url(session.get('role') or 'user'))
    return render_template('login.html', next=request.args.get('next', ''))


@app.route('/lupa-password', methods=['GET', 'POST'])
@csrf_protect_if_post
def lupa_password():
    """Reset password mandiri untuk akun user (verifikasi nomor anggota yang terhubung)."""
    ensure_users_schema()
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        no_anggota = (request.form.get('no_anggota') or '').strip()
        pwd1 = request.form.get('password') or ''
        pwd2 = request.form.get('password_confirm') or ''

        user = get_user_by_username(username)
        if not user or user.get('role') != 'user' or not user.get('id_anggota'):
            flash('Data tidak cocok atau akun tidak memenuhi syarat reset mandiri. Hubungi admin.', 'danger')
            return render_template('lupa_password.html'), 400

        anggota = get_anggota_by_id(user.get('id_anggota'))
        if (not anggota or
                (anggota.get('no_anggota') or '').strip().upper() != no_anggota.strip().upper()):
            flash('Data tidak cocok atau akun tidak memenuhi syarat reset mandiri. Hubungi admin.', 'danger')
            return render_template('lupa_password.html'), 400

        if pwd1 != pwd2:
            flash('Konfirmasi password tidak sama.', 'danger')
            return render_template('lupa_password.html'), 400

        ok, msg = validate_password_policy(pwd1)
        if not ok:
            flash(msg, 'danger')
            return render_template('lupa_password.html'), 400

        users = baca_csv(FILE_USERS)
        for u in users:
            if u.get('id_user') == user.get('id_user'):
                u['password_hash'] = generate_password_hash(pwd1)
                break
        tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
        flash('Password berhasil diubah. Silakan login.', 'success')
        return redirect(url_for('login'))

    return render_template('lupa_password.html')


@app.route('/logout')
@login_required
def logout():
    session.clear()
    flash('Anda berhasil logout.', 'success')
    return redirect(url_for('landing'))


# ══════════════════════════════════════════════
#  ROUTE: MANAJEMEN USER (ADMIN)
# ══════════════════════════════════════════════
@app.route('/users')
@admin_required
@permission_required('users.manage')
def users_index():
    ensure_users_schema()
    users = baca_csv(FILE_USERS)
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_map = {a.get('id_anggota'): a for a in anggota}
    users.sort(key=lambda u: ((u.get('role') or '') not in ADMIN_PANEL_ROLES, (u.get('username') or '').lower()))
    
    # Jika DB mode aktif, ambil role options dari DB
    role_options = ROLE_OPTIONS
    if _is_db_mode_enabled():
        db_roles = get_all_roles_from_db()
        if db_roles:
            role_options = [(r['role_name'], r['role_name'].replace('_', ' ').title()) for r in db_roles]
    
    return render_template(
        'users.html',
        users=users,
        anggota=anggota,
        anggota_map=anggota_map,
        role_options=role_options,
        role_labels=ROLE_LABELS,
        _is_db_mode=_is_db_mode_enabled()
    )


@app.route('/users/tambah', methods=['POST'])
@admin_required
@permission_required('users.manage')
@csrf_protect
def users_tambah():
    ensure_users_schema()
    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    role = request.form.get('role') or 'user'
    id_anggota = request.form.get('id_anggota') or ''

    if not username or not password:
        flash('Username dan password wajib diisi.', 'danger')
        return redirect(url_for('users_index'))
    ok_pol, msg_pol = validate_password_policy(password)
    if not ok_pol:
        flash(msg_pol, 'danger')
        return redirect(url_for('users_index'))
    if role not in {value for value, _ in ROLE_OPTIONS}:
        flash('Role tidak valid.', 'danger')
        return redirect(url_for('users_index'))
    if role not in MEMBER_ROLES and not has_permission('roles.manage'):
        flash('Hanya Super Admin yang dapat membuat role staf/non-anggota.', 'danger')
        return redirect(url_for('users_index'))
    if get_user_by_username(username):
        flash('Username sudah dipakai. Gunakan username lain.', 'danger')
        return redirect(url_for('users_index'))

    if role in MEMBER_ROLES and not id_anggota:
        flash('Untuk role anggota/user, wajib pilih anggota.', 'danger')
        return redirect(url_for('users_index'))

    if id_anggota:
        anggota = baca_csv(FILE_ANGGOTA)
        if not any(a.get('id_anggota') == id_anggota for a in anggota):
            flash('Anggota tidak ditemukan.', 'danger')
            return redirect(url_for('users_index'))

    users = baca_csv(FILE_USERS)
    users.append({
        'id_user': str(uuid.uuid4()),
        'username': username,
        'password_hash': generate_password_hash(password),
        'role': role,
        'id_anggota': id_anggota if role in MEMBER_ROLES else '',
        'created_at': datetime.now().strftime('%Y-%m-%d')
    })
    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    flash('Akun berhasil dibuat.', 'success')
    return redirect(url_for('users_index'))


@app.route('/users/reset_password/<id_user>', methods=['POST'])
@admin_required
@permission_required('users.manage')
@csrf_protect
def users_reset_password(id_user):
    ensure_users_schema()
    new_password = request.form.get('new_password') or ''
    if not new_password:
        flash('Password baru wajib diisi.', 'danger')
        return redirect(url_for('users_index'))
    ok_pol, msg_pol = validate_password_policy(new_password)
    if not ok_pol:
        flash(msg_pol, 'danger')
        return redirect(url_for('users_index'))

    users = baca_csv(FILE_USERS)
    found = False
    for u in users:
        if u.get('id_user') == id_user:
            u['password_hash'] = generate_password_hash(new_password)
            found = True
            break
    if not found:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('users_index'))

    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    flash('Password berhasil direset.', 'success')
    return redirect(url_for('users_index'))


@app.route('/users/hapus/<id_user>', methods=['POST'])
@admin_required
@permission_required('users.manage')
@csrf_protect
def users_hapus(id_user):
    ensure_users_schema()
    current_id = session.get('id_user')
    if current_id and current_id == id_user:
        flash('Tidak bisa menghapus akun yang sedang login.', 'danger')
        return redirect(url_for('users_index'))

    users = baca_csv(FILE_USERS)
    before = len(users)
    users = [u for u in users if u.get('id_user') != id_user]
    if len(users) == before:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('users_index'))
    tulis_csv(FILE_USERS, users, ['id_user', 'username', 'password_hash', 'role', 'id_anggota', 'created_at'])
    flash('User berhasil dihapus.', 'success')
    return redirect(url_for('users_index'))


# ══════════════════════════════════════════════
#  ROUTE: MANAJEMEN BERITA (ADMIN)
# ══════════════════════════════════════════════
@app.route('/berita-admin')
@admin_required
@permission_required('news.manage')
def berita_admin_index():
    berita_items = _read_berita_items()
    edit_id = (request.args.get('edit_id') or '').strip()
    edit_item = next((item for item in berita_items if item.get('id') == edit_id), None)
    if edit_id and not edit_item:
        flash('Data berita yang ingin diedit tidak ditemukan.', 'warning')
    return render_template('admin_berita.html', berita_items=berita_items, edit_item=edit_item)


@app.route('/berita-admin/simpan', methods=['POST'])
@admin_required
@permission_required('news.manage')
@csrf_protect
def berita_admin_simpan():
    berita_items = _read_berita_items()
    berita_id = (request.form.get('id') or '').strip()
    judul = (request.form.get('judul') or '').strip()
    kategori = (request.form.get('kategori') or 'Pengumuman').strip() or 'Pengumuman'
    isi = (request.form.get('isi') or '').strip()
    tanggal = (request.form.get('tanggal') or '').strip() or datetime.now().strftime('%Y-%m-%d')
    status = (request.form.get('status') or 'Aktif').strip() or 'Aktif'
    foto_file = request.files.get('foto')

    if not judul or not isi:
        flash('Judul dan isi berita wajib diisi.', 'danger')
        if berita_id:
            return redirect(url_for('berita_admin_index', edit_id=berita_id))
        return redirect(url_for('berita_admin_index'))

    if status not in ('Aktif', 'Nonaktif'):
        status = 'Aktif'

    if berita_id:
        target_item = None
        for item in berita_items:
            if item.get('id') == berita_id:
                target_item = item
                break
        if not target_item:
            flash('Berita yang ingin diedit tidak ditemukan.', 'danger')
            return redirect(url_for('berita_admin_index'))

        if foto_file and getattr(foto_file, 'filename', ''):
            new_foto = _save_berita_image(foto_file, berita_id)
            if not new_foto:
                flash('Upload foto gagal. Gunakan PNG/JPG/JPEG/WEBP maksimal 2MB.', 'danger')
                return redirect(url_for('berita_admin_index', edit_id=berita_id))
            old_foto = (target_item.get('foto') or '').strip()
            target_item['foto'] = new_foto
            if old_foto and old_foto != new_foto:
                _delete_static_file_relpath(old_foto)

        target_item['judul'] = judul
        target_item['kategori'] = kategori
        target_item['isi'] = isi
        target_item['tanggal'] = tanggal
        target_item['status'] = status
        message = 'Berita berhasil diperbarui.'
    else:
        new_id = str(uuid.uuid4())
        foto = ''
        if foto_file and getattr(foto_file, 'filename', ''):
            foto = _save_berita_image(foto_file, new_id)
            if not foto:
                flash('Upload foto gagal. Gunakan PNG/JPG/JPEG/WEBP maksimal 2MB.', 'danger')
                return redirect(url_for('berita_admin_index'))
        berita_items.append({
            'id': new_id,
            'judul': judul,
            'kategori': kategori,
            'isi': isi,
            'tanggal': tanggal,
            'status': status,
            'foto': foto,
        })
        message = 'Berita berhasil ditambahkan.'

    _write_berita_items(berita_items)
    flash(message, 'success')
    return redirect(url_for('berita_admin_index'))


@app.route('/berita-admin/hapus/<berita_id>', methods=['POST'])
@admin_required
@permission_required('news.manage')
@csrf_protect
def berita_admin_hapus(berita_id):
    berita_items = _read_berita_items()
    target_item = next((item for item in berita_items if item.get('id') == berita_id), None)
    if not target_item:
        flash('Berita tidak ditemukan.', 'danger')
        return redirect(url_for('berita_admin_index'))

    foto = (target_item.get('foto') or '').strip()
    berita_items = [item for item in berita_items if item.get('id') != berita_id]
    if foto:
        _delete_static_file_relpath(foto)

    _write_berita_items(berita_items)
    flash('Berita berhasil dihapus.', 'success')
    return redirect(url_for('berita_admin_index'))


@app.route('/berita-admin/hapus-foto/<berita_id>', methods=['POST'])
@admin_required
@permission_required('news.manage')
@csrf_protect
def berita_admin_hapus_foto(berita_id):
    berita_items = _read_berita_items()
    target_item = next((item for item in berita_items if item.get('id') == berita_id), None)
    if not target_item:
        flash('Berita tidak ditemukan.', 'danger')
        return redirect(url_for('berita_admin_index'))

    foto = (target_item.get('foto') or '').strip()
    if not foto:
        flash('Berita ini belum memiliki foto.', 'warning')
        return redirect(url_for('berita_admin_index', edit_id=berita_id))

    _delete_static_file_relpath(foto)
    target_item['foto'] = ''
    _write_berita_items(berita_items)
    flash('Foto berita berhasil dihapus.', 'success')
    return redirect(url_for('berita_admin_index', edit_id=berita_id))


def bunga_untuk_jenis_pinjaman(jenis: str, tenor_bulan: int) -> float:
    """Bunga bulanan per produk + validasi tenor tiap jenis pinjaman."""
    j = (jenis or '').strip()
    if j == 'Solusi Cepat':
        if 1 <= tenor_bulan <= 2:
            return 2.0
        raise ValueError('Tenor Solusi Cepat wajib 1-2 bulan.')
    if j == 'Jangka Pendek':
        if 1 <= tenor_bulan <= 12:
            return 1.5
        raise ValueError('Tenor Jangka Pendek wajib 1-12 bulan.')
    if j == 'Jangka Panjang':
        if 13 <= tenor_bulan <= 24:
            return 0.80
        if 25 <= tenor_bulan <= 120:
            return 0.75
        raise ValueError('Tenor Jangka Panjang wajib 13-24 bulan atau 25-120 bulan.')
    if j == 'Modal Usaha':
        if 1 <= tenor_bulan <= 120:
            return 0.50
        raise ValueError('Tenor Modal Usaha wajib 1-120 bulan.')
    if j in JENIS_PINJAMAN:
        return float(JENIS_PINJAMAN[j]['bunga'])
    return bunga_dari_tenor(tenor_bulan)


# ══════════════════════════════════════════════
#  ROUTE: DASHBOARD
# ══════════════════════════════════════════════
@app.route('/dashboard')
@login_required
def dashboard():
    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    ensure_pinjaman_cicilan_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    simpanan = baca_csv(FILE_SIMPANAN)
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    pinjaman = baca_csv(FILE_PINJAMAN)
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    current_role = _current_role()

    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]
        simpanan = [s for s in simpanan if s.get('id_anggota') == id_anggota]
        simpanan_transaksi = [t for t in simpanan_transaksi if t.get('id_anggota') == id_anggota]
        pinjaman = [p for p in pinjaman if p.get('id_anggota') == id_anggota]
        cicilan = [c for c in cicilan if c.get('id_anggota') == id_anggota]

    total_anggota = len(anggota)
    total_simpanan = sum(float(s.get('total_simpanan') or 0) for s in simpanan)
    pv_aktif = [p for p in pinjaman if p.get('status') == 'Disetujui']
    total_pinjaman = sum(float(p.get('plafon') or 0) for p in pv_aktif)
    total_pinjaman_beredar = sum(saldo_pinjaman_aktual(p) for p in pv_aktif)
    pinjaman_per_jenis = {jenis: 0.0 for jenis in JENIS_PINJAMAN_CHOICES}
    for p in pv_aktif:
        jenis = (p.get('jenis_pinjaman') or '').strip()
        if jenis not in pinjaman_per_jenis:
            continue
        try:
            pinjaman_per_jenis[jenis] += float(p.get('plafon') or 0)
        except (TypeError, ValueError):
            continue

    # Saldo simpanan per jenis dihitung dari transaksi, dengan fallback untuk data lama.
    simpanan_per_jenis = {j: 0.0 for j in SIMPANAN_AKUMULASI_CHOICES}
    total_transaksi_simpanan = 0.0
    for t in simpanan_transaksi:
        jenis = normalize_jenis_simpanan(t.get('jenis_simpanan'))
        if jenis not in simpanan_per_jenis:
            continue
        try:
            nominal = float(t.get('jumlah') or 0)
        except (TypeError, ValueError):
            nominal = 0.0
        simpanan_per_jenis[jenis] += nominal
        total_transaksi_simpanan += nominal
    if total_simpanan > total_transaksi_simpanan:
        simpanan_per_jenis[SIMPANAN_DEFAULT_AKUMULASI] += (total_simpanan - total_transaksi_simpanan)

    # Ringkasan aktivitas: gabungan semua transaksi simpanan, pinjaman, dan cicilan.
    transaksi_terakhir = []

    def append_transaksi(tipe: str, tanggal: str, nama: str, keterangan: str, status: str = ''):
        transaksi_terakhir.append({
            'tipe': tipe,
            'tanggal': tanggal or '-',
            'nama': nama or '',
            'keterangan': keterangan,
            'status': status or '-',
        })

    for t in simpanan_transaksi:
        append_transaksi(
            'Simpanan',
            t.get('tanggal', '-'),
            t.get('nama_anggota', ''),
            f"Simpanan {t.get('jenis_simpanan', 'Manasuka')} — Rp {float(t.get('jumlah') or 0):,.0f}",
            'Disetujui',
        )

    for p in pinjaman:
        tanggal_evt = p.get('tanggal_lunas') or p.get('tanggal_pengajuan') or '-'
        status_evt = p.get('status') or '-'
        jenis_evt = p.get('jenis_pinjaman') or 'Pinjaman'
        keterangan_evt = f"{jenis_evt} — Rp {float(p.get('plafon') or 0):,.0f}"
        if status_evt == 'Menunggu':
            keterangan_evt = f"Pengajuan {keterangan_evt}"
        elif status_evt == 'Disetujui':
            keterangan_evt = f"Pinjaman disetujui {keterangan_evt}"
        elif status_evt == 'Lunas':
            keterangan_evt = f"Pinjaman lunas {keterangan_evt}"
        elif status_evt == 'Ditolak':
            keterangan_evt = f"Pengajuan ditolak {keterangan_evt}"
        append_transaksi('Pinjaman', tanggal_evt, p.get('nama_anggota', ''), keterangan_evt, status_evt)

    for c in cicilan:
        status_evt = c.get('status') or '-'
        tanggal_evt = c.get('tanggal_konfirmasi') or c.get('tanggal_pengajuan') or '-'
        if status_evt == 'Menunggu':
            label_evt = 'Pengajuan cicilan'
        elif status_evt == 'Disetujui':
            label_evt = 'Cicilan diterima'
        elif status_evt == 'Gagal Bayar':
            label_evt = 'Cicilan gagal bayar'
        elif status_evt == 'Ditolak':
            label_evt = 'Cicilan ditolak'
        else:
            label_evt = 'Cicilan'
        keterangan_evt = f"{label_evt} — Rp {float(c.get('jumlah') or 0):,.0f}"
        if c.get('metode_pembayaran'):
            keterangan_evt += f" ({c.get('metode_pembayaran')})"
        status_tampil = 'Cicilan diterima' if status_evt == 'Disetujui' else status_evt
        append_transaksi('Cicilan', tanggal_evt, c.get('nama_anggota', ''), keterangan_evt, status_tampil)

    transaksi_terakhir.sort(key=lambda x: x.get('tanggal', ''), reverse=True)

    pinjaman_dashboard = []
    for p in sorted(pinjaman, key=lambda x: x.get('tanggal_pengajuan', ''), reverse=True):
        try:
            plafon = float(p.get('plafon') or 0)
        except (TypeError, ValueError):
            plafon = 0.0
        try:
            tenor = int(float(p.get('tenor_bulan') or p.get('tenor_awal') or 0))
        except (TypeError, ValueError):
            tenor = 0
        pinjaman_dashboard.append({
            'id_pinjaman': p.get('id_pinjaman', ''),
            'no_anggota': p.get('no_anggota', ''),
            'nama_anggota': p.get('nama_anggota', ''),
            'jenis_pinjaman': p.get('jenis_pinjaman', '-') or '-',
            'plafon': plafon,
            'sisa_pinjaman': saldo_pinjaman_aktual(p),
            'tenor_bulan': tenor,
            'status': p.get('status', '-') or '-',
            'tanggal_pengajuan': p.get('tanggal_pengajuan', '-') or '-',
            'tanggal_lunas': p.get('tanggal_lunas', '-') or '-',
        })

    return render_template('dashboard.html',
                           total_anggota=total_anggota,
                           total_simpanan=total_simpanan,
                           total_pinjaman=total_pinjaman,
                           total_pinjaman_beredar=total_pinjaman_beredar,
                           simpanan_per_jenis=simpanan_per_jenis,
                           pinjaman_per_jenis=pinjaman_per_jenis,
                           transaksi_terakhir=transaksi_terakhir,
                           pinjaman_dashboard=pinjaman_dashboard,
                           is_auditor_dashboard=(current_role == 'auditor'))


@app.route('/pengajuan-anggota/konfirmasi/<id_pengajuan>', methods=['POST'])
@admin_required
@permission_required('members.approve')
@csrf_protect
def konfirmasi_pengajuan_anggota(id_pengajuan):
    ensure_pendaftaran_schema()
    pengajuan = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    found = False
    for p in pengajuan:
        if p.get('id_pengajuan') == id_pengajuan:
            p['status'] = 'Disetujui'
            p['catatan_admin'] = (
                f"Dikonfirmasi manual oleh {session.get('user', 'admin')} pada {datetime.now().strftime('%Y-%m-%d')}"
            )
            found = True
            break
    if not found:
        flash('Pengajuan anggota tidak ditemukan.', 'danger')
        return redirect(url_for('dashboard'))
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, pengajuan, PENDAFTARAN_FIELDNAMES)
    flash('Pengajuan anggota disetujui secara manual.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/pengajuan-anggota/tolak/<id_pengajuan>', methods=['POST'])
@admin_required
@permission_required('members.approve')
@csrf_protect
def tolak_pengajuan_anggota(id_pengajuan):
    ensure_pendaftaran_schema()
    pengajuan = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
    found = False
    for p in pengajuan:
        if p.get('id_pengajuan') == id_pengajuan:
            p['status'] = 'Ditolak'
            p['catatan_admin'] = f"Ditolak {session.get('user', 'admin')} pada {datetime.now().strftime('%Y-%m-%d')}"
            found = True
            break
    if not found:
        flash('Pengajuan anggota tidak ditemukan.', 'danger')
        return redirect(url_for('dashboard'))
    tulis_csv(FILE_PENDAFTARAN_ANGGOTA, pengajuan, PENDAFTARAN_FIELDNAMES)
    flash('Pengajuan anggota ditolak.', 'warning')
    return redirect(url_for('dashboard'))


# ══════════════════════════════════════════════
#  ROUTE: ANGGOTA
# ══════════════════════════════════════════════
@app.route('/anggota')
@app.route('/anggota/')
@login_required
@permission_required('members.view', 'members.self.view')
def halaman_anggota():
    ensure_anggota_schema()
    anggota_all = baca_csv(FILE_ANGGOTA)
    anggota = anggota_all
    current_role = _current_role()

    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]
    else:
        for a in anggota:
            info = info_pinjaman_dsr_anggota(a)
            a['_dsr_persen'] = info['dsr_persen']
            a['_kap_cicilan'] = info['kapasitas_cicilan']

    auditor_transaksi = []
    auditor_riwayat_pinjaman = []
    if current_role == 'auditor':
        ensure_simpanan_transaksi_schema()
        ensure_pinjaman_cicilan_schema()
        simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
        pinjaman = baca_csv(FILE_PINJAMAN)
        cicilan = baca_csv(FILE_PINJAMAN_CICILAN)

        def append_transaksi(tipe: str, tanggal: str, nama: str, keterangan: str, status: str = ''):
            auditor_transaksi.append({
                'tipe': tipe,
                'tanggal': tanggal or '-',
                'nama': nama or '',
                'keterangan': keterangan,
                'status': status or '-',
            })

        for t in simpanan_transaksi:
            append_transaksi(
                'Simpanan',
                t.get('tanggal', '-'),
                t.get('nama_anggota', ''),
                f"Simpanan {t.get('jenis_simpanan', 'Manasuka')} - Rp {float(t.get('jumlah') or 0):,.0f}",
                'Disetujui',
            )

        for p in pinjaman:
            tanggal_evt = p.get('tanggal_lunas') or p.get('tanggal_pengajuan') or '-'
            status_evt = p.get('status') or '-'
            jenis_evt = p.get('jenis_pinjaman') or 'Pinjaman'
            keterangan_evt = f"{jenis_evt} - Rp {float(p.get('plafon') or 0):,.0f}"
            if status_evt == 'Menunggu':
                keterangan_evt = f"Pengajuan {keterangan_evt}"
            elif status_evt == 'Disetujui':
                keterangan_evt = f"Pinjaman disetujui {keterangan_evt}"
            elif status_evt == 'Lunas':
                keterangan_evt = f"Pinjaman lunas {keterangan_evt}"
            elif status_evt == 'Ditolak':
                keterangan_evt = f"Pengajuan ditolak {keterangan_evt}"
            append_transaksi('Pinjaman', tanggal_evt, p.get('nama_anggota', ''), keterangan_evt, status_evt)

        for c in cicilan:
            status_evt = c.get('status') or '-'
            tanggal_evt = c.get('tanggal_konfirmasi') or c.get('tanggal_pengajuan') or '-'
            if status_evt == 'Menunggu':
                label_evt = 'Pengajuan cicilan'
            elif status_evt == 'Disetujui':
                label_evt = 'Cicilan diterima'
            elif status_evt == 'Gagal Bayar':
                label_evt = 'Cicilan gagal bayar'
            elif status_evt == 'Ditolak':
                label_evt = 'Cicilan ditolak'
            else:
                label_evt = 'Cicilan'
            keterangan_evt = f"{label_evt} - Rp {float(c.get('jumlah') or 0):,.0f}"
            if c.get('metode_pembayaran'):
                keterangan_evt += f" ({c.get('metode_pembayaran')})"
            status_tampil = 'Cicilan diterima' if status_evt == 'Disetujui' else status_evt
            append_transaksi('Cicilan', tanggal_evt, c.get('nama_anggota', ''), keterangan_evt, status_tampil)

        auditor_transaksi.sort(key=lambda x: x.get('tanggal', ''), reverse=True)
        auditor_transaksi = auditor_transaksi[:120]

        riwayat_map = {}
        for p in pinjaman:
            id_a = p.get('id_anggota', '')
            if not id_a:
                continue
            tenor_raw = p.get('tenor_bulan', p.get('tenor_awal', '0'))
            try:
                tenor_norm = int(float(tenor_raw or 0))
            except (TypeError, ValueError):
                tenor_norm = 0
            jenis = (p.get('jenis_pinjaman') or '').strip() or kategori_pinjaman_dari_tenor(tenor_norm)
            if not jenis:
                continue

            row = riwayat_map.setdefault(id_a, {
                'id_anggota': id_a,
                'no_anggota': p.get('no_anggota', ''),
                'nama_anggota': p.get('nama_anggota', ''),
                'total_pengajuan': 0,
                'total_aktif': 0,
                'jenis_list': [],
            })
            if not row['no_anggota']:
                row['no_anggota'] = p.get('no_anggota', '')
            if not row['nama_anggota']:
                row['nama_anggota'] = p.get('nama_anggota', '')

            row['total_pengajuan'] += 1
            if (p.get('status') or '').strip() == 'Disetujui' and saldo_pinjaman_aktual(p) > 0:
                row['total_aktif'] += 1
            if jenis not in row['jenis_list']:
                row['jenis_list'].append(jenis)

        anggota_map = {a.get('id_anggota'): a for a in anggota_all}
        for item in riwayat_map.values():
            a = anggota_map.get(item['id_anggota'], {})
            no_anggota = item['no_anggota'] or a.get('no_anggota', '-')
            nama_anggota = item['nama_anggota'] or a.get('nama_lengkap', '-')
            jenis_items = item['jenis_list']
            auditor_riwayat_pinjaman.append({
                'id_anggota': item['id_anggota'],
                'no_anggota': no_anggota,
                'nama_anggota': nama_anggota,
                'total_pengajuan': item['total_pengajuan'],
                'total_aktif': item['total_aktif'],
                'jenis_text': ', '.join(jenis_items),
            })
        auditor_riwayat_pinjaman.sort(key=lambda x: (x.get('nama_anggota') or '').lower())

    pengajuan_anggota = []
    jumlah_pengajuan_menunggu = 0
    if current_role == 'super_admin':
        ensure_pendaftaran_schema()
        pengajuan_anggota = baca_csv(FILE_PENDAFTARAN_ANGGOTA)
        pengajuan_anggota = [p for p in pengajuan_anggota if (p.get('status') or '') == 'Menunggu']
        pengajuan_anggota.sort(key=lambda x: x.get('tanggal_pengajuan', ''), reverse=True)
        jumlah_pengajuan_menunggu = len(pengajuan_anggota)
        pengajuan_anggota = pengajuan_anggota[:10]

    return render_template(
        'anggota.html',
        anggota=anggota,
        auditor_transaksi=auditor_transaksi,
        auditor_riwayat_pinjaman=auditor_riwayat_pinjaman,
        pengajuan_anggota=pengajuan_anggota,
        jumlah_pengajuan_menunggu=jumlah_pengajuan_menunggu,
        is_auditor_view=(current_role == 'auditor'),
        is_ketua_view=False,
        is_super_admin_view=(current_role == 'super_admin'),
    )


@app.route('/anggota/tambah', methods=['POST'])
@admin_required
@permission_required('members.manage')
@csrf_protect
def tambah_anggota():
    ensure_anggota_schema()
    data = baca_csv(FILE_ANGGOTA)
    simpanan = baca_csv(FILE_SIMPANAN)
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    nik = normalize_nik(request.form.get('nik'))
    if nik and not is_valid_nik(nik):
        flash('NIK harus terdiri dari tepat 16 digit angka.', 'danger')
        return redirect('/anggota')
    nama_lengkap = (request.form.get('nama_lengkap') or '').strip()
    email = (request.form.get('email') or '').strip()
    alamat = (request.form.get('alamat') or '').strip()
    no_hp = (request.form.get('no_hp') or '').strip()
    kategori_anggota = (request.form.get('kategori_anggota') or '').strip()
    no_rekening = (request.form.get('no_rekening') or '').strip()
    nama_bank = (request.form.get('nama_bank') or '').strip()
    status_kredit = (request.form.get('status_kredit') or 'Lancar').strip()
    foto_ktp = (request.form.get('foto_ktp') or '').strip()
    penghasilan_bersih = str(parse_rupiah_to_float(request.form.get('penghasilan_bersih', '0')))
    cicilan_lain = str(parse_rupiah_to_float(request.form.get('cicilan_lain', '0')))
    simpanan_pokok = parse_rupiah_to_float(request.form.get('simpanan_pokok', '500000'))
    simpanan_pokok_fix = float(SIMPANAN_RULES['Simpanan Pokok']['fixed'])
    no_anggota = (request.form.get('no_anggota') or '').strip()
    if not no_anggota:
        no_anggota = generate_no_anggota_berikutnya(data)
    if simpanan_pokok != simpanan_pokok_fix:
        flash(f'Simpanan pokok wajib Rp {int(simpanan_pokok_fix):,}.'.replace(',', '.'), 'danger')
        return redirect('/anggota')
    target = None
    if nik:
        target = next((a for a in data if (a.get('nik') or '').strip().upper() == nik.upper()), None)
    if target is None and no_anggota:
        target = next((a for a in data if (a.get('no_anggota') or '').strip().upper() == no_anggota.upper()), None)

    if target:
        target['no_anggota'] = no_anggota or target.get('no_anggota', '')
        target['nik'] = nik or target.get('nik', '')
        target['nama_lengkap'] = nama_lengkap or target.get('nama_lengkap', '')
        target['email'] = email or target.get('email', '')
        target['alamat'] = alamat or target.get('alamat', '')
        target['no_hp'] = no_hp or target.get('no_hp', '')
        target['kategori_anggota'] = kategori_anggota or target.get('kategori_anggota', '')
        target['no_rekening'] = no_rekening or target.get('no_rekening', '')
        target['nama_bank'] = nama_bank or target.get('nama_bank', '')
        target['status_kredit'] = status_kredit
        target['foto_ktp'] = foto_ktp or target.get('foto_ktp', '')
        target['penghasilan_bersih'] = penghasilan_bersih
        target['cicilan_lain'] = cicilan_lain
        target['simpanan_pokok'] = str(int(simpanan_pokok))
        catat_simpanan_pokok_awal(
            target,
            simpanan,
            simpanan_transaksi,
            simpanan_pokok,
            session.get('user') or '',
            'Simpanan Pokok validasi anggota baru',
        )
        flash('Data anggota berhasil diperbarui tanpa membuat data ganda.', 'success')
    else:
        anggota_baru = {
            'id_anggota': str(uuid.uuid4()),
            'no_anggota': no_anggota,
            'nik': nik,
            'nama_lengkap': nama_lengkap,
            'email': email,
            'alamat': alamat,
            'no_hp': no_hp,
            'kategori_anggota': kategori_anggota,
            'no_rekening': no_rekening,
            'nama_bank': nama_bank,
            'status_kredit': status_kredit,
            'foto_ktp': foto_ktp,
            'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
            'status_anggota': 'Aktif',
            'penghasilan_bersih': penghasilan_bersih,
            'cicilan_lain': cicilan_lain,
            'simpanan_pokok': str(int(simpanan_pokok)),
        }
        data.append(anggota_baru)
        catat_simpanan_pokok_awal(
            anggota_baru,
            simpanan,
            simpanan_transaksi,
            simpanan_pokok,
            session.get('user') or '',
            'Simpanan Pokok validasi anggota baru',
        )
        flash('Data anggota berhasil ditambahkan.', 'success')
    tulis_csv(FILE_ANGGOTA, data, ANGGOTA_FIELDNAMES)
    tulis_csv(FILE_SIMPANAN, simpanan, SIMPANAN_FIELDNAMES)
    tulis_csv(FILE_SIMPANAN_TRANSAKSI, simpanan_transaksi, SIMPANAN_TRANSAKSI_FIELDNAMES)
    return redirect('/anggota')


@app.route('/anggota/hapus/<id_anggota>', methods=['POST'])
@admin_required
@permission_required('members.manage')
@csrf_protect
def hapus_anggota(id_anggota):
    data = baca_csv(FILE_ANGGOTA)
    data = [a for a in data if a['id_anggota'] != id_anggota]
    tulis_csv(FILE_ANGGOTA, data, ANGGOTA_FIELDNAMES)
    return redirect('/anggota')


@app.route('/anggota/edit/<id_anggota>', methods=['GET', 'POST'])
@login_required
@permission_required('members.manage', 'members.manage.limited', 'members.self.edit.limited')
@csrf_protect_if_post
def edit_anggota(id_anggota):
    ensure_anggota_schema()
    data = baca_csv(FILE_ANGGOTA)
    idx = next((i for i, a in enumerate(data) if a.get('id_anggota') == id_anggota), None)
    if idx is None:
        flash('Anggota tidak ditemukan.', 'danger')
        return redirect(url_for('halaman_anggota'))

    can_full_edit = has_permission('members.manage')
    can_limited_edit = has_permission('members.manage.limited')
    can_self_edit = has_permission('members.self.edit.limited')
    if (can_limited_edit or can_self_edit) and not can_full_edit:
        target_id = data[idx].get('id_anggota')
        if can_self_edit:
            restrict_id_anggota_or_forbid(target_id)
    if request.method == 'POST':
        nik = normalize_nik(request.form.get('nik'))
        if nik and not is_valid_nik(nik):
            flash('NIK harus terdiri dari tepat 16 digit angka.', 'danger')
            return redirect(url_for('edit_anggota', id_anggota=id_anggota))
        if can_full_edit:
            data[idx]['nik'] = nik
            data[idx]['nama_lengkap'] = (request.form.get('nama_lengkap') or '').strip()
            data[idx]['email'] = (request.form.get('email') or '').strip()
            data[idx]['alamat'] = (request.form.get('alamat') or '').strip()
            data[idx]['no_hp'] = (request.form.get('no_hp') or '').strip()
            data[idx]['kategori_anggota'] = (request.form.get('kategori_anggota') or '').strip()
            data[idx]['status_kredit'] = (request.form.get('status_kredit') or '').strip()
            data[idx]['no_rekening'] = (request.form.get('no_rekening') or '').strip()
            data[idx]['nama_bank'] = (request.form.get('nama_bank') or '').strip()
            data[idx]['penghasilan_bersih'] = str(int(parse_rupiah_to_float(request.form.get('penghasilan_bersih', '0'))))
            data[idx]['cicilan_lain'] = str(int(parse_rupiah_to_float(request.form.get('cicilan_lain', '0'))))
        else:
            # Edit terbatas: hanya kontak dan rekening.
            data[idx]['email'] = (request.form.get('email') or '').strip()
            data[idx]['alamat'] = (request.form.get('alamat') or '').strip()
            data[idx]['no_hp'] = (request.form.get('no_hp') or '').strip()
            data[idx]['no_rekening'] = (request.form.get('no_rekening') or '').strip()
            data[idx]['nama_bank'] = (request.form.get('nama_bank') or '').strip()
        tulis_csv(FILE_ANGGOTA, data, ANGGOTA_FIELDNAMES)
        flash('Data anggota berhasil diperbarui.', 'success')
        return redirect(url_for('halaman_anggota'))

    return render_template('anggota_edit.html', anggota=data[idx])


@app.route('/anggota/import-excel', methods=['POST'])
@permission_required('excel.import')
@csrf_protect
def import_anggota_excel():
    """Import anggota dari Excel dengan smart upsert berdasarkan NIK, fallback No Anggota."""
    ensure_anggota_schema()
    if load_workbook is None:
        flash('Fitur import Excel membutuhkan openpyxl.', 'danger')
        return redirect(url_for('halaman_anggota'))

    file = request.files.get('file_excel')
    if not file or not file.filename:
        flash('File Excel belum dipilih.', 'danger')
        return redirect(url_for('halaman_anggota'))
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('Format file tidak didukung. Gunakan .xlsx atau .xlsm', 'danger')
        return redirect(url_for('halaman_anggota'))

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        flash(f'Gagal membaca file Excel: {e}', 'danger')
        return redirect(url_for('halaman_anggota'))

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        flash('File Excel kosong.', 'danger')
        return redirect(url_for('halaman_anggota'))

    def norm(h):
        return str(h or '').strip().lower().replace(' ', '_')

    header_map = {norm(h): i for i, h in enumerate(headers_raw)}

    def get_val(row, *aliases):
        for a in aliases:
            idx = header_map.get(a)
            if idx is not None and idx < len(row):
                v = row[idx]
                return '' if v is None else str(v).strip()
        return ''

    anggota = baca_csv(FILE_ANGGOTA)
    by_nik = {
        normalize_nik(a.get('nik')): a
        for a in anggota
        if normalize_nik(a.get('nik'))
    }
    by_no = {(a.get('no_anggota') or '').strip(): a for a in anggota if (a.get('no_anggota') or '').strip()}

    added = 0
    updated = 0
    max_no = 0
    for a in anggota:
        no = (a.get('no_anggota') or '').strip()
        if no.startswith('AGT-'):
            num = no.replace('AGT-', '')
            if num.isdigit():
                max_no = max(max_no, int(num))

    invalid_nik_rows = 0
    for row in rows_iter:
        nik = normalize_nik(get_val(row, 'nik'))
        no_anggota = get_val(row, 'no_anggota')
        nama = get_val(row, 'nama')
        alamat = get_val(row, 'alamat')
        no_telp = get_val(row, 'no_telp', 'no_hp', 'hp', 'telp', 'no_telepon', 'telepon')
        no_rekening = get_val(row, 'no_rekening', 'norek', 'rekening')
        nama_bank = get_val(row, 'nama_bank', 'bank')
        penghasilan = get_val(row, 'penghasilan_bersih')
        cicilan = get_val(row, 'cicilan_lain')
        if nik and not is_valid_nik(nik):
            invalid_nik_rows += 1
            continue
        if not (nik or no_anggota or nama):
            continue

        target = None
        if nik and nik in by_nik:
            target = by_nik[nik]
        elif no_anggota and no_anggota in by_no:
            target = by_no[no_anggota]

        if target:
            target['nik'] = nik or target.get('nik', '')
            target['nama_lengkap'] = nama or target.get('nama_lengkap', '')
            target['alamat'] = alamat or target.get('alamat', '')
            target['no_telp'] = no_telp or target.get('no_telp', '')
            target['no_rekening'] = no_rekening or target.get('no_rekening', '')
            target['nama_bank'] = nama_bank or target.get('nama_bank', '')
            target['penghasilan_bersih'] = str(parse_rupiah_to_float(penghasilan or target.get('penghasilan_bersih', '0')))
            target['cicilan_lain'] = str(parse_rupiah_to_float(cicilan or target.get('cicilan_lain', '0')))
            updated += 1
            continue

        if not no_anggota:
            max_no += 1
            no_anggota = f"AGT-{max_no:04d}"

        new_row = {
            'id_anggota': str(uuid.uuid4()),
            'no_anggota': no_anggota,
            'nik': nik,
            'nama': nama,
            'alamat': alamat,
            'no_telp': no_telp,
            'no_rekening': no_rekening,
            'nama_bank': nama_bank,
            'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
            'penghasilan_bersih': str(parse_rupiah_to_float(penghasilan or '0')),
            'cicilan_lain': str(parse_rupiah_to_float(cicilan or '0')),
        }
        anggota.append(new_row)
        if nik:
            by_nik[nik] = new_row
        if no_anggota:
            by_no[no_anggota] = new_row
        added += 1

    tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
    flash(f'Import selesai: {added} anggota baru ditambahkan, {updated} data anggota diperbarui.', 'success')
    if invalid_nik_rows:
        flash(f'{invalid_nik_rows} baris dilewati karena NIK tidak valid (harus 16 digit angka).', 'warning')
    return redirect(url_for('halaman_anggota'))


def _norm_csv_header(h):
    return str(h or '').strip().lower().replace(' ', '_').replace('-', '_')


def _map_import_csv_headers(fieldnames):
    """Map normalized header -> canonical key."""
    if not fieldnames:
        return None
    norm = {_norm_csv_header(h): h for h in fieldnames}
    aliases = {
        'nama': ('nama', 'name'),
        'nik': ('nik',),
        'no_anggota': ('no_anggota', 'no anggota'),
        'no_hp': ('no_hp', 'nohp', 'no_telp', 'telepon', 'telp', 'hp'),
        'no_rekening': ('no_rekening', 'norek', 'rekening'),
        'nama_bank': ('nama_bank', 'bank'),
        'alamat': ('alamat', 'address'),
        'jenis_pinjaman': ('jenis_pinjaman', 'pilihan_pinjaman', 'kategori_pinjaman'),
        'jenis_simpanan': ('jenis_simpanan', 'pilihan_simpanan', 'kategori_simpanan'),
        'simpanan': ('simpanan', 'tabungan'),
        'pinjaman': ('pinjaman', 'hutang'),
        'tenor_bulan': ('tenor_bulan', 'tenor', 'jangka_waktu', 'lama_bulan'),
    }
    out = {}
    for key, alist in aliases.items():
        for a in alist:
            if a in norm:
                out[key] = norm[a]
                break
    if 'nik' not in out:
        return None
    # Hindari file "satu kolom".
    if len(out.keys()) < 2:
        return None
    return out


def _parse_opt_float_cell(raw):
    s = (raw or '').strip()
    if not s:
        return None
    return parse_rupiah_to_float(s)


def parse_anggota_csv_upload(file_storage):
    """Validasi & parse file Excel ringkasan anggota. Mengembalikan dict preview atau raise ValueError."""
    if not file_storage or not file_storage.filename:
        raise ValueError('File belum dipilih.')
    ext = os.path.splitext(file_storage.filename.lower())[1]
    if ext not in ('.xlsx', '.xlsm'):
        raise ValueError('Gunakan file Excel (.xlsx atau .xlsm).')
    file_storage.stream.seek(0)
    raw = file_storage.read()
    if len(raw) > 2 * 1024 * 1024:
        raise ValueError('Ukuran file maksimal 2 MB.')
    rows_src = []
    detected_columns = []
    if load_workbook is None:
        raise ValueError('Import Excel membutuhkan openpyxl.')
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError('File kosong.')
    headers = [str(h or '').strip() for h in all_rows[0]]
    header_non_empty = [h for h in headers if h]
    if len(header_non_empty) < 2:
        raise ValueError('Format Excel tidak valid: header harus terpisah di beberapa kolom, bukan satu kolom gabungan.')
    fieldnames = headers
    for rr in all_rows[1:]:
        row_map = {}
        for i, h in enumerate(headers):
            v = rr[i] if i < len(rr) else ''
            row_map[h] = '' if v is None else str(v).strip()
        # Cegah pola data gabungan dalam satu kolom (mis: "Nama,NIK,No HP,...")
        isi_terisi = [v for v in row_map.values() if str(v or '').strip()]
        if len(isi_terisi) == 1 and any(sep in isi_terisi[0] for sep in (',', ';', '|', '\t')):
            raise ValueError('Format Excel tidak valid: data terdeteksi gabungan dalam satu kolom. Pisahkan per kolom.')
        rows_src.append(row_map)
    header_map = _map_import_csv_headers(fieldnames)
    if not header_map:
        raise ValueError(
            'Header tidak valid. Wajib ada kolom nik; disarankan: nama,nik,no hp,simpanan,pinjaman,tenor'
        )
    detected_columns = [k for k in ('nama', 'nik', 'no_anggota', 'no_hp', 'no_rekening', 'nama_bank', 'alamat', 'jenis_pinjaman', 'jenis_simpanan', 'simpanan', 'pinjaman', 'tenor_bulan') if k in header_map]
    rows_out = []
    line_no = 1
    for row in rows_src:
        line_no += 1
        nama = (row.get(header_map['nama_lengkap']) or '').strip() if 'nama_lengkap' in header_map else ''
        nik = normalize_nik(row.get(header_map['nik']))
        no_anggota = (row.get(header_map['no_anggota']) or '').strip() if 'no_anggota' in header_map else ''
        no_hp = (row.get(header_map['no_hp']) or '').strip() if 'no_hp' in header_map else ''
        no_rekening = (row.get(header_map['no_rekening']) or '').strip() if 'no_rekening' in header_map else ''
        nama_bank = (row.get(header_map['nama_bank']) or '').strip() if 'nama_bank' in header_map else ''
        alamat = (row.get(header_map['alamat']) or '').strip() if 'alamat' in header_map else ''
        jenis_pinjaman = (row.get(header_map['jenis_pinjaman']) or '').strip() if 'jenis_pinjaman' in header_map else ''
        jenis_simpanan = (row.get(header_map['jenis_simpanan']) or '').strip() if 'jenis_simpanan' in header_map else ''
        jenis_simpanan_norm = normalize_jenis_simpanan(jenis_simpanan)
        simp_raw = row.get(header_map['simpanan']) if 'simpanan' in header_map else ''
        pin_raw = row.get(header_map['pinjaman']) if 'pinjaman' in header_map else ''
        tenor_raw = row.get(header_map['tenor_bulan']) if 'tenor_bulan' in header_map else ''
        err = []
        if not nik:
            err.append('NIK wajib diisi')
        elif not is_valid_nik(nik):
            err.append('NIK harus 16 digit angka')
        if jenis_pinjaman and jenis_pinjaman not in JENIS_PINJAMAN_CHOICES:
            err.append('jenis_pinjaman tidak valid')
        if jenis_simpanan and jenis_simpanan_norm not in SIMPANAN_CHOICES and jenis_simpanan not in JENIS_SIMPANAN:
            err.append('jenis_simpanan tidak valid')
        simpanan = None
        pinjaman = None
        tenor_bulan = None
        try:
            if (simp_raw or '').strip():
                simpanan = _parse_opt_float_cell(simp_raw)
        except Exception:
            err.append('simpanan bukan angka valid')
        try:
            if (pin_raw or '').strip():
                pinjaman = _parse_opt_float_cell(pin_raw)
        except Exception:
            err.append('pinjaman bukan angka valid')
        try:
            if (tenor_raw or '').strip():
                tenor_bulan = int(float(str(tenor_raw).strip()))
                if tenor_bulan < 1 or tenor_bulan > 120:
                    err.append('tenor harus 1-120 bulan')
        except Exception:
            err.append('tenor bukan angka valid')
        rows_out.append({
            'line': line_no,
            'nama': nama,
            'nik': nik,
            'no_anggota': no_anggota,
            'no_hp': no_hp,
            'no_rekening': no_rekening,
            'nama_bank': nama_bank,
            'alamat': alamat,
            'jenis_pinjaman': jenis_pinjaman,
            'jenis_simpanan': jenis_simpanan_norm or jenis_simpanan,
            'simpanan': simpanan,
            'pinjaman': pinjaman,
            'tenor_bulan': tenor_bulan,
            'errors': err,
        })
    if not rows_out:
        raise ValueError('Tidak ada baris data (selain header).')
    if len(rows_out) > 500:
        raise ValueError('Maksimal 500 baris data per unggah.')
    return {'rows': rows_out, 'filename': file_storage.filename, 'detected_columns': detected_columns}


def _append_import_log(berhasil: int, gagal: int, mode: str, nama_file: str, catatan: str):
    # Mode DB: append langsung ke tabel import_log.
    if _is_db_mode_enabled():
        try:
            _db_insert_import_log(berhasil=berhasil, gagal=gagal, mode=mode, nama_file=nama_file, catatan=catatan)
            return
        except SQLAlchemyError as e:
            # fallback ke file jika DB bermasalah
            print(f"Error menulis import_log ke DB, fallback ke Excel: {e}")

    ensure_import_log_schema()
    rows = baca_csv(FILE_IMPORT_LOG)
    rows.append({
        'waktu': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user': session.get('user') or '',
        'mode': mode,
        'nama_file': nama_file,
        'berhasil': str(berhasil),
        'gagal': str(gagal),
        'catatan': catatan[:2000],
    })
    tulis_csv(FILE_IMPORT_LOG, rows, ['waktu', 'user', 'mode', 'nama_file', 'berhasil', 'gagal', 'catatan'])


def jalankan_import_csv_ringkasan(preview_rows: list, mode: str) -> tuple:
    """Legacy preview import (dinonaktifkan di UI)."""
    return 0, len(preview_rows or []), 'Import ringkasan tidak lagi digunakan.'


def _float_impor_csv(val) -> float:
    if val is None or str(val).strip() == '':
        return 0.0
    return float(str(val).strip().replace(',', '').replace(' ', ''))


def _append_simpanan_transaksi_import(
    simpanan_transaksi: list,
    anggota_map: dict,
    id_anggota: str,
    jumlah: float,
    jenis_simpanan: str = '',
    keterangan: str = 'Import Excel',
) -> None:
    """Catat riwayat transaksi simpanan dari proses import agar dashboard selalu sinkron."""
    if abs(float(jumlah or 0)) < 0.000001:
        return
    anggota_data = anggota_map.get(id_anggota) or {}
    jenis_norm = normalize_jenis_simpanan(jenis_simpanan or '')
    if jenis_norm not in SIMPANAN_CHOICES:
        jenis_norm = normalize_jenis_simpanan(JENIS_SIMPANAN_IMPORT)
    if jenis_norm not in SIMPANAN_CHOICES:
        jenis_norm = SIMPANAN_DEFAULT_AKUMULASI
    simpanan_transaksi.append({
        'id_transaksi': str(uuid.uuid4()),
        'id_anggota': id_anggota,
        'no_anggota': anggota_data.get('no_anggota', ''),
        'nama_anggota': anggota_data.get('nama_lengkap', ''),
        'tanggal': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'jenis_simpanan': jenis_norm,
        'jumlah': str(round(float(jumlah), 2)),
        'keterangan': keterangan,
        'diajukan_oleh': session.get('user') or '',
    })


def upsert_anggota_dari_baris_impor(
    anggota_list: list,
    id_a: str,
    nama: str,
    nik: str,
    no_hp: str,
    alamat: str,
    no_anggota: str = '',
    no_rekening: str = '',
    nama_bank: str = '',
) -> None:
    no_anggota = (no_anggota or '').strip()
    for a in anggota_list:
        if a.get('id_anggota') == id_a:
            if nama:
                a['nama_lengkap'] = nama
            if nik:
                a['nik'] = nik
            if no_hp:
                a['no_hp'] = no_hp
            if alamat:
                a['alamat'] = alamat
            if no_rekening:
                a['no_rekening'] = no_rekening
            if nama_bank:
                a['nama_bank'] = nama_bank
            if no_anggota:
                bentrok = next(
                    (
                        x for x in anggota_list
                        if x.get('id_anggota') != id_a
                        and (x.get('no_anggota') or '').strip().upper() == no_anggota.upper()
                    ),
                    None,
                )
                if not bentrok:
                    a['no_anggota'] = no_anggota
            return
    no_baru = no_anggota or generate_no_anggota_berikutnya(anggota_list)
    bentrok_baru = next(
        (x for x in anggota_list if (x.get('no_anggota') or '').strip().upper() == no_baru.upper()),
        None,
    )
    if bentrok_baru:
        no_baru = generate_no_anggota_berikutnya(anggota_list)
    anggota_list.append({
        'id_anggota': id_a,
        'no_anggota': no_baru,
        'nik': nik or '',
        'nama': nama or 'Anggota',
        'alamat': alamat or '',
        'no_telp': no_hp or '',
        'no_rekening': no_rekening or '',
        'nama_bank': nama_bank or '',
        'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
        'penghasilan_bersih': '0',
        'cicilan_lain': '0',
    })


IMPORT_CSV_FIELDNAMES = [
    'nik', 'no_hp',
]


def _pick_import_value(row: dict, *keys: str) -> str:
    """Ambil nilai kolom dengan dukungan alias header (spasi/underscore/case-insensitive)."""
    if not row:
        return ''
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() != '':
            return str(v).strip()
    norm_map = {str(k or '').strip().lower().replace(' ', '_'): k for k in row.keys()}
    for k in keys:
        nk = str(k or '').strip().lower().replace(' ', '_')
        rk = norm_map.get(nk)
        if rk is None:
            continue
        v = row.get(rk)
        if v is not None and str(v).strip() != '':
            return str(v).strip()
    return ''


def _resolve_id_anggota_by_nik_nohp(anggota_list: list, nik: str, no_hp: str) -> str:
    """Key import anggota berdasarkan kombinasi NIK + No HP."""
    nik_key = (nik or '').strip()
    hp_key = (no_hp or '').strip()
    if not nik_key or not hp_key:
        return ''
    # Kecocokan utama: NIK + No HP
    for a in anggota_list:
        if (a.get('nik') or '').strip() == nik_key and (a.get('no_hp') or '').strip() == hp_key:
            return (a.get('id_anggota') or '').strip()
    # Fallback aman: NIK sama dan nomor lama kosong -> isi nomor dari import
    for a in anggota_list:
        if (a.get('nik') or '').strip() == nik_key and not (a.get('no_hp') or '').strip():
            a['no_hp'] = hp_key
            return (a.get('id_anggota') or '').strip()
    return str(uuid.uuid4())


def _resolve_id_anggota_by_identity(anggota_list: list, nik: str, no_hp: str, no_anggota: str = '') -> str:
    """Resolve anggota by no_anggota first, then by NIK + No HP."""
    no_anggota_key = (no_anggota or '').strip()
    if no_anggota_key:
        for a in anggota_list:
            if (a.get('no_anggota') or '').strip() == no_anggota_key:
                return (a.get('id_anggota') or '').strip()
    return _resolve_id_anggota_by_nik_nohp(anggota_list, nik, no_hp)


def kategori_pinjaman_dari_tenor(tenor_bulan: int) -> str:
    """Kelompokkan pinjaman impor menjadi Jangka Pendek atau Jangka Panjang."""
    tenor = max(int(tenor_bulan or 0), 0)
    if tenor >= PROVISI_MIN_TENOR_BULAN:
        return 'Jangka Panjang'
    return 'Jangka Pendek'


@app.route('/import_csv', methods=['POST'])
@permission_required('excel.import')
@csrf_protect
def import_csv_unified():
    """Unggah satu file (CSV/Excel): merge anggota, akumulasi simpanan & pinjaman (tenor max)."""
    redir = request.referrer or url_for('halaman_anggota')
    try:
        if 'file' not in request.files:
            flash('Tidak ada berkas yang diunggah.', 'danger')
            return redirect(redir)
        f = request.files['file']
        if not f.filename or not str(f.filename).lower().endswith(('.csv', '.xlsx', '.xlsm')):
            flash('Unggah file dengan ekstensi .csv atau .xlsx.', 'danger')
            return redirect(redir)
        
        # Parse file (CSV atau Excel)
        ext = os.path.splitext(f.filename.lower())[1]
        f.stream.seek(0)
        raw = f.read()
        rows_data = []
        fieldnames = []
        
        if ext == '.csv':
            text = raw.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            rows_data = list(reader) if reader else []
            fieldnames = reader.fieldnames if reader else []
        else:
            # Excel format
            if load_workbook is None:
                flash('Fitur Excel membutuhkan openpyxl.', 'danger')
                return redirect(redir)
            try:
                wb = load_workbook(io.BytesIO(raw), data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    flash('File Excel kosong.', 'danger')
                    return redirect(redir)
                headers = [str(h or '').strip() for h in all_rows[0]]
                fieldnames = headers
                for rr in all_rows[1:]:
                    row_map = {}
                    for i, h in enumerate(headers):
                        v = rr[i] if i < len(rr) else ''
                        row_map[h] = '' if v is None else str(v).strip()
                    rows_data.append(row_map)
            except Exception as e:
                flash(f'Gagal membaca Excel: {str(e)}', 'danger')
                return redirect(redir)
        
        if not fieldnames:
            flash('File tidak memiliki header.', 'danger')
            return redirect(redir)
        
        header_norm = {str(h or '').strip().lower().replace(' ', '_') for h in fieldnames if h}
        has_nik = 'nik' in header_norm
        has_nohp = ('no_hp' in header_norm) or ('no_telp' in header_norm)
        if not has_nik or not has_nohp:
            flash('Kolom wajib tidak lengkap: nik dan no_hp/no_telp.', 'danger')
            return redirect(redir)

        invalid_count = 0
        for row in rows_data:
            nik_chk = _pick_import_value(row, 'nik')
            no_hp_chk = _pick_import_value(row, 'no_hp', 'no_telp', 'no hp', 'no telp')
            if not nik_chk or not no_hp_chk:
                invalid_count += 1
                continue
            try:
                v_simp_chk = _float_impor_csv(_pick_import_value(row, 'simpanan'))
                v_pin_chk = _float_impor_csv(_pick_import_value(row, 'pinjaman'))
                v_ten_chk = int(float(str(_pick_import_value(row, 'tenor', 'tenor_bulan') or '0').strip()))
            except (TypeError, ValueError):
                invalid_count += 1
                continue
            if v_simp_chk < 0 or v_pin_chk < 0 or v_ten_chk < 0:
                invalid_count += 1
                continue
            jp_chk = _pick_import_value(row, 'jenis_pinjaman')
            if jp_chk and jp_chk not in JENIS_PINJAMAN_CHOICES:
                invalid_count += 1
                continue
            js_raw_chk = _pick_import_value(row, 'jenis_simpanan')
            if js_raw_chk:
                js_norm_chk = normalize_jenis_simpanan(js_raw_chk)
                if js_norm_chk not in SIMPANAN_CHOICES and js_raw_chk not in JENIS_SIMPANAN:
                    invalid_count += 1
                    continue

        if invalid_count > 0:
            flash(f'Import dibatalkan: ditemukan {invalid_count} baris tidak valid.', 'danger')
            return redirect(redir)

        ensure_anggota_schema()
        ensure_simpanan_schema()
        ensure_simpanan_transaksi_schema()
        ensure_pinjaman_plafon_schema()
        anggota = baca_csv(FILE_ANGGOTA)
        simpanan = baca_csv(FILE_SIMPANAN)
        simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
        pinjaman = baca_csv(FILE_PINJAMAN)
        anggota_map = {a.get('id_anggota'): a for a in anggota}

        diproses = 0
        dilewati = 0
        
        for row in rows_data:
            nik = _pick_import_value(row, 'nik')
            no_hp = _pick_import_value(row, 'no_hp', 'no_telp', 'no hp', 'no telp')
            if not nik or not no_hp:
                dilewati += 1
                continue
            no_anggota = _pick_import_value(row, 'no_anggota', 'no anggota')
            id_a = _resolve_id_anggota_by_identity(anggota, nik, no_hp, no_anggota)
            if not id_a:
                dilewati += 1
                continue
            try:
                v_simp = _float_impor_csv(_pick_import_value(row, 'simpanan'))
                v_pin = _float_impor_csv(_pick_import_value(row, 'pinjaman'))
                v_ten = int(float(str(_pick_import_value(row, 'tenor', 'tenor_bulan') or '0').strip()))
            except (TypeError, ValueError):
                dilewati += 1
                continue
            if v_simp < 0 or v_pin < 0 or v_ten < 0:
                dilewati += 1
                continue

            nama = _pick_import_value(row, 'nama')
            alamat = _pick_import_value(row, 'alamat')
            no_rekening = _pick_import_value(row, 'no_rekening', 'norek', 'rekening')
            nama_bank = _pick_import_value(row, 'nama_bank', 'bank')
            jenis_simpanan = normalize_jenis_simpanan(_pick_import_value(row, 'jenis_simpanan'))
            upsert_anggota_dari_baris_impor(
                anggota,
                id_a,
                nama,
                nik,
                no_hp,
                alamat,
                no_anggota,
                no_rekening,
                nama_bank,
            )
            anggota_map = {a.get('id_anggota'): a for a in anggota}
            if v_simp > 0:
                merge_akumulasi(simpanan, id_a, v_simp)
                _append_simpanan_transaksi_import(
                    simpanan_transaksi,
                    anggota_map,
                    id_a,
                    v_simp,
                    jenis_simpanan,
                    'Setoran via import file',
                )
            if v_pin > 0 or v_ten > 0:
                merge_pinjaman_akumulasi(
                    pinjaman,
                    id_a,
                    v_pin,
                    v_ten,
                    kategori_pinjaman_dari_tenor(v_ten),
                )
            diproses += 1

        simpanan[:] = _dedupe_rows_simpanan(simpanan)
        pinjaman[:] = _dedupe_rows_pinjaman(pinjaman)
        tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
        tulis_csv(FILE_SIMPANAN, simpanan, SIMPANAN_FIELDNAMES)
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, simpanan_transaksi, SIMPANAN_TRANSAKSI_FIELDNAMES)
        amap_fin = {a.get('id_anggota'): a for a in anggota}
        for p in pinjaman:
            ag = amap_fin.get(p.get('id_anggota'))
            if ag:
                p['nama_anggota'] = ag.get('nama_lengkap', '')
                p['no_anggota'] = ag.get('no_anggota', '')
        tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
        flash(f'Import selesai: {diproses} baris berhasil diproses.', 'success')
    except Exception as ex:
        flash(f'Import gagal: {ex}', 'danger')
    return redirect(redir)


@app.route('/anggota/import-csv', methods=['GET'])
@admin_required
@permission_required('excel.import')
def halaman_import_csv_anggota():
    # Mode DB: ambil langsung 50 terbaru dari PostgreSQL (lebih efisien).
    if _is_db_mode_enabled():
        logs = _db_select_import_log_latest(limit=50)
    else:
        ensure_import_log_schema()
        logs = baca_csv(FILE_IMPORT_LOG)
        logs = sorted(logs, key=lambda x: x.get('waktu', ''), reverse=True)[:50]
    return render_template('import_csv_anggota.html', import_log=logs)


@app.route('/anggota/import-csv/sample')
@admin_required
@permission_required('excel.import')
def download_sample_csv_anggota():
    if Workbook is None:
        flash('Fitur unduh contoh Excel membutuhkan openpyxl.', 'danger')
        return redirect(url_for('halaman_import_csv_anggota'))

    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Import'
    ws.append([
        'nama', 'nik', 'no_anggota', 'no_hp', 'alamat',
        'no_rekening', 'nama_bank', 'jenis_pinjaman', 'jenis_simpanan',
        'simpanan', 'pinjaman', 'tenor_bulan'
    ])
    ws.append([
        'Budi Santoso', '3174120101010001', 'AG0001', '081234567890', 'Jakarta',
        '1234567890', 'BRI', 'Jangka Panjang', 'Manasuka', '1500000', '40000000', '40'
    ])
    ws.append([
        'Siti Aulia', '3273010202020002', 'AG0002', '081298765432', 'Bandung',
        '9876543210', 'BCA', 'Jangka Pendek', 'Pendidikan', '500000', '0', ''
    ])

    pinjaman_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(JENIS_PINJAMAN_CHOICES) + '"',
        allow_blank=True,
    )
    simpanan_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(JENIS_SIMPANAN) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(pinjaman_dv)
    ws.add_data_validation(simpanan_dv)
    pinjaman_dv.add('H2:H500')
    simpanan_dv.add('I2:I500')
    ws.freeze_panes = 'A2'

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name='template_import_gabungan.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/anggota/import-csv/preview', methods=['POST'])
@admin_required
@permission_required('excel.import')
@csrf_protect
def preview_import_csv_anggota():
    try:
        if 'file_csv' not in request.files:
            flash('Tidak ada file yang dipilih.', 'danger')
            return redirect(url_for('halaman_import_csv_anggota'))
        payload = parse_anggota_csv_upload(request.files['file_csv'])
        token = token_hex(16)
        preview_path = os.path.join(IMPORT_PREVIEW_DIR, f'{token}.json')
        with open(preview_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False)
        return render_template(
            'import_csv_preview.html',
            token=token,
            filename=payload.get('filename', ''),
            rows=payload.get('rows', []),
            detected_columns=payload.get('detected_columns', []),
            has_errors=any((r.get('errors') or []) for r in (payload.get('rows', []) or [])),
        )
    except ValueError as ex:
        flash(str(ex), 'danger')
        return redirect(url_for('halaman_import_csv_anggota'))
    except Exception as ex:
        flash(f'Gagal membuat preview import: {ex}', 'danger')
        return redirect(url_for('halaman_import_csv_anggota'))


@app.route('/anggota/import-csv/preview', methods=['GET'])
@admin_required
@permission_required('excel.import')
def tampil_preview_import_csv():
    token = (request.args.get('token') or '').strip()
    if not token:
        flash('Token preview tidak ditemukan.', 'warning')
        return redirect(url_for('halaman_import_csv_anggota'))
    preview_path = os.path.join(IMPORT_PREVIEW_DIR, f'{token}.json')
    if not os.path.exists(preview_path):
        flash('Preview tidak ditemukan atau sudah kedaluwarsa.', 'warning')
        return redirect(url_for('halaman_import_csv_anggota'))
    with open(preview_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)
    return render_template(
        'import_csv_preview.html',
        token=token,
        filename=payload.get('filename', ''),
        rows=payload.get('rows', []),
        detected_columns=payload.get('detected_columns', []),
        has_errors=any((r.get('errors') or []) for r in (payload.get('rows', []) or [])),
    )


@app.route('/anggota/import-csv/execute', methods=['POST'])
@admin_required
@permission_required('excel.import')
@csrf_protect
def execute_import_csv_anggota():
    token = (request.form.get('token') or '').strip()
    mode = (request.form.get('mode') or 'append').strip().lower()
    if mode not in ('append', 'overwrite'):
        mode = 'append'
    if not token:
        flash('Token preview tidak valid.', 'danger')
        return redirect(url_for('halaman_import_csv_anggota'))

    preview_path = os.path.join(IMPORT_PREVIEW_DIR, f'{token}.json')
    if not os.path.exists(preview_path):
        flash('Preview tidak ditemukan atau sudah kedaluwarsa.', 'warning')
        return redirect(url_for('halaman_import_csv_anggota'))

    try:
        with open(preview_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        rows = payload.get('rows', []) or []
        filename = payload.get('filename', 'import.xlsx')

        ensure_anggota_schema()
        ensure_simpanan_schema()
        ensure_simpanan_transaksi_schema()
        ensure_pinjaman_plafon_schema()

        anggota = baca_csv(FILE_ANGGOTA)
        simpanan = baca_csv(FILE_SIMPANAN)
        simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
        pinjaman = baca_csv(FILE_PINJAMAN)

        invalid_rows = [r for r in rows if r.get('errors')]
        if invalid_rows:
            flash(f'Import dibatalkan: ditemukan {len(invalid_rows)} baris tidak valid. Perbaiki file lalu upload ulang.', 'danger')
            return redirect(url_for('tampil_preview_import_csv', token=token))

        berhasil = 0
        gagal = 0

        for r in rows:
            if r.get('errors'):
                gagal += 1
                continue

            nik = (r.get('nik') or '').strip()
            no_hp = (r.get('no_hp') or '').strip()
            no_anggota = (r.get('no_anggota') or '').strip()
            jenis_pinjaman = (r.get('jenis_pinjaman') or '').strip()
            jenis_simpanan = normalize_jenis_simpanan((r.get('jenis_simpanan') or '').strip())
            nama = (r.get('nama_lengkap') or '').strip()
            alamat = (r.get('alamat') or '').strip()

            id_a = _resolve_id_anggota_by_identity(anggota, nik, no_hp, no_anggota)
            if not id_a and nik:
                found_by_nik = next((a for a in anggota if (a.get('nik') or '').strip() == nik), None)
                if found_by_nik:
                    id_a = (found_by_nik.get('id_anggota') or '').strip()
            if not id_a:
                id_a = str(uuid.uuid4())

            upsert_anggota_dari_baris_impor(anggota, id_a, nama, nik, no_hp, alamat, no_anggota)

            try:
                v_simp = float(r.get('simpanan') or 0)
            except (TypeError, ValueError):
                v_simp = 0.0
            try:
                v_pin = float(r.get('pinjaman') or 0)
            except (TypeError, ValueError):
                v_pin = 0.0
            try:
                v_ten = int(float(r.get('tenor_bulan') or 0))
            except (TypeError, ValueError):
                v_ten = 0

            if v_simp < 0 or v_pin < 0 or v_ten < 0:
                gagal += 1
                continue

            if mode == 'overwrite':
                if v_simp > 0:
                    target = next((s for s in simpanan if s.get('id_anggota') == id_a), None)
                    prev_saldo = 0.0
                    if target:
                        try:
                            prev_saldo = float(target.get('total_simpanan') or 0)
                        except (TypeError, ValueError):
                            prev_saldo = 0.0
                    if target:
                        target['total_simpanan'] = str(round(v_simp, 2))
                    else:
                        simpanan.append({'id_anggota': id_a, 'total_simpanan': str(round(v_simp, 2))})
                    delta = round(v_simp - prev_saldo, 2)
                    if abs(delta) > 0:
                        _append_simpanan_transaksi_import(
                            simpanan_transaksi,
                            {a.get('id_anggota'): a for a in anggota},
                            id_a,
                            delta,
                            jenis_simpanan,
                            'Penyesuaian saldo via import Excel (overwrite)',
                        )
                if v_pin > 0 or v_ten > 0:
                    tenor_input = v_ten if v_ten > 0 else DEFAULT_TENOR_IMPORT_PINJAMAN
                    jenis = jenis_pinjaman if jenis_pinjaman in JENIS_PINJAMAN_CHOICES else kategori_pinjaman_dari_tenor(tenor_input)
                    p_exist = next(
                        (
                            p for p in pinjaman
                            if p.get('id_anggota') == id_a
                            and (p.get('jenis_pinjaman') or '').strip() == jenis
                            and (p.get('status') or '').strip() == 'Disetujui'
                        ),
                        None,
                    )
                    if p_exist:
                        try:
                            old_plaf = float(p_exist.get('plafon') or 0)
                            old_sisa = saldo_pinjaman_aktual(p_exist)
                        except (TypeError, ValueError):
                            old_plaf = 0.0
                            old_sisa = 0.0
                        paid_nominal = max(old_plaf - old_sisa, 0.0)
                        tenor_data = int(float(p_exist.get('tenor_bulan') or p_exist.get('tenor_awal') or tenor_input))
                        tenor_data = max(tenor_data, 1)
                        bunga = float(p_exist.get('bunga_persen') or bunga_untuk_jenis_pinjaman(jenis, tenor_data))
                        total_bayar = hitung_total_bayar_tanpa_provisi(v_pin, bunga, tenor_data, jenis)
                        cic = hitung_cicilan_bulanan(v_pin, bunga, tenor_data)
                        p_exist['plafon'] = str(round(v_pin, 2))
                        p_exist['tenor_awal'] = p_exist.get('tenor_awal') or str(tenor_data)
                        p_exist['tenor_bulan'] = str(tenor_data)
                        p_exist['total_bayar'] = str(round(total_bayar, 2))
                        p_exist['cicilan_per_bulan'] = str(round(cic, 2))
                        p_exist['sisa_pinjaman'] = str(round(max(v_pin - paid_nominal, 0.0), 2))
                    else:
                        merge_pinjaman_akumulasi(pinjaman, id_a, v_pin, tenor_input, jenis)
            else:
                if v_simp > 0:
                    merge_akumulasi(simpanan, id_a, v_simp)
                    _append_simpanan_transaksi_import(
                        simpanan_transaksi,
                        {a.get('id_anggota'): a for a in anggota},
                        id_a,
                        v_simp,
                        jenis_simpanan,
                        'Setoran via import Excel (append)',
                    )
                if v_pin > 0 or v_ten > 0:
                    tenor_input = v_ten if v_ten > 0 else DEFAULT_TENOR_IMPORT_PINJAMAN
                    jenis = jenis_pinjaman if jenis_pinjaman in JENIS_PINJAMAN_CHOICES else kategori_pinjaman_dari_tenor(tenor_input)
                    merge_pinjaman_akumulasi(
                        pinjaman,
                        id_a,
                        v_pin,
                        tenor_input,
                        jenis,
                    )

            berhasil += 1

        simpanan[:] = _dedupe_rows_simpanan(simpanan)
        pinjaman[:] = _dedupe_rows_pinjaman(pinjaman)
        tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
        tulis_csv(FILE_SIMPANAN, simpanan, SIMPANAN_FIELDNAMES)
        tulis_csv(FILE_SIMPANAN_TRANSAKSI, simpanan_transaksi, SIMPANAN_TRANSAKSI_FIELDNAMES)

        amap_fin = {a.get('id_anggota'): a for a in anggota}
        for p in pinjaman:
            ag = amap_fin.get(p.get('id_anggota'))
            if ag:
                p['nama_anggota'] = ag.get('nama_lengkap', '')
                p['no_anggota'] = ag.get('no_anggota', '')
        tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)

        _append_import_log(
            berhasil=berhasil,
            gagal=gagal,
            mode=f'preview-{mode}',
            nama_file=filename,
            catatan='Import gabungan via preview Excel',
        )

        try:
            os.remove(preview_path)
        except OSError:
            pass

        flash(f'Import selesai: {berhasil} baris berhasil, {gagal} baris gagal/dilewati.', 'success')
        return redirect(url_for('halaman_anggota'))
    except Exception as ex:
        flash(f'Eksekusi import gagal: {ex}', 'danger')
        return redirect(url_for('halaman_import_csv_anggota'))


@app.route('/anggota/cari_nama', methods=['GET'])
@login_required
def cari_nama():
    """Autocomplete nama anggota berdasarkan No Anggota."""
    no_anggota = request.args.get('no_anggota', '')
    data = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        data = [a for a in data if a.get('id_anggota') == id_anggota]
    for a in data:
        if a['no_anggota'] == no_anggota:
            return jsonify({'nama': a.get('nama_lengkap', ''), 'id_anggota': a['id_anggota']})
    return jsonify({'nama': '', 'id_anggota': ''})


@app.route('/anggota/cari_anggota', methods=['GET'])
@login_required
def cari_anggota_autocomplete():
    """Cari anggota berdasarkan nama (autocomplete)."""
    query = request.args.get('q', '').lower()
    data = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        data = [a for a in data if a.get('id_anggota') == id_anggota]
    hasil = [a for a in data if query in (a.get('nama_lengkap') or '').lower()]
    return jsonify(hasil)


# ══════════════════════════════════════════════
#  ROUTE: SIMPANAN
# ══════════════════════════════════════════════
@app.route('/simpanan')
@login_required
@permission_required('savings.deposit.input', 'savings.deposit.request', 'savings.deposit.validate', 'savings.withdraw.validate')
def halaman_simpanan():
    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    ensure_simpanan_pengajuan_schema()
    ensure_iuran_sosial_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    pengajuan_simpanan = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    iuran_sosial = baca_csv(FILE_IURAN_SOSIAL)
    anggota = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        simpanan = [s for s in simpanan if s.get('id_anggota') == id_anggota]
        simpanan_transaksi = [t for t in simpanan_transaksi if t.get('id_anggota') == id_anggota]
        pengajuan_simpanan = [p for p in pengajuan_simpanan if p.get('id_anggota') == id_anggota]
        iuran_sosial = [i for i in iuran_sosial if i.get('id_anggota') == id_anggota]
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]

    simpanan_tampil = enrich_simpanan_untuk_tampilan(simpanan, anggota)
    simpanan_tampil.reverse()

    saldo_per_jenis_map = _saldo_per_jenis_akumulasi(simpanan_transaksi)

    for s in simpanan_tampil:
        id_a = s.get('id_anggota', '')
        by_jenis = saldo_per_jenis_map.get(id_a, {k: 0.0 for k in SIMPANAN_AKUMULASI_CHOICES})
        saldo_pokok = float(by_jenis.get('Simpanan Pokok', 0.0))
        saldo_wajib = float(by_jenis.get('Simpanan Wajib', 0.0))
        saldo_hari_koperasi = float(by_jenis.get('Simpanan Hari Koperasi', by_jenis.get('Hari Koperasi', 0.0)))
        saldo_pensiun = float(by_jenis.get('Simpanan Pensiun', by_jenis.get('Simpan Pensiun', by_jenis.get('Pensiun', 0.0))))
        saldo_hari_raya = float(by_jenis.get('Simpanan Hari Raya', 0.0))
        saldo_manasuka = float(by_jenis.get('Simpanan Manasuka', 0.0))
        saldo_pendidikan = float(by_jenis.get('Simpanan Pendidikan', 0.0))
        # Jika ada data lama (total tanpa rincian jenis), alokasikan selisih ke Simpanan Wajib.
        try:
            total_saldo = float(s.get('jumlah') or 0)
        except (TypeError, ValueError):
            total_saldo = 0.0
        subtotal_jenis = (
            saldo_pokok + saldo_wajib + saldo_hari_koperasi + saldo_pensiun
            + saldo_hari_raya + saldo_manasuka + saldo_pendidikan
        )
        if total_saldo > subtotal_jenis:
            saldo_wajib += (total_saldo - subtotal_jenis)
        s['saldo_pokok'] = round(saldo_pokok, 2)
        s['saldo_wajib'] = round(saldo_wajib, 2)
        s['saldo_hari_koperasi'] = round(saldo_hari_koperasi, 2)
        s['saldo_pensiun'] = round(saldo_pensiun, 2)
        s['saldo_hari_raya'] = round(saldo_hari_raya, 2)
        s['saldo_manasuka'] = round(saldo_manasuka, 2)
        s['saldo_pendidikan'] = round(saldo_pendidikan, 2)

    saldo_anggota = {}
    for a in anggota:
        saldo_anggota[a['id_anggota']] = 0.0
    for s in simpanan:
        if s['id_anggota'] in saldo_anggota:
            saldo_anggota[s['id_anggota']] += float(s.get('total_simpanan') or 0)

    transaksi_tampil = sorted(simpanan_transaksi, key=lambda x: x.get('tanggal', ''), reverse=True)
    catatan_iuran_sosial = sorted(iuran_sosial, key=lambda x: x.get('tanggal', ''), reverse=True)

    return render_template(
        'simpanan.html',
        simpanan=simpanan_tampil,
        simpanan_transaksi=transaksi_tampil,
        pengajuan_simpanan=sorted(pengajuan_simpanan, key=lambda x: x.get('tanggal_pengajuan', ''), reverse=True),
        catatan_iuran_sosial=catatan_iuran_sosial,
        anggota=anggota,
        saldo_anggota=saldo_anggota,
        jenis_simpanan_choices=[j for j in SIMPANAN_CHOICES if j != 'Simpanan Pokok'],
    )


@app.route('/simpanan/tambah', methods=['POST'])
@permission_required('savings.deposit.input', 'savings.deposit.request')
@csrf_protect
def tambah_simpanan():
    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    ensure_simpanan_pengajuan_schema()
    ensure_iuran_sosial_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    pengajuan_simpanan = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    iuran_sosial = baca_csv(FILE_IURAN_SOSIAL)
    if is_current_user_admin():
        id_anggota = request.form['id_anggota']
    else:
        id_anggota = get_current_user_id_anggota()
        if not id_anggota:
            abort(403)
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_data = next((a for a in anggota if a['id_anggota'] == id_anggota), None)

    if not anggota_data:
        return "Anggota tidak ditemukan", 400

    try:
        jumlah = float(request.form['jumlah'].replace(',', '').replace('.', ''))
    except ValueError:
        flash('Jumlah simpanan tidak valid.', 'danger')
        return redirect('/simpanan')
    if jumlah < 0:
        flash('Jumlah simpanan tidak boleh negatif.', 'danger')
        return redirect('/simpanan')

    jenis_simpanan = normalize_jenis_simpanan(request.form.get('jenis_simpanan'))
    if jenis_simpanan == 'Simpanan Pokok':
        flash('Simpanan Pokok dicatat saat validasi anggota baru, bukan dari menu simpanan.', 'danger')
        return redirect('/simpanan')
    if jenis_simpanan not in SIMPANAN_RULES:
        flash('Jenis simpanan tidak valid.', 'danger')
        return redirect('/simpanan')

    rules = SIMPANAN_RULES[jenis_simpanan]
    nominal_fix = rules.get('fixed')
    nominal_min = rules.get('min')
    nominal_max = rules.get('max')
    if nominal_fix is not None and jumlah != float(nominal_fix):
        flash(f"Nominal {jenis_simpanan} wajib tepat Rp {int(nominal_fix):,}.".replace(',', '.'), 'danger')
        return redirect('/simpanan')
    if nominal_min is not None and jumlah < float(nominal_min):
        flash(f"Nominal {jenis_simpanan} minimal Rp {int(nominal_min):,}.".replace(',', '.'), 'danger')
        return redirect('/simpanan')
    if nominal_max is not None and jumlah > float(nominal_max):
        flash(f"Nominal {jenis_simpanan} maksimal Rp {int(nominal_max):,}.".replace(',', '.'), 'danger')
        return redirect('/simpanan')

    riwayat_jenis = iuran_sosial if jenis_simpanan == 'Iuran Sosial' else simpanan_transaksi

    if rules.get('one_time') and (
        _jenis_sudah_tercatat_anggota(riwayat_jenis, id_anggota, jenis_simpanan)
        or _jenis_simpanan_menunggu_pengajuan(pengajuan_simpanan, id_anggota, jenis_simpanan)
    ):
        flash(f'{jenis_simpanan} hanya boleh dibayar sekali per anggota.', 'danger')
        return redirect('/simpanan')

    month_key = datetime.now().strftime('%Y-%m')
    jenis_default = 'Iuran Sosial' if jenis_simpanan == 'Iuran Sosial' else ''
    if rules.get('monthly') and (
        _jenis_sudah_dibayar_bulan_ini(
            riwayat_jenis,
            id_anggota,
            jenis_simpanan,
            month_key,
            jenis_default,
        ) or _jenis_simpanan_menunggu_pengajuan(
            pengajuan_simpanan,
            id_anggota,
            jenis_simpanan,
            month_key,
            jenis_default,
        )
    ):
        flash(f'{jenis_simpanan} bulan ini sudah tercatat untuk anggota tersebut.', 'warning')
        return redirect('/simpanan')

    if jenis_simpanan == 'Iuran Sosial':
        iuran_sosial.append({
            'id_iuran': str(uuid.uuid4()),
            'id_anggota': id_anggota,
            'no_anggota': anggota_data.get('no_anggota', ''),
            'nama_anggota': anggota_data.get('nama_lengkap', ''),
            'tanggal': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'jumlah': str(round(jumlah, 2)),
            'keterangan': 'Catatan iuran sosial (tidak menambah saldo simpanan)',
            'diajukan_oleh': session.get('user') or '',
        })
        tulis_csv(FILE_IURAN_SOSIAL, iuran_sosial, IURAN_SOSIAL_FIELDNAMES)
        flash('Iuran sosial berhasil dicatat dan tidak menambah saldo simpanan.', 'success')
    else:
        pengajuan_simpanan.append({
            'id_pengajuan': str(uuid.uuid4()),
            'id_anggota': id_anggota,
            'no_anggota': anggota_data.get('no_anggota', ''),
            'nama_anggota': anggota_data.get('nama_lengkap', ''),
            'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'jenis_simpanan': jenis_simpanan,
            'jumlah': str(round(jumlah, 2)),
            'keterangan': f'Pengajuan {jenis_simpanan} menunggu konfirmasi admin',
            'status': 'Menunggu',
            'tanggal_konfirmasi': '',
            'dikonfirmasi_oleh': '',
            'diajukan_oleh': session.get('user') or '',
        })
        tulis_csv(FILE_SIMPANAN_PENGAJUAN, pengajuan_simpanan, SIMPANAN_PENGAJUAN_FIELDNAMES)
        flash(f'Pengajuan {jenis_simpanan} berhasil dikirim dan menunggu konfirmasi admin.', 'success')
    return redirect('/simpanan')


@app.route('/simpanan/konfirmasi/<id_simpanan>', methods=['POST'])
@permission_required('savings.deposit.validate')
@csrf_protect
def konfirmasi_simpanan(id_simpanan):
    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    ensure_simpanan_pengajuan_schema()
    pengajuan_simpanan = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    simpanan = baca_csv(FILE_SIMPANAN)
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)

    ditemukan = None
    for item in pengajuan_simpanan:
        if item.get('id_pengajuan') == id_simpanan:
            ditemukan = item
            break

    if not ditemukan:
        flash('Pengajuan simpanan tidak ditemukan.', 'danger')
        return redirect('/simpanan')

    if (ditemukan.get('status') or '').strip() != 'Menunggu':
        flash('Pengajuan simpanan sudah diproses sebelumnya.', 'warning')
        return redirect('/simpanan')

    try:
        jumlah = float(ditemukan.get('jumlah') or 0)
    except (TypeError, ValueError):
        flash('Nominal pengajuan simpanan tidak valid.', 'danger')
        return redirect('/simpanan')

    id_anggota = ditemukan.get('id_anggota') or ''
    jenis_simpanan = normalize_jenis_simpanan(ditemukan.get('jenis_simpanan'))
    merge_akumulasi(simpanan, id_anggota, jumlah)
    tulis_csv(FILE_SIMPANAN, simpanan, SIMPANAN_FIELDNAMES)

    simpanan_transaksi.append({
        'id_transaksi': str(uuid.uuid4()),
        'id_anggota': id_anggota,
        'no_anggota': ditemukan.get('no_anggota', ''),
        'nama_anggota': ditemukan.get('nama_anggota', ''),
        'tanggal': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'jenis_simpanan': jenis_simpanan,
        'jumlah': str(round(jumlah, 2)),
        'keterangan': 'Simpanan disetujui admin',
        'diajukan_oleh': ditemukan.get('diajukan_oleh', ''),
    })
    tulis_csv(FILE_SIMPANAN_TRANSAKSI, simpanan_transaksi, SIMPANAN_TRANSAKSI_FIELDNAMES)

    ditemukan['status'] = 'Disetujui'
    ditemukan['tanggal_konfirmasi'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ditemukan['dikonfirmasi_oleh'] = session.get('user') or ''
    tulis_csv(FILE_SIMPANAN_PENGAJUAN, pengajuan_simpanan, SIMPANAN_PENGAJUAN_FIELDNAMES)

    flash('Pengajuan simpanan berhasil disetujui dan saldo sudah diperbarui.', 'success')
    return redirect('/simpanan')


@app.route('/simpanan/tolak/<id_simpanan>', methods=['POST'])
@permission_required('savings.deposit.validate')
@csrf_protect
def tolak_simpanan(id_simpanan):
    ensure_simpanan_pengajuan_schema()
    pengajuan_simpanan = baca_csv(FILE_SIMPANAN_PENGAJUAN)
    ditemukan = None
    for item in pengajuan_simpanan:
        if item.get('id_pengajuan') == id_simpanan:
            ditemukan = item
            break

    if not ditemukan:
        flash('Pengajuan simpanan tidak ditemukan.', 'danger')
        return redirect('/simpanan')

    if (ditemukan.get('status') or '').strip() != 'Menunggu':
        flash('Pengajuan simpanan sudah diproses sebelumnya.', 'warning')
        return redirect('/simpanan')

    ditemukan['status'] = 'Ditolak'
    ditemukan['tanggal_konfirmasi'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ditemukan['dikonfirmasi_oleh'] = session.get('user') or ''
    tulis_csv(FILE_SIMPANAN_PENGAJUAN, pengajuan_simpanan, SIMPANAN_PENGAJUAN_FIELDNAMES)

    flash('Pengajuan simpanan ditolak.', 'warning')
    return redirect('/simpanan')


@app.route('/simpanan/hapus/<id_anggota>', methods=['POST'])
@permission_required('savings.deposit.validate')
@csrf_protect
def hapus_simpanan(id_anggota):
    data = [s for s in baca_csv(FILE_SIMPANAN) if s.get('id_anggota') != id_anggota]
    tulis_csv(FILE_SIMPANAN, data, SIMPANAN_FIELDNAMES)
    flash('Saldo simpanan anggota dihapus dari daftar.', 'success')
    return redirect('/simpanan')


@app.route('/simpanan/saldo/<id_anggota>')
@login_required
def cek_saldo(id_anggota):
    """Cek total saldo simpanan anggota."""
    restrict_id_anggota_or_forbid(id_anggota)
    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    simpanan_transaksi = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    total = 0.0
    for s in simpanan:
        if s.get('id_anggota') == id_anggota:
            total = float(s.get('total_simpanan') or 0)
            break
    per_jenis = {k: 0.0 for k in SIMPANAN_AKUMULASI_CHOICES}
    per_jenis_map = _saldo_per_jenis_akumulasi(simpanan_transaksi)
    if id_anggota in per_jenis_map:
        per_jenis.update(per_jenis_map[id_anggota])
    subtotal = sum(per_jenis.values())
    if total > subtotal:
        per_jenis[SIMPANAN_DEFAULT_AKUMULASI] += (total - subtotal)

    jenis_req = normalize_jenis_simpanan(request.args.get('jenis_simpanan') or SIMPANAN_DEFAULT_AKUMULASI)
    jenis_req = (jenis_req or SIMPANAN_DEFAULT_AKUMULASI).strip() or SIMPANAN_DEFAULT_AKUMULASI
    if jenis_req not in per_jenis:
        jenis_req = SIMPANAN_DEFAULT_AKUMULASI

    return jsonify({
        'saldo': total,
        'saldo_per_jenis': {k: round(v, 2) for k, v in per_jenis.items()},
        'jenis_simpanan': jenis_req,
        'saldo_jenis': round(per_jenis.get(jenis_req, 0.0), 2),
    })


# ══════════════════════════════════════════════
#  ROUTE: PINJAMAN
# ══════════════════════════════════════════════
@app.route('/pinjaman')
@login_required
@permission_required('loan.documents.review', 'loan.eligibility.analyze', 'loans.approve', 'loan.request', 'loan.disbursement.input')
def halaman_pinjaman():
    ensure_pinjaman_plafon_schema()
    pinjaman = baca_csv(FILE_PINJAMAN)
    anggota_full = baca_csv(FILE_ANGGOTA)
    cicilan_menunggu = []
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        pinjaman = [p for p in pinjaman if p.get('id_anggota') == id_anggota]
        anggota = [a for a in anggota_full if a.get('id_anggota') == id_anggota]
    else:
        anggota = anggota_full
        if _mark_expired_cicilan(cicilan):
            tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
        cicilan_menunggu = [c for c in cicilan if c.get('status') == 'Menunggu']

    riwayat_map = {}
    for p in pinjaman:
        id_a = p.get('id_anggota', '')
        if not id_a:
            continue
        tenor_raw = p.get('tenor_bulan', p.get('tenor', '0'))
        try:
            tenor_norm = int(float(tenor_raw or 0))
        except (TypeError, ValueError):
            tenor_norm = 0
        jenis = (p.get('jenis_pinjaman') or '').strip()
        if not jenis or jenis == JENIS_IMPORT_CSV:
            jenis = kategori_pinjaman_dari_tenor(tenor_norm)
        if not jenis:
            continue

        row = riwayat_map.setdefault(id_a, {
            'id_anggota': id_a,
            'no_anggota': p.get('no_anggota', ''),
            'nama_anggota': p.get('nama_anggota', ''),
            'total_pengajuan': 0,
            'total_aktif': 0,
            'jenis_list': [],
        })

        if not row['no_anggota']:
            row['no_anggota'] = p.get('no_anggota', '')
        if not row['nama_anggota']:
            row['nama_anggota'] = p.get('nama_anggota', '')

        row['total_pengajuan'] += 1
        if (p.get('status') or '').strip() == 'Disetujui' and saldo_pinjaman_aktual(p) > 0:
            row['total_aktif'] += 1
        if jenis not in row['jenis_list']:
            row['jenis_list'].append(jenis)

    anggota_map = {a.get('id_anggota'): a for a in anggota_full}
    riwayat_pinjaman_anggota = []
    for item in riwayat_map.values():
        a = anggota_map.get(item['id_anggota'], {})
        no_anggota = item['no_anggota'] or a.get('no_anggota', '-')
        nama_anggota = item['nama_anggota'] or a.get('nama_lengkap', '-')
        jenis_items = item['jenis_list']
        riwayat_pinjaman_anggota.append({
            'id_anggota': item['id_anggota'],
            'no_anggota': no_anggota,
            'nama_anggota': nama_anggota,
            'total_pengajuan': item['total_pengajuan'],
            'total_aktif': item['total_aktif'],
            'jenis_items': jenis_items,
            'jenis_text': ', '.join(jenis_items),
        })

    riwayat_pinjaman_anggota.sort(key=lambda x: (x.get('nama_anggota') or '').lower())

    pinjaman_tampil = enrich_pinjaman_untuk_tampilan(pinjaman, anggota_full)
    periode_gagal_bayar = _geser_bulan_tanggal(datetime.now().date(), -1) or datetime.now().date()
    for p in pinjaman_tampil:
        p['gagal_bayar_bulan_ini'] = _sudah_gagal_bayar_bulan_periode(
            cicilan,
            p.get('id_pinjaman', ''),
            periode_gagal_bayar,
        )
    pinjaman_tampil.reverse()
    return render_template(
        'pinjaman.html',
        pinjaman=pinjaman_tampil,
        riwayat_pinjaman_anggota=riwayat_pinjaman_anggota,
        anggota=anggota,
        cicilan_menunggu=cicilan_menunggu,
        metode_bayar_choices=METODE_BAYAR_CHOICES,
        koperasi_rekening=KOPERASI_REKENING_BANK,
        qris_dana_target=(KOPERASI_REKENING_BANK.get('no_rekening', '') or ''),
        jenis_pinjaman_choices=[j for j in JENIS_PINJAMAN_CHOICES if j in JENIS_PINJAMAN],
    )


@app.route('/pinjaman/hitung', methods=['GET'])
@login_required
@permission_required('loan.request', 'loan.eligibility.analyze', 'loan.documents.review')
def hitung_pinjaman():
    """Hitung estimasi pinjaman berdasarkan plafon + tenor fleksibel."""
    try:
        raw_plafon = request.args.get('plafon') or request.args.get('Pinjaman', 0)
        raw_tenor = request.args.get('tenor')
        id_anggota = request.args.get('id_anggota')
        jenis_pm = (request.args.get('jenis_pinjaman') or '').strip()
        plafon = float(raw_plafon)
        if not raw_tenor:
            return jsonify({'error': 'Tenor wajib diisi'})
        tenor = int(raw_tenor)
        if tenor <= 0:
            return jsonify({'error': 'Tenor harus lebih dari 0'}), 400
        bunga_persen = bunga_untuk_jenis_pinjaman(jenis_pm, tenor)
        total_bayar = hitung_total_bayar_tanpa_provisi(plafon, bunga_persen, tenor, jenis_pm)
        provisi_nominal = provisi_nominal_pinjaman(jenis_pm, plafon, tenor)
        provisi_persen = provisi_persen_dari_pinjaman(jenis_pm, tenor)
        # Hitung cicilan bulanan: cicilan = (p / b) + (p × j%)
        cicilan = hitung_cicilan_bulanan(plafon, bunga_persen, tenor)

        kapasitas_cicilan = None
        plafon_maks = None
        dsr = None
        dsr_persen = None
        if id_anggota and not session.get('user'):
            id_anggota = ''
        if id_anggota and not is_current_user_admin():
            current_id_anggota = get_current_user_id_anggota()
            if not current_id_anggota or id_anggota != current_id_anggota:
                abort(403)
        if id_anggota:
            ensure_anggota_schema()
            anggota = get_anggota_by_id(id_anggota)
            if anggota:
                penghasilan = float(anggota.get('penghasilan_bersih') or 0)
                cicilan_lain = float(anggota.get('cicilan_lain') or 0)
                dsr = dsr_otomatis_dari_penghasilan(penghasilan)
                dsr_persen = round(dsr * 100)
                kapasitas_cicilan = hitung_kapasitas_cicilan(penghasilan, cicilan_lain, dsr)
                plafon_maks = hitung_plafon_maks_anuitas(kapasitas_cicilan, bunga_persen, tenor, jenis_pm)
                if jenis_pm == 'Solusi Cepat':
                    plafon_maks = min(plafon_maks, 3_000_000.0)
        return jsonify({
            'bunga_persen': bunga_persen,
            'tenor': tenor,
            'total_bayar': round(total_bayar, 2),
            'cicilan_per_bulan': round(cicilan, 2),
            'provisi_nominal': round(provisi_nominal, 2),
            'provisi_persen': round(provisi_persen, 2),
            'kapasitas_cicilan': round(kapasitas_cicilan, 2) if kapasitas_cicilan is not None else None,
            'plafon_maks': round(plafon_maks, 2) if plafon_maks is not None else None,
            'dsr_persen': dsr_persen,
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/pinjaman/cek_tunggakan/<id_anggota>')
@login_required
def cek_tunggakan(id_anggota):
    """Cek apakah anggota masih punya tunggakan."""
    restrict_id_anggota_or_forbid(id_anggota)
    pinjaman = baca_csv(FILE_PINJAMAN)
    tunggakan = [
        p for p in pinjaman
        if p.get('id_anggota') == id_anggota
        and p.get('status') == 'Disetujui'
        and saldo_pinjaman_aktual(p) > 0
    ]
    ada_solusi_cepat_aktif = ada_pinjaman_solusi_cepat_aktif(pinjaman, id_anggota)
    if tunggakan:
        return jsonify({
            'ada_tunggakan': True,
            'pinjaman_aktif': len(tunggakan),
            'ada_solusi_cepat_aktif': ada_solusi_cepat_aktif,
        })
    return jsonify({'ada_tunggakan': False, 'pinjaman_aktif': 0, 'ada_solusi_cepat_aktif': ada_solusi_cepat_aktif})


@app.route('/pinjaman/tambah', methods=['POST'])
@permission_required('loan.request', 'loan.disbursement.input')
@csrf_protect
def tambah_pinjaman():
    pinjaman = baca_csv(FILE_PINJAMAN)
    if is_current_user_admin():
        id_anggota = request.form['id_anggota']
    else:
        id_anggota = get_current_user_id_anggota()
        if not id_anggota:
            abort(403)
    ensure_anggota_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_data = next((a for a in anggota if a['id_anggota'] == id_anggota), None)

    if not anggota_data:
        return "Anggota tidak ditemukan", 400

    if not is_current_user_admin():
        pinjaman_telat = [
            p for p in pinjaman
            if p.get('id_anggota') == id_anggota
            and p.get('status') == 'Disetujui'
            and pinjaman_telat_bayar_aktual(p)
        ]
        if pinjaman_telat:
            flash('Pengajuan ditolak: masih ada pinjaman yang telat bayar melewati tanggal jatuh tempo.', 'danger')
            return redirect('/pinjaman')

    jenis = (request.form.get('jenis_pinjaman') or 'Jangka Pendek').strip()
    try:
        plafon = float(request.form['plafon'].replace(',', '').replace('.', ''))
        tenor = int(request.form['tenor_bulan'])
        bunga_persen = bunga_untuk_jenis_pinjaman(jenis, tenor)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect('/pinjaman')

    if jenis == 'Solusi Cepat' and ada_pinjaman_solusi_cepat_aktif(pinjaman, id_anggota):
        flash('Pengajuan Solusi Cepat ditolak: Solusi Cepat sebelumnya belum lunas. Jenis pinjaman lain tetap bisa diajukan.', 'danger')
        return redirect('/pinjaman')

    jenis_simpanan = (request.form.get('jenis_simpanan') or 'Manasuka').strip() or 'Manasuka'
    if jenis_simpanan not in JENIS_SIMPANAN:
        jenis_simpanan = 'Manasuka'

    if jenis == 'Solusi Cepat' and plafon > 3_000_000:
        flash('Pengajuan ditolak: plafon Solusi Cepat maksimal Rp 3.000.000.', 'danger')
        return redirect('/pinjaman')

    # Hitung total_bayar TANPA provisi untuk pengajuan awal
    total_bayar = hitung_total_bayar_tanpa_provisi(plafon, bunga_persen, tenor, jenis)
    cicilan = hitung_cicilan_bulanan(plafon, bunga_persen, tenor)
    
    # Hitung provisi untuk informasi user (khusus Jangka Panjang, dipotong saat pencairan)
    provisi_nominal = provisi_nominal_pinjaman(jenis, plafon, tenor)

    penghasilan = float(anggota_data.get('penghasilan_bersih') or 0)
    cicilan_lain = float(anggota_data.get('cicilan_lain') or 0)
    dsr = dsr_otomatis_dari_penghasilan(penghasilan)
    kapasitas_cicilan = hitung_kapasitas_cicilan(penghasilan, cicilan_lain, dsr)
    plafon_maks = hitung_plafon_maks_anuitas(kapasitas_cicilan, bunga_persen, tenor, jenis)
    if jenis == 'Solusi Cepat':
        plafon_maks = min(plafon_maks, 3_000_000.0)
    if kapasitas_cicilan <= 0:
        flash('Pengajuan ditolak: kapasitas cicilan anggota tidak mencukupi.', 'danger')
        return redirect('/pinjaman')
    if cicilan > kapasitas_cicilan:
        flash(
            f"Pengajuan ditolak: cicilan bulanan Rp {cicilan:,.0f} melebihi kapasitas Rp {kapasitas_cicilan:,.0f}.",
            'danger'
        )
        return redirect('/pinjaman')
    if plafon > plafon_maks:
        flash(
            f"Pengajuan melebihi plafon maksimal anggota. Maksimal Rp {plafon_maks:,.0f} "
            f"(kapasitas cicilan Rp {kapasitas_cicilan:,.0f}/bln).",
            'danger'
        )
        return redirect('/pinjaman')

    tgl = datetime.now().strftime('%Y-%m-%d')
    pid = str(uuid.uuid4())
    pinjaman.append({
        'id_pinjaman': pid,
        'id_anggota': id_anggota,
        'nama_anggota': anggota_data.get('nama_lengkap', ''),
        'no_anggota': anggota_data.get('no_anggota', ''),
        'jenis_pinjaman': jenis,
        'jenis_simpanan': jenis_simpanan,
        'plafon': str(round(plafon, 2)),
        'tenor_awal': str(tenor),
        'tenor_bulan': str(tenor),
        'bunga_persen': str(bunga_persen),
        'total_bayar': str(round(total_bayar, 2)),
        'cicilan_per_bulan': str(round(cicilan, 2)),
        'sisa_pinjaman': str(round(plafon, 2)),
        'tanggal_pengajuan': tgl,
        'status': 'Menunggu',
        'tanggal_lunas': '',
    })
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    provisi_info = (
        f' Provisi Rp {provisi_nominal:,.0f} akan dipotong dari dana pencairan saat disetujui admin.'
        if provisi_nominal > 0 else ''
    )
    flash(f'Pengajuan pinjaman tercatat dan menunggu konfirmasi admin.{provisi_info}', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/konfirmasi/<id_pinjaman>', methods=['POST'])
@permission_required('loans.approve')
@csrf_protect
def konfirmasi_pinjaman(id_pinjaman):
    pinjaman = baca_csv(FILE_PINJAMAN)
    found = False
    provisi_nominal = 0.0
    dana_cair = 0.0
    for i, p in enumerate(pinjaman):
        if p.get('id_pinjaman') == id_pinjaman and p.get('status') == 'Menunggu':
            # Ubah status menjadi Disetujui
            p['status'] = 'Disetujui'
            
            # Provisi dipotong dari dana cair, bukan ditambahkan ke cicilan.
            try:
                jenis_pinjaman = (p.get('jenis_pinjaman') or '').strip()
                plafon = float(p.get('plafon') or 0)
                tenor = int(float(p.get('tenor_bulan') or 0))
                bunga_persen = float(p.get('bunga_persen') or 0)
                
                # Kewajiban cicilan tetap tanpa provisi (provisi dipotong saat pencairan)
                total_bayar = hitung_total_bayar_tanpa_provisi(plafon, bunga_persen, tenor, jenis_pinjaman)
                cicilan = hitung_cicilan_bulanan(plafon, bunga_persen, tenor)
                provisi_nominal = provisi_nominal_pinjaman(jenis_pinjaman, plafon, tenor)
                dana_cair = max(plafon - provisi_nominal, 0.0)
                
                p['total_bayar'] = str(round(total_bayar, 2))
                p['cicilan_per_bulan'] = str(round(cicilan, 2))
                p['tenor_awal'] = str(tenor)
                p['sisa_pinjaman'] = str(round(plafon, 2))
            except (TypeError, ValueError):
                # Jika gagal hitung, gunakan nilai existing
                plaf = float(p.get('plafon') or 0)
                p['tenor_awal'] = str(int(float(p.get('tenor_bulan') or 0)))
                p['sisa_pinjaman'] = str(round(plaf, 2))
            
            found = True
            break
    
    if not found:
        flash('Pengajuan tidak ditemukan atau sudah diproses.', 'warning')
        return redirect('/pinjaman')
    
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    
    # Cari id_anggota dari pinjaman yang baru dikonfirmasi
    id_anggota = None
    for p in pinjaman:
        if p.get('id_pinjaman') == id_pinjaman:
            id_anggota = p.get('id_anggota')
            break
    
    # Merge dengan pinjaman approved lain dari anggota sama (jika ada)
    if id_anggota:
        merge_pinjaman_sama_anggota(pinjaman, id_pinjaman, id_anggota)
        tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    
    if provisi_nominal > 0:
        flash(
            f'Pinjaman disetujui. Provisi Rp {provisi_nominal:,.0f} dipotong saat pencairan. '
            f'Dana cair bersih: Rp {dana_cair:,.0f}.',
            'success'
        )
    else:
        flash('Pinjaman disetujui.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/tolak/<id_pinjaman>', methods=['POST'])
@permission_required('loans.approve')
@csrf_protect
def tolak_pinjaman(id_pinjaman):
    pinjaman = baca_csv(FILE_PINJAMAN)
    found = False
    for p in pinjaman:
        if p.get('id_pinjaman') == id_pinjaman and p.get('status') == 'Menunggu':
            p['status'] = 'Ditolak'
            found = True
            break
    if not found:
        flash('Pengajuan tidak ditemukan atau sudah diproses.', 'warning')
        return redirect('/pinjaman')
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)
    flash('Pengajuan pinjaman ditolak.', 'warning')
    return redirect('/pinjaman')


def _proses_angsur_pinjaman(id_pinjaman: str):
    """Kurangi sisa kewajiban satu kali cicilan dan turunkan tenor tersisa 1 bulan."""
    pinjaman = baca_csv(FILE_PINJAMAN)
    for p in pinjaman:
        if p.get('id_pinjaman') != id_pinjaman:
            continue
        if p.get('status') != 'Disetujui':
            return
        sisa = saldo_pinjaman_aktual(p)
        cic = nominal_cicilan_aktual(p)
        baru = max(sisa - cic, 0)
        p['sisa_pinjaman'] = str(round(baru, 2))
        try:
            tenor_now = int(float(p.get('tenor_bulan') or p.get('tenor_awal') or 0))
        except (TypeError, ValueError):
            tenor_now = 0
        p['tenor_bulan'] = str(max(tenor_now - 1, 0))
        if baru <= 0:
            p['status'] = 'Lunas'
            p['tanggal_lunas'] = datetime.now().strftime('%Y-%m-%d')
            p['tenor_bulan'] = '0'
        break
    tulis_csv(FILE_PINJAMAN, pinjaman, PINJAMAN_FIELDNAMES)


def _pinjaman_row_dengan_nama(id_pinjaman: str):
    pinjaman = baca_csv(FILE_PINJAMAN)
    target = next((p for p in pinjaman if p.get('id_pinjaman') == id_pinjaman), None)
    if not target:
        return None, None
    ag = get_anggota_by_id(target.get('id_anggota', '')) or {}
    cic = float(target.get('cicilan_per_bulan') or 0) or cicilan_per_bulan_saldo(target)
    sisa = saldo_pinjaman_aktual(target)
    return target, {
        'no_anggota': target.get('no_anggota') or ag.get('no_anggota', ''),
        'nama_anggota': target.get('nama_anggota') or ag.get('nama_lengkap', ''),
        'cicilan_per_bulan': str(round(cic, 2)),
        'saldo_pinjaman': str(round(sisa, 2)),
    }


def _jumlah_cicilan_terhitung(cicilan_rows: list, id_pinjaman: str) -> int:
    """Hitung cicilan yang sudah memakan kuota tenor (disetujui + gagal bayar)."""
    total = 0
    for c in cicilan_rows:
        if c.get('id_pinjaman') != id_pinjaman:
            continue
        status = (c.get('status') or '').strip()
        if status in ('Disetujui', 'Gagal Bayar'):
            total += 1
    return total


def _sudah_gagal_bayar_bulan_periode(cicilan_rows: list, id_pinjaman: str, periode) -> bool:
    """Cek apakah pinjaman sudah pernah ditandai Gagal Bayar di bulan periode yang sama."""
    for c in cicilan_rows:
        if c.get('id_pinjaman') != id_pinjaman:
            continue
        if (c.get('status') or '').strip() != 'Gagal Bayar':
            continue
        tanggal_str = (c.get('tanggal_konfirmasi') or c.get('tanggal_pengajuan') or '').strip()
        if not tanggal_str:
            continue
        parsed = _parse_tanggal_iso(tanggal_str)
        if not parsed:
            continue
        if parsed.year == periode.year and parsed.month == periode.month:
            return True
    return False


@app.route('/pinjaman/angsur/<id_pinjaman>', methods=['POST'])
@permission_required('installments.manage')
@csrf_protect
def angsur_pinjaman(id_pinjaman):
    """Bayar cicilan langsung oleh admin; mencatat metode Admin di riwayat cicilan."""
    ensure_pinjaman_cicilan_schema()
    ket = (request.form.get('keterangan') or '').strip() or 'Pembayaran cicilan melalui admin (kas koperasi)'
    target, meta = _pinjaman_row_dengan_nama(id_pinjaman)
    if not target:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    if target.get('status') != 'Disetujui' or saldo_pinjaman_aktual(target) <= 0:
        flash('Pinjaman ini tidak memiliki cicilan yang dapat dibayar.', 'warning')
        return redirect('/pinjaman')

    id_anggota = target.get('id_anggota', '')
    jumlah_bayar = nominal_cicilan_aktual(target)
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    cicilan.append({
        'id_cicilan': str(uuid.uuid4()),
        'id_pinjaman': id_pinjaman,
        'id_anggota': id_anggota,
        'no_anggota': meta['no_anggota'],
        'nama_anggota': meta['nama_anggota'],
        'jumlah': str(round(jumlah_bayar, 2)),
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Disetujui',
        'tanggal_konfirmasi': datetime.now().strftime('%Y-%m-%d'),
        'dikonfirmasi_oleh': session.get('user') or '',
        'diajukan_oleh': session.get('user') or '',
        'keterangan': ket,
        'metode_pembayaran': 'Admin',
        'detail_pembayaran': '',
    })
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    _proses_angsur_pinjaman(id_pinjaman)
    flash('Cicilan dicatat (bayar melalui admin). Saldo pinjaman diperbarui.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/lunasi/<id_pinjaman>', methods=['POST'])
@permission_required('installments.manage')
@csrf_protect
def lunasi_pinjaman(id_pinjaman):
    """Lunasi pinjaman aktif sekaligus dan catat riwayat cicilan admin."""
    ensure_pinjaman_cicilan_schema()
    ket = (request.form.get('keterangan') or '').strip() or 'Pelunasan pinjaman melalui admin (bayar lunas)'
    target, meta = _pinjaman_row_dengan_nama(id_pinjaman)
    if not target:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    if target.get('status') != 'Disetujui':
        flash('Pinjaman ini tidak dalam status aktif.', 'warning')
        return redirect('/pinjaman')

    sisa = saldo_pinjaman_aktual(target)
    if sisa <= 0:
        flash('Pinjaman ini sudah lunas.', 'info')
        return redirect('/pinjaman')

    id_anggota = target.get('id_anggota', '')
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    cicilan.append({
        'id_cicilan': str(uuid.uuid4()),
        'id_pinjaman': id_pinjaman,
        'id_anggota': id_anggota,
        'no_anggota': meta['no_anggota'],
        'nama_anggota': meta['nama_anggota'],
        'jumlah': str(round(sisa, 2)),
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Disetujui',
        'tanggal_konfirmasi': datetime.now().strftime('%Y-%m-%d'),
        'dikonfirmasi_oleh': session.get('user') or '',
        'diajukan_oleh': session.get('user') or '',
        'keterangan': ket,
        'metode_pembayaran': 'Admin',
        'detail_pembayaran': 'Pelunasan',
    })
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)

    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    for p in pinjaman_rows:
        if p.get('id_pinjaman') != id_pinjaman:
            continue
        p['sisa_pinjaman'] = '0'
        p['tenor_bulan'] = '0'
        p['status'] = 'Lunas'
        p['tanggal_lunas'] = datetime.now().strftime('%Y-%m-%d')
        break
    tulis_csv(FILE_PINJAMAN, pinjaman_rows, PINJAMAN_FIELDNAMES)

    flash('Pinjaman berhasil dilunasi sekaligus (bayar lunas).', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/cicilan/gagal-bayar/<id_pinjaman>', methods=['POST'])
@permission_required('installments.manage')
@csrf_protect
def gagal_bayar_pinjaman(id_pinjaman):
    """Catat cicilan gagal bayar dan akumulasikan ke sisa pinjaman."""
    ensure_pinjaman_cicilan_schema()
    ket = (request.form.get('keterangan') or '').strip() or 'Gagal bayar cicilan (terlambat / tidak tepat waktu)'
    target, meta = _pinjaman_row_dengan_nama(id_pinjaman)
    if not target:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    if target.get('status') != 'Disetujui' or saldo_pinjaman_aktual(target) <= 0:
        flash('Pinjaman ini tidak memiliki cicilan aktif yang bisa ditandai gagal bayar.', 'warning')
        return redirect('/pinjaman')

    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    try:
        tenor_awal = int(float(target.get('tenor_awal') or target.get('tenor_bulan') or 0))
    except (TypeError, ValueError):
        tenor_awal = 0
    total_terhitung = _jumlah_cicilan_terhitung(cicilan, id_pinjaman)
    if tenor_awal > 0 and total_terhitung >= tenor_awal:
        flash('Gagal bayar tidak bisa dicatat: kuota cicilan sudah mencapai tenor awal.', 'warning')
        return redirect('/pinjaman')

    try:
        plafon_pinjaman = max(float(target.get('plafon') or target.get('total_pinjaman') or 0), 0.0)
    except (TypeError, ValueError):
        plafon_pinjaman = 0.0
    sisa_sekarang = saldo_pinjaman_aktual(target)
    if plafon_pinjaman > 0 and sisa_sekarang >= plafon_pinjaman:
        flash('Gagal bayar tidak bisa dicatat: sisa pinjaman sudah mencapai batas plafon.', 'warning')
        return redirect('/pinjaman')

    jumlah_bayar = nominal_cicilan_aktual(target)
    tanggal_periode = _geser_bulan_tanggal(datetime.now().date(), -1) or datetime.now().date()
    tanggal_periode_str = tanggal_periode.strftime('%Y-%m-%d')
    if _sudah_gagal_bayar_bulan_periode(cicilan, id_pinjaman, tanggal_periode):
        flash('Gagal bayar untuk periode bulan ini sudah pernah dicatat. Ulangi lagi di bulan berikutnya.', 'warning')
        return redirect('/pinjaman')
    cicilan.append({
        'id_cicilan': str(uuid.uuid4()),
        'id_pinjaman': id_pinjaman,
        'id_anggota': target.get('id_anggota', ''),
        'no_anggota': meta['no_anggota'],
        'nama_anggota': meta['nama_anggota'],
        'jumlah': str(round(jumlah_bayar, 2)),
        'tanggal_pengajuan': tanggal_periode_str,
        'status': 'Gagal Bayar',
        'tanggal_konfirmasi': tanggal_periode_str,
        'dikonfirmasi_oleh': session.get('user') or '',
        'diajukan_oleh': session.get('user') or '',
        'keterangan': ket,
        'metode_pembayaran': 'Admin',
        'detail_pembayaran': '',
    })
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)

    # Akumulasi tunggakan: cicilan gagal bayar ditambahkan ke sisa pinjaman.
    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    for p in pinjaman_rows:
        if p.get('id_pinjaman') != id_pinjaman:
            continue
        sisa_sekarang = saldo_pinjaman_aktual(p)
        try:
            plafon_total = max(float(p.get('plafon') or p.get('total_pinjaman') or 0), 0.0)
        except (TypeError, ValueError):
            plafon_total = 0.0
        sisa_baru = max(sisa_sekarang + jumlah_bayar, 0.0)
        if plafon_total > 0:
            sisa_baru = min(sisa_baru, plafon_total)
        p['sisa_pinjaman'] = str(round(sisa_baru, 2))
        try:
            tenor_now = int(float(p.get('tenor_bulan') or 0))
        except (TypeError, ValueError):
            tenor_now = 0
        try:
            tenor_awal = int(float(p.get('tenor_awal') or p.get('tenor_bulan') or 0))
        except (TypeError, ValueError):
            tenor_awal = tenor_now
        tenor_cap = max(tenor_awal, 0)
        p['tenor_bulan'] = str(min(max(tenor_now + 1, 0), tenor_cap))
        if p.get('status') == 'Lunas':
            p['status'] = 'Disetujui'
            p['tanggal_lunas'] = ''
        break
    tulis_csv(FILE_PINJAMAN, pinjaman_rows, PINJAMAN_FIELDNAMES)

    flash('Cicilan ditandai gagal bayar. Nilai cicilan dan tenor dikembalikan ke akumulasi periode sebelumnya.', 'warning')
    return redirect('/pinjaman')


@app.route('/pinjaman/ajukan-cicilan/<id_pinjaman>', methods=['POST'])
@permission_required('installments.proof.upload', 'installments.manage')
@csrf_protect
def ajukan_cicilan(id_pinjaman):
    """Pengajuan bayar cicilan oleh user; perlu dikonfirmasi admin sebelum mengurangi sisa pinjaman."""
    ensure_pinjaman_cicilan_schema()
    target, meta = _pinjaman_row_dengan_nama(id_pinjaman)
    if not target:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')

    id_anggota = target.get('id_anggota', '')
    if not is_current_user_admin():
        id_anggota_user = get_current_user_id_anggota()
        if not id_anggota_user or id_anggota != id_anggota_user:
            abort(403)

    if target.get('status') != 'Disetujui' or saldo_pinjaman_aktual(target) <= 0:
        flash('Pinjaman ini tidak memiliki cicilan yang perlu dibayar.', 'warning')
        return redirect('/pinjaman')

    if (target.get('jenis_pinjaman') or '').strip() == 'Solusi Cepat':
        flash('Solusi Cepat hanya bisa dibayar melalui admin.', 'warning')
        return redirect('/pinjaman')

    metode = (request.form.get('metode_pembayaran') or 'Lainnya').strip()
    if metode not in METODE_BAYAR_CHOICES:
        metode = 'Lainnya'
    detail = (request.form.get('detail_pembayaran') or request.form.get('detail_pembayaran_transfer') or '').strip()
    idempotency_key = (request.form.get('idempotency_key') or '').strip() or uuid.uuid4().hex
    jumlah_bayar = nominal_cicilan_aktual(target)
    bukti_path = _save_bukti_transfer(request.files.get('bukti_transfer'), id_pinjaman)
    if bukti_path:
        detail = (detail + (' | ' if detail else '') + f"Bukti: /static/{bukti_path}").strip()
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    if _mark_expired_cicilan(cicilan):
        tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)

    # Anti-duplikasi: periodik bulanan per pinjaman + idempotency key.
    periode_tagihan = datetime.now().strftime('%Y-%m')
    for c in cicilan:
        if (c.get('idempotency_key') or '').strip() == idempotency_key:
            flash('Permintaan terdeteksi duplikat (idempotency). Silakan refresh halaman.', 'warning')
            return redirect('/pinjaman')
        if (c.get('id_pinjaman') or '').strip() != id_pinjaman:
            continue
        if (c.get('periode_tagihan') or '').strip() != periode_tagihan:
            continue
        trx_status = (c.get('status_transaksi') or '').strip()
        if trx_status in (PAYMENT_STATUS_WAITING_PAYMENT, PAYMENT_STATUS_WAITING_VERIFICATION, PAYMENT_STATUS_SUCCESS):
            flash('Ajuan pembayaran untuk periode ini sudah ada. Hindari submit ganda.', 'warning')
            return redirect('/pinjaman')

    va_number = _generate_va_number(meta.get('no_anggota', ''), id_pinjaman)
    expires_dt = datetime.now() + timedelta(minutes=10 if metode == 'QRIS' else 24 * 60)
    if metode == 'QRIS':
        status_transaksi = PAYMENT_STATUS_WAITING_PAYMENT
    else:
        status_transaksi = PAYMENT_STATUS_WAITING_VERIFICATION

    ket_parts = [f'Pengajuan pembayaran via {metode}', f'VA: {va_number}']
    if metode == 'QRIS':
        qris_payload = _build_qris_payload_dana(KOPERASI_REKENING_BANK.get('no_rekening', ''), jumlah_bayar, id_pinjaman)
        detail = (detail + (' | ' if detail else '') + f"QRIS_PAYLOAD: {qris_payload}").strip()
    if detail:
        ket_parts.append(detail)
    keterangan = ' — '.join(ket_parts)

    cicilan.append({
        'id_cicilan': str(uuid.uuid4()),
        'id_pinjaman': id_pinjaman,
        'id_anggota': id_anggota,
        'no_anggota': meta['no_anggota'],
        'nama_anggota': meta['nama_anggota'],
        'jumlah': str(round(jumlah_bayar, 2)),
        'tanggal_pengajuan': datetime.now().strftime('%Y-%m-%d'),
        'status': 'Menunggu',
        'tanggal_konfirmasi': '',
        'dikonfirmasi_oleh': '',
        'diajukan_oleh': session.get('user') or '',
        'keterangan': keterangan,
        'metode_pembayaran': metode,
        'detail_pembayaran': detail,
        'status_transaksi': status_transaksi,
        'va_number': va_number,
        'idempotency_key': idempotency_key,
        'periode_tagihan': periode_tagihan,
        'expires_at': expires_dt.strftime('%Y-%m-%d %H:%M:%S'),
    })
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    _send_email_notification(
        subject='[Koperasi] Ajuan bayar cicilan baru',
        body=(
            f"Ajuan baru dari {meta.get('nama_anggota', '-')}\n"
            f"No Anggota: {meta.get('no_anggota', '-')}\n"
            f"Metode: {metode}\n"
            f"Nominal: {_format_currency_idr(jumlah_bayar)}\n"
            f"Status transaksi: {status_transaksi}\n"
            f"VA: {va_number}\n"
            f"Periode: {periode_tagihan}\n"
        ),
    )
    flash('Pengajuan bayar cicilan berhasil dikirim. Menunggu konfirmasi admin.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/cicilan/konfirmasi/<id_cicilan>', methods=['POST'])
@permission_required('installments.manage')
@csrf_protect
def konfirmasi_cicilan(id_cicilan):
    """Admin menyetujui pengajuan cicilan dan langsung mengurangi sisa pinjaman."""
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    if _mark_expired_cicilan(cicilan):
        tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    target = None
    for c in cicilan:
        if c.get('id_cicilan') == id_cicilan:
            target = c
            break
    if not target:
        flash('Data pengajuan cicilan tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    if target.get('status') != 'Menunggu':
        flash('Pengajuan cicilan ini sudah diproses sebelumnya.', 'warning')
        return redirect('/pinjaman')

    pinjaman, _ = _pinjaman_row_dengan_nama(target['id_pinjaman'])
    if not pinjaman:
        flash('Data pinjaman tidak ditemukan.', 'danger')
        return redirect('/pinjaman')

    jumlah_bayar = nominal_cicilan_aktual(pinjaman)

    target['status'] = 'Disetujui'
    target['status_transaksi'] = PAYMENT_STATUS_SUCCESS
    target['tanggal_konfirmasi'] = datetime.now().strftime('%Y-%m-%d')
    target['dikonfirmasi_oleh'] = session.get('user') or ''
    target['jumlah'] = str(round(jumlah_bayar, 2))
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)

    _proses_angsur_pinjaman(target['id_pinjaman'])
    user_email = _resolve_user_email_for_cicilan(target)
    if user_email:
        _send_email_notification(
            subject='[Koperasi] Pembayaran cicilan selesai',
            body=(
                f"Halo {target.get('nama_anggota', 'Anggota')},\n\n"
                f"Pembayaran cicilan Anda telah selesai diproses admin.\n"
                f"ID Ajuan: {target.get('id_cicilan', '-')}\n"
                f"No Anggota: {target.get('no_anggota', '-')}\n"
                f"Nominal: {_format_currency_idr(jumlah_bayar)}\n"
                f"Status: {PAYMENT_STATUS_SUCCESS}\n\n"
                f"Salam,\nAdmin Koperasi"
            ),
            to_email=user_email,
        )
    _send_email_notification(
        subject='[Koperasi] Ajuan bayar disetujui',
        body=(
            f"Ajuan ID: {target.get('id_cicilan', '-')}\n"
            f"Nama: {target.get('nama_anggota', '-')}\n"
            f"No Anggota: {target.get('no_anggota', '-')}\n"
            f"Nominal: {_format_currency_idr(jumlah_bayar)}\n"
            f"Status transaksi: {PAYMENT_STATUS_SUCCESS}\n"
            f"Diproses oleh: {session.get('user') or '-'}\n"
        ),
    )
    flash('Pengajuan bayar cicilan diterima dan sisa pinjaman telah diperbarui.', 'success')
    return redirect('/pinjaman')


@app.route('/pinjaman/cicilan/menunggu', methods=['GET'])
@permission_required('installments.manage')
def cicilan_menunggu_json():
    """Data pengajuan cicilan menunggu untuk refresh otomatis di halaman pinjaman."""
    ensure_pinjaman_cicilan_schema()
    rows = baca_csv(FILE_PINJAMAN_CICILAN)
    if _mark_expired_cicilan(rows):
        tulis_csv(FILE_PINJAMAN_CICILAN, rows, CICILAN_FIELDNAMES)
    rows = [c for c in rows if (c.get('status') or '').strip() == 'Menunggu']
    rows.sort(key=lambda x: x.get('tanggal_pengajuan', ''), reverse=True)

    items = []
    for c in rows:
        try:
            jumlah = float(c.get('jumlah') or 0)
        except (TypeError, ValueError):
            jumlah = 0.0
        items.append({
            'id_cicilan': c.get('id_cicilan', ''),
            'tanggal_pengajuan': c.get('tanggal_pengajuan', '-'),
            'no_anggota': c.get('no_anggota', '-') or '-',
            'nama_anggota': c.get('nama_anggota', '-') or '-',
            'jumlah': round(jumlah, 2),
            'metode_pembayaran': c.get('metode_pembayaran', '-') or '-',
            'status_transaksi': c.get('status_transaksi', '-') or '-',
            'va_number': c.get('va_number', '-') or '-',
            'keterangan': c.get('keterangan', '-') or '-',
        })

    return jsonify({'count': len(items), 'items': items})


@app.route('/pinjaman/cicilan/tolak/<id_cicilan>', methods=['POST'])
@permission_required('installments.manage')
@csrf_protect
def tolak_cicilan(id_cicilan):
    """Admin menolak pengajuan cicilan (tanpa mengubah sisa pinjaman)."""
    cicilan = baca_csv(FILE_PINJAMAN_CICILAN)
    if _mark_expired_cicilan(cicilan):
        tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    found = False
    for c in cicilan:
        if c.get('id_cicilan') == id_cicilan:
            c['status'] = 'Ditolak'
            c['status_transaksi'] = PAYMENT_STATUS_FAILED
            c['tanggal_konfirmasi'] = datetime.now().strftime('%Y-%m-%d')
            c['dikonfirmasi_oleh'] = session.get('user') or ''
            found = True
            break
    if not found:
        flash('Data pengajuan cicilan tidak ditemukan.', 'danger')
        return redirect('/pinjaman')
    tulis_csv(FILE_PINJAMAN_CICILAN, cicilan, CICILAN_FIELDNAMES)
    _send_email_notification(
        subject='[Koperasi] Ajuan bayar ditolak',
        body=(
            f"Ajuan ID: {id_cicilan}\n"
            f"Status transaksi: {PAYMENT_STATUS_FAILED}\n"
            f"Diproses oleh: {session.get('user') or '-'}\n"
        ),
    )
    flash('Pengajuan bayar cicilan ditolak.', 'warning')
    return redirect('/pinjaman')


@app.route('/pinjaman/hapus/<id_pinjaman>', methods=['POST'])
@permission_required('installments.manage')
@csrf_protect
def hapus_pinjaman(id_pinjaman):
    data = [p for p in baca_csv(FILE_PINJAMAN) if p.get('id_pinjaman') != id_pinjaman]
    tulis_csv(FILE_PINJAMAN, data, PINJAMAN_FIELDNAMES)
    flash('Data pinjaman dihapus.', 'success')
    return redirect('/pinjaman')


# ══════════════════════════════════════════════
#  ROUTE: LAPORAN & EXPORT
# ══════════════════════════════════════════════

def upsert_anggota_from_riwayat(pinjaman_rows: list, simpanan_rows: list) -> dict:
    """Sinkronkan anggota dari riwayat dengan key wajib: no anggota + NIK + no HP."""
    ensure_anggota_schema()
    anggota = baca_csv(FILE_ANGGOTA)
    by_id = {a.get('id_anggota'): a for a in anggota}
    by_key = {}
    for a in anggota:
        key = (
            (a.get('no_anggota') or '').strip().upper(),
            normalize_nik(a.get('nik')),
            re.sub(r'\D', '', a.get('no_hp') or ''),
        )
        if all(key):
            by_key[key] = a

    added = 0
    failed = 0

    def key_from_row(row: dict) -> tuple:
        return (
            (row.get('no_anggota') or '').strip().upper(),
            normalize_nik(row.get('nik')),
            re.sub(r'\D', '', row.get('no_hp') or '')
        )

    def ensure_id(row: dict):
        nonlocal added, failed
        id_a = (row.get('id_anggota') or '').strip()
        if id_a and id_a in by_id:
            return

        no_anggota, nik, no_hp = key_from_row(row)
        if not (no_anggota and nik and no_hp):
            failed += 1
            return

        matched = by_key.get((no_anggota, nik, no_hp))
        if matched:
            row['id_anggota'] = matched.get('id_anggota', '')
            return

        id_baru = id_a or str(uuid.uuid4())
        anggota.append({
            'id_anggota': id_baru,
            'no_anggota': no_anggota,
            'nik': nik,
            'nama': row.get('nama_anggota', '') or row.get('nama', ''),
            'alamat': row.get('alamat', ''),
            'no_telp': row.get('no_hp', '') or row.get('no_telp', ''),
            'tgl_bergabung': datetime.now().strftime('%Y-%m-%d'),
            'no_rekening': row.get('no_rekening', ''),
            'nama_bank': row.get('nama_bank', ''),
            'penghasilan_bersih': '0',
            'cicilan_lain': '0',
        })
        by_id[id_baru] = anggota[-1]
        by_key[(no_anggota, nik, no_hp)] = anggota[-1]
        row['id_anggota'] = id_baru
        added += 1

    for p in pinjaman_rows:
        ensure_id(p)
    for s in simpanan_rows:
        ensure_id(s)

    tulis_csv(FILE_ANGGOTA, anggota, ANGGOTA_FIELDNAMES)
    return {'added': added, 'updated': 0, 'failed': failed}


def hitung_shu_pinjaman() -> dict:
    """SHU perkiraan: proporsi dari nominal cicilan yang sudah disetujui (saldo pokok)."""
    ensure_pinjaman_cicilan_schema()
    cicilan_rows = [c for c in baca_csv(FILE_PINJAMAN_CICILAN) if c.get('status') == 'Disetujui']
    anggota_map = {a.get('id_anggota'): a for a in baca_csv(FILE_ANGGOTA)}

    bunga_per_anggota = {}
    total_bunga_dibayar = 0.0
    for c in cicilan_rows:
        id_anggota = (c.get('id_anggota') or '').strip()
        j = float(c.get('jumlah') or 0)
        if j <= 0:
            continue
        total_bunga_dibayar += j
        ag = anggota_map.get(id_anggota, {})
        if id_anggota not in bunga_per_anggota:
            bunga_per_anggota[id_anggota] = {
                'id_anggota': id_anggota,
                'no_anggota': ag.get('no_anggota', ''),
                'nama_anggota': ag.get('nama_lengkap', ''),
                'bunga_dibayar': 0.0,
            }
        bunga_per_anggota[id_anggota]['bunga_dibayar'] += j

    shu_kotor = total_bunga_dibayar
    shu_anggota_65 = shu_kotor * 0.65

    alokasi = []
    for row in bunga_per_anggota.values():
        prop = (row['bunga_dibayar'] / shu_kotor) if shu_kotor > 0 else 0.0
        shu_member = shu_anggota_65 * prop
        alokasi.append({
            'no_anggota': row['no_anggota'],
            'nama_anggota': row['nama_anggota'],
            'bunga_dibayar': row['bunga_dibayar'],
            'shu_anggota': shu_member,
        })

    alokasi.sort(key=lambda x: x['shu_anggota'], reverse=True)
    return {
        'shu_kotor': shu_kotor,
        'shu_anggota_65': shu_anggota_65,
        'shu_lain': shu_kotor - shu_anggota_65,
        'alokasi': alokasi
    }


def ensure_shu_schema():
    for path, fieldnames in (
        (FILE_SHU_TAHUNAN, SHU_TAHUNAN_FIELDNAMES),
        (FILE_SHU_ALOKASI, SHU_ALOKASI_FIELDNAMES),
    ):
        if os.path.exists(path):
            continue
        try:
            tulis_csv(path, [], fieldnames)
        except Exception:
            pass


def _parse_float(value, default: float = 0.0) -> float:
    try:
        if value is None:
            return float(default)
        text_value = str(value).strip().replace(',', '')
        if not text_value:
            return float(default)
        return float(text_value)
    except Exception:
        return float(default)


def _extract_year(value: str) -> str:
    raw = (value or '').strip()
    if not raw:
        return ''
    return raw[:4] if len(raw) >= 4 else ''


def _hitung_pendapatan_provisi_tahun(tahun: str) -> dict:
    target_tahun = (tahun or '').strip()
    rows = baca_csv(FILE_PINJAMAN)
    total_provisi = 0.0
    total_pinjaman_disetujui = 0
    anggota_provisi = {}

    for row in rows:
        status = (row.get('status') or '').strip().lower()
        if status not in {'disetujui', 'lunas'}:
            continue
        tahun_row = _extract_year(row.get('tanggal_pengajuan') or '')
        if target_tahun and tahun_row and tahun_row != target_tahun:
            continue

        jenis = (row.get('jenis_pinjaman') or '').strip()
        plafon = _parse_float(row.get('plafon'))
        tenor = int(_parse_float(row.get('tenor_bulan')))
        id_anggota = (row.get('id_anggota') or '').strip()

        provisi = _parse_float(row.get('provisi_nominal'))
        if provisi <= 0:
            provisi = provisi_nominal_pinjaman(jenis, plafon, tenor)
        if provisi <= 0:
            continue

        total_pinjaman_disetujui += 1
        total_provisi += provisi
        if id_anggota:
            anggota_provisi[id_anggota] = anggota_provisi.get(id_anggota, 0.0) + provisi

    return {
        'tahun': target_tahun,
        'total_provisi': total_provisi,
        'total_pinjaman_disetujui': total_pinjaman_disetujui,
        'anggota_provisi': anggota_provisi,
    }


def _rincian_pinjaman_sumber_shu(tahun: str) -> list[dict]:
    target_tahun = (tahun or '').strip()
    rows = baca_csv(FILE_PINJAMAN)
    detail = []

    for row in rows:
        status = (row.get('status') or '').strip().lower()
        if status not in {'disetujui', 'lunas'}:
            continue
        tahun_row = _extract_year(row.get('tanggal_pengajuan') or '')
        if target_tahun and tahun_row and tahun_row != target_tahun:
            continue

        jenis = (row.get('jenis_pinjaman') or '').strip()
        plafon = _parse_float(row.get('plafon'))
        tenor = int(_parse_float(row.get('tenor_bulan')))
        provisi_persen = provisi_persen_dari_pinjaman(jenis, tenor)
        provisi = _parse_float(row.get('provisi_nominal'))
        if provisi <= 0:
            provisi = provisi_nominal_pinjaman(jenis, plafon, tenor)
        if provisi <= 0:
            continue

        detail.append({
            'id_pinjaman': row.get('id_pinjaman', ''),
            'tanggal_pengajuan': row.get('tanggal_pengajuan', ''),
            'id_anggota': row.get('id_anggota', ''),
            'no_anggota': row.get('no_anggota', ''),
            'nama_anggota': row.get('nama_anggota', ''),
            'jenis_pinjaman': jenis,
            'plafon': plafon,
            'tenor_bulan': tenor,
            'provisi_persen': provisi_persen,
            'provisi_nominal': provisi,
            'status': row.get('status', ''),
        })

    detail.sort(key=lambda x: (x.get('tanggal_pengajuan') or '', x.get('id_pinjaman') or ''), reverse=True)
    return detail


def _hitung_jasa_anggota_shu(tahun: str) -> tuple[list[dict], dict]:
    anggota_rows = [a for a in baca_csv(FILE_ANGGOTA) if (a.get('id_anggota') or '').strip()]
    provisi_ctx = _hitung_pendapatan_provisi_tahun(tahun)
    provisi_map = provisi_ctx.get('anggota_provisi', {})

    hasil = []
    for anggota in anggota_rows:
        status_anggota = (anggota.get('status_anggota') or 'Aktif').strip().lower()
        if status_anggota and status_anggota not in {'aktif', 'active'}:
            continue
        id_anggota = (anggota.get('id_anggota') or '').strip()
        nilai_provisi = _parse_float(provisi_map.get(id_anggota, 0.0))
        hasil.append({
            'id_anggota': id_anggota,
            'no_anggota': anggota.get('no_anggota', ''),
            'nama_anggota': anggota.get('nama', ''),
            'pendapatan_provisi': nilai_provisi,
            'jasa_anggota': nilai_provisi,
        })
    return hasil, provisi_ctx


def hitung_shu_tahunan(total_shu: float, tahun: str, catatan: str = '') -> dict:
    tahun = (tahun or str(datetime.now().year)).strip()
    anggota_jasa, provisi_ctx = _hitung_jasa_anggota_shu(tahun)
    total_shu_input = max(_parse_float(total_shu), 0.0)
    total_shu = total_shu_input if total_shu_input > 0 else max(_parse_float(provisi_ctx.get('total_provisi')), 0.0)
    sumber_total_shu = 'Manual' if total_shu_input > 0 else 'Provisi Pinjaman'
    tahun = (tahun or str(datetime.now().year)).strip()

    cadangan_umum = total_shu * 0.15
    shu_pasif_total = total_shu * 0.15
    shu_aktif_total = total_shu * 0.50
    dana_kesejahteraan = total_shu * 0.05
    dana_pendidikan = total_shu * 0.005
    dana_sosial = total_shu * 0.02
    dana_pembangunan = total_shu * 0.005
    dana_pengurus = total_shu * 0.10
    dana_risiko = total_shu * 0.02

    jumlah_anggota = len(anggota_jasa)
    total_jasa = sum(item['jasa_anggota'] for item in anggota_jasa)
    shu_pasif_per_anggota = (shu_pasif_total / jumlah_anggota) if jumlah_anggota else 0.0

    alokasi = []
    for item in anggota_jasa:
        proporsi = (item['jasa_anggota'] / total_jasa) if total_jasa > 0 else (1.0 / jumlah_anggota if jumlah_anggota else 0.0)
        shu_aktif = shu_aktif_total * proporsi
        total_member = shu_pasif_per_anggota + shu_aktif
        alokasi.append({
            'id_anggota': item['id_anggota'],
            'no_anggota': item['no_anggota'],
            'nama_anggota': item['nama_anggota'],
            'jasa_anggota': item['jasa_anggota'],
            'nilai_jasa': item['jasa_anggota'],
            'shu_pasif': shu_pasif_per_anggota,
            'shu_aktif': shu_aktif,
            'total_shu': total_member,
            'status': 'Draft',
            'keterangan': catatan,
        })

    alokasi.sort(key=lambda row: row['total_shu'], reverse=True)
    catatan_sistem = (
        f"Sumber SHU: {sumber_total_shu}. "
        f"Tahun: {tahun}. "
        f"Pendapatan provisi pinjaman disetujui/lunas: Rp {provisi_ctx.get('total_provisi', 0):,.0f}. "
        f"Jumlah pinjaman dihitung: {int(provisi_ctx.get('total_pinjaman_disetujui', 0))}."
    )
    catatan_final = f"{catatan_sistem}\n{catatan}".strip() if catatan else catatan_sistem

    for row in alokasi:
        row['keterangan'] = catatan_final

    return {
        'tahun': tahun,
        'total_shu': total_shu,
        'cadangan_umum': cadangan_umum,
        'shu_pasif_total': shu_pasif_total,
        'shu_aktif_total': shu_aktif_total,
        'dana_kesejahteraan': dana_kesejahteraan,
        'dana_pendidikan': dana_pendidikan,
        'dana_sosial': dana_sosial,
        'dana_pembangunan': dana_pembangunan,
        'dana_pengurus': dana_pengurus,
        'dana_risiko': dana_risiko,
        'alokasi': alokasi,
        'sumber_total_shu': sumber_total_shu,
        'catatan_perhitungan': catatan_final,
    }


def _load_shu_records() -> list[dict]:
    ensure_shu_schema()
    if DATABASE_URL:
        try:
            with db_session() as conn:
                rows = conn.execute(text("SELECT * FROM shu_tahunan ORDER BY tahun DESC")).mappings().all()
            results = []
            for r in rows:
                results.append({
                    'id_shu': r.get('id_shu'),
                    'tahun': str(r.get('tahun')) if r.get('tahun') is not None else '',
                    'tanggal_input': (r.get('tanggal_input').strftime('%Y-%m-%d %H:%M:%S') if r.get('tanggal_input') else ''),
                    'total_shu': float(r.get('total_shu') or 0),
                    'cadangan_umum': float(r.get('cadangan_umum') or 0),
                    'shu_pasif_total': float(r.get('shu_pasif_total') or 0),
                    'shu_aktif_total': float(r.get('shu_aktif_total') or 0),
                    'dana_kesejahteraan': float(r.get('dana_kesejahteraan') or 0),
                    'dana_pendidikan': float(r.get('dana_pendidikan') or 0),
                    'dana_sosial': float(r.get('dana_sosial') or 0),
                    'dana_pembangunan': float(r.get('dana_pembangunan') or 0),
                    'dana_pengurus': float(r.get('dana_pengurus') or 0),
                    'dana_risiko': float(r.get('dana_risiko') or 0),
                    'status': r.get('status') or 'Draft',
                    'dikonfirmasi_oleh': r.get('dikonfirmasi_oleh') or '',
                    'tanggal_konfirmasi': (r.get('tanggal_konfirmasi').strftime('%Y-%m-%d %H:%M:%S') if r.get('tanggal_konfirmasi') else ''),
                    'catatan': r.get('catatan') or '',
                })
            return results
        except Exception:
            return baca_csv(FILE_SHU_TAHUNAN)
    return baca_csv(FILE_SHU_TAHUNAN)


def _load_shu_allocation() -> list[dict]:
    ensure_shu_schema()
    if DATABASE_URL:
        try:
            with db_session() as conn:
                rows = conn.execute(text("SELECT * FROM shu_anggota ORDER BY created_at DESC")).mappings().all()
            results = []
            for r in rows:
                results.append({
                    'id_alokasi': r.get('id_shu_anggota'),
                    'id_shu': r.get('id_shu'),
                    'id_anggota': r.get('id_anggota'),
                    'no_anggota': r.get('no_anggota'),
                    'nama_anggota': r.get('nama_anggota'),
                    'jasa_anggota': float(r.get('jasa_anggota') or 0),
                    'nilai_jasa': float(r.get('nilai_jasa') or 0),
                    'shu_pasif': float(r.get('shu_pasif') or 0),
                    'shu_aktif': float(r.get('shu_aktif') or 0),
                    'total_shu': float(r.get('shu_total') or 0),
                    'status': r.get('status_ambil') or r.get('status') or 'Draft',
                    'keterangan': r.get('catatan') or '',
                })
            return results
        except Exception:
            return baca_csv(FILE_SHU_ALOKASI)
    return baca_csv(FILE_SHU_ALOKASI)


def _save_shu_data(record: dict, alokasi_rows: list[dict]) -> None:
    # If DATABASE_URL is set, persist to Postgres; otherwise use file-based storage
    if DATABASE_URL:
        try:
            with db_session() as conn:
                # insert or update shu_tahunan
                conn.execute(text(
                    """
                    INSERT INTO shu_tahunan (id_shu, tahun, tanggal_input, total_shu, cadangan_umum,
                        shu_pasif_total, shu_aktif_total, dana_kesejahteraan, dana_pendidikan,
                        dana_sosial, dana_pembangunan, dana_pengurus, dana_risiko, status, dikonfirmasi_oleh, tanggal_konfirmasi, catatan)
                    VALUES (:id_shu, :tahun, :tanggal_input, :total_shu, :cadangan_umum,
                        :shu_pasif_total, :shu_aktif_total, :dana_kesejahteraan, :dana_pendidikan,
                        :dana_sosial, :dana_pembangunan, :dana_pengurus, :dana_risiko, :status, :dikonfirmasi_oleh, :tanggal_konfirmasi, :catatan)
                    ON CONFLICT (id_shu) DO UPDATE SET
                        tahun = EXCLUDED.tahun,
                        total_shu = EXCLUDED.total_shu,
                        cadangan_umum = EXCLUDED.cadangan_umum,
                        shu_pasif_total = EXCLUDED.shu_pasif_total,
                        shu_aktif_total = EXCLUDED.shu_aktif_total,
                        dana_kesejahteraan = EXCLUDED.dana_kesejahteraan,
                        dana_pendidikan = EXCLUDED.dana_pendidikan,
                        dana_sosial = EXCLUDED.dana_sosial,
                        dana_pembangunan = EXCLUDED.dana_pembangunan,
                        dana_pengurus = EXCLUDED.dana_pengurus,
                        dana_risiko = EXCLUDED.dana_risiko,
                        status = EXCLUDED.status,
                        dikonfirmasi_oleh = EXCLUDED.dikonfirmasi_oleh,
                        tanggal_konfirmasi = EXCLUDED.tanggal_konfirmasi,
                        catatan = EXCLUDED.catatan
                    """,
                ), {
                    'id_shu': record.get('id_shu'),
                    'tahun': int(record.get('tahun')) if record.get('tahun') else None,
                    'tanggal_input': record.get('tanggal_input') or datetime.now(),
                    'total_shu': float(record.get('total_shu') or 0),
                    'cadangan_umum': float(record.get('cadangan_umum') or 0),
                    'shu_pasif_total': float(record.get('shu_pasif_total') or 0),
                    'shu_aktif_total': float(record.get('shu_aktif_total') or 0),
                    'dana_kesejahteraan': float(record.get('dana_kesejahteraan') or 0),
                    'dana_pendidikan': float(record.get('dana_pendidikan') or 0),
                    'dana_sosial': float(record.get('dana_sosial') or 0),
                    'dana_pembangunan': float(record.get('dana_pembangunan') or 0),
                    'dana_pengurus': float(record.get('dana_pengurus') or 0),
                    'dana_risiko': float(record.get('dana_risiko') or 0),
                    'status': record.get('status') or 'Draft',
                    'dikonfirmasi_oleh': record.get('dikonfirmasi_oleh') or '',
                    'tanggal_konfirmasi': record.get('tanggal_konfirmasi') or None,
                    'catatan': record.get('catatan') or '',
                })

                # remove existing anggota allocations for this id_shu and re-insert
                conn.execute(text("DELETE FROM shu_anggota WHERE id_shu = :id_shu"), {'id_shu': record.get('id_shu')})
                for a in alokasi_rows:
                    conn.execute(text(
                        """
                        INSERT INTO shu_anggota (id_shu_anggota, id_shu, id_anggota, no_anggota, nama_anggota,
                            jasa_anggota, nilai_jasa, shu_pasif, shu_aktif, shu_total, status_ambil, catatan)
                        VALUES (:id_shu_anggota, :id_shu, :id_anggota, :no_anggota, :nama_anggota,
                            :jasa_anggota, :nilai_jasa, :shu_pasif, :shu_aktif, :shu_total, :status_ambil, :catatan)
                        """
                    ), {
                        'id_shu_anggota': a.get('id_alokasi') or str(uuid.uuid4()),
                        'id_shu': a.get('id_shu'),
                        'id_anggota': a.get('id_anggota') or '',
                        'no_anggota': a.get('no_anggota') or '',
                        'nama_anggota': a.get('nama_anggota') or '',
                        'jasa_anggota': float(a.get('jasa_anggota') or 0),
                        'nilai_jasa': float(a.get('nilai_jasa') or 0),
                        'shu_pasif': float(a.get('shu_pasif') or 0),
                        'shu_aktif': float(a.get('shu_aktif') or 0),
                        'shu_total': float(a.get('total_shu') or 0),
                        'status_ambil': a.get('status') or 'Draft',
                        'catatan': a.get('keterangan') or a.get('catatan') or '',
                    })
        except Exception:
            # fallback to file storage on any DB error
            records = _load_shu_records()
            existing_ids = {row.get('id_shu', ''): i for i, row in enumerate(records)}
            idx = existing_ids.get(record['id_shu'])
            if idx is not None:
                records[idx] = record
            else:
                records.append(record)
            tulis_csv(FILE_SHU_TAHUNAN, records, SHU_TAHUNAN_FIELDNAMES)

            allocations = [row for row in _load_shu_allocation() if row.get('id_shu') != record['id_shu']]
            allocations.extend(alokasi_rows)
            tulis_csv(FILE_SHU_ALOKASI, allocations, SHU_ALOKASI_FIELDNAMES)
        return
    # file fallback
    records = _load_shu_records()
    existing_ids = {row.get('id_shu', ''): i for i, row in enumerate(records)}
    idx = existing_ids.get(record['id_shu'])
    if idx is not None:
        records[idx] = record
    else:
        records.append(record)
    tulis_csv(FILE_SHU_TAHUNAN, records, SHU_TAHUNAN_FIELDNAMES)

    allocations = [row for row in _load_shu_allocation() if row.get('id_shu') != record['id_shu']]
    allocations.extend(alokasi_rows)
    tulis_csv(FILE_SHU_ALOKASI, allocations, SHU_ALOKASI_FIELDNAMES)


def _format_currency(value) -> str:
    try:
        return f"Rp {float(value):,.0f}"
    except Exception:
        return 'Rp 0'


@app.route('/shu')
@permission_required('shu.self.view', 'shu.view')
def halaman_shu():
    ensure_shu_schema()
    records = sorted(_load_shu_records(), key=lambda row: (row.get('tahun') or ''), reverse=True)
    allocations = _load_shu_allocation()
    current_role = _current_role()
    current_id_anggota = get_current_user_id_anggota()
    selected_id = (request.args.get('id_shu') or '').strip()

    selected_record = None
    if selected_id:
        selected_record = next((row for row in records if row.get('id_shu') == selected_id), None)
    if not selected_record and records:
        selected_record = records[0]

    selected_allocation = []
    if selected_record:
        selected_allocation = [row for row in allocations if row.get('id_shu') == selected_record.get('id_shu')]

    tahun_rincian = (selected_record.get('tahun') if selected_record else str(datetime.now().year)) if selected_record else str(datetime.now().year)
    provisi_details = _rincian_pinjaman_sumber_shu(str(tahun_rincian))

    can_manage = has_permission('shu.manage', current_role)
    can_validate = has_permission('shu.validate', current_role)
    can_view_all = has_permission('shu.view', current_role)

    current_no_anggota = ''
    if current_id_anggota:
        anggota_rows = baca_csv(FILE_ANGGOTA)
        anggota_row = next((a for a in anggota_rows if a.get('id_anggota') == current_id_anggota), None)
        current_no_anggota = (anggota_row.get('no_anggota') or '').strip() if anggota_row else ''

    if not can_view_all and (current_id_anggota or current_no_anggota):
        selected_allocation = [
            row for row in selected_allocation
            if (
                (current_id_anggota and row.get('id_anggota') == current_id_anggota)
                or (current_no_anggota and (row.get('no_anggota') or '').strip() == current_no_anggota)
            )
        ]
        provisi_details = [row for row in provisi_details if row.get('id_anggota') == current_id_anggota]

    my_allocation = None
    if selected_allocation:
        my_allocation = selected_allocation[0]
    elif current_id_anggota or current_no_anggota:
        for row in allocations:
            if (
                (current_id_anggota and row.get('id_anggota') == current_id_anggota)
                or (current_no_anggota and (row.get('no_anggota') or '').strip() == current_no_anggota)
            ):
                my_allocation = row
                break

    summary = {
        'total_shu': sum(_parse_float(row.get('total_shu')) for row in records),
        'shu_pasif_total': sum(_parse_float(row.get('shu_pasif_total')) for row in records),
        'shu_aktif_total': sum(_parse_float(row.get('shu_aktif_total')) for row in records),
        'jumlah_periode': len(records),
    }

    provisi_summary = {
        'tahun': str(tahun_rincian),
        'jumlah_pinjaman': len(provisi_details),
        'total_provisi': sum(_parse_float(row.get('provisi_nominal')) for row in provisi_details),
    }

    return render_template(
        'shu.html',
        records=records,
        selected_record=selected_record,
        selected_allocation=selected_allocation,
        my_allocation=my_allocation,
        can_manage=can_manage,
        can_validate=can_validate,
        summary=summary,
        provisi_details=provisi_details,
        provisi_summary=provisi_summary,
        format_currency=_format_currency,
        csrf_token=_get_or_create_csrf_token(),
        now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    )


@app.route('/shu/simpan', methods=['POST'])
@permission_required('shu.manage')
@csrf_protect
def simpan_shu():
    ensure_shu_schema()
    tahun = (request.form.get('tahun') or str(datetime.now().year)).strip()
    mode = (request.form.get('mode') or '').strip().lower()
    total_shu = _parse_float(request.form.get('total_shu'))
    catatan = (request.form.get('catatan') or '').strip()

    if not tahun:
        flash('Tahun SHU wajib diisi.', 'danger')
        return redirect(url_for('halaman_shu'))

    if mode == 'auto_provisi':
        total_shu = 0.0
        catatan = (f"[AUTO] SHU dihitung otomatis dari pendapatan provisi pinjaman tahun {tahun}. " + catatan).strip()

    hasil = hitung_shu_tahunan(total_shu, tahun, catatan)
    if _parse_float(hasil.get('total_shu')) <= 0:
        flash('Pendapatan provisi pinjaman pada tahun ini masih 0, sehingga SHU belum dapat dihitung.', 'warning')
        return redirect(url_for('halaman_shu'))

    id_shu = str(uuid.uuid4())
    now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    catatan_perhitungan = hasil.get('catatan_perhitungan') or catatan
    record = {
        'id_shu': id_shu,
        'tahun': hasil['tahun'],
        'tanggal_input': now_text,
        'total_shu': hasil['total_shu'],
        'cadangan_umum': hasil['cadangan_umum'],
        'shu_pasif_total': hasil['shu_pasif_total'],
        'shu_aktif_total': hasil['shu_aktif_total'],
        'dana_kesejahteraan': hasil['dana_kesejahteraan'],
        'dana_pendidikan': hasil['dana_pendidikan'],
        'dana_sosial': hasil['dana_sosial'],
        'dana_pembangunan': hasil['dana_pembangunan'],
        'dana_pengurus': hasil['dana_pengurus'],
        'dana_risiko': hasil['dana_risiko'],
        'status': 'Draft',
        'dikonfirmasi_oleh': '',
        'tanggal_konfirmasi': '',
        'catatan': catatan_perhitungan,
    }

    alokasi_rows = []
    for item in hasil['alokasi']:
        alokasi_rows.append({
            'id_alokasi': str(uuid.uuid4()),
            'id_shu': id_shu,
            'id_anggota': item['id_anggota'],
            'no_anggota': item['no_anggota'],
            'nama_anggota': item['nama_anggota'],
            'jasa_anggota': item['jasa_anggota'],
            'nilai_jasa': item['nilai_jasa'],
            'shu_pasif': item['shu_pasif'],
            'shu_aktif': item['shu_aktif'],
            'total_shu': item['total_shu'],
            'status': 'Draft',
            'keterangan': catatan_perhitungan,
        })

    _save_shu_data(record, alokasi_rows)
    flash(f"SHU tahun {tahun} berhasil dihitung dari sumber: {hasil.get('sumber_total_shu', 'Manual')} dan disimpan.", 'success')
    return redirect(url_for('halaman_shu', id_shu=id_shu))


@app.route('/shu/konfirmasi/<id_shu>', methods=['POST'])
@permission_required('shu.validate')
@csrf_protect
def konfirmasi_shu(id_shu: str):
    ensure_shu_schema()
    now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if DATABASE_URL:
        try:
            with db_session() as conn:
                res = conn.execute(text(
                    "UPDATE shu_tahunan SET status = 'Disetujui', dikonfirmasi_oleh = :user, tanggal_konfirmasi = :ts WHERE id_shu = :id_shu"
                ), {'user': session.get('user', ''), 'ts': now_text, 'id_shu': id_shu})
                conn.execute(text("UPDATE shu_anggota SET status_ambil = 'Disetujui' WHERE id_shu = :id_shu"), {'id_shu': id_shu})
                if res.rowcount and res.rowcount > 0:
                    flash('Data SHU berhasil dikonfirmasi.', 'success')
                else:
                    flash('Data SHU tidak ditemukan.', 'warning')
        except Exception:
            # fallback to file mode
            records = _load_shu_records()
            allocations = _load_shu_allocation()
            now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            changed = False
            for row in records:
                if row.get('id_shu') != id_shu:
                    continue
                row['status'] = 'Disetujui'
                row['dikonfirmasi_oleh'] = session.get('user', '')
                row['tanggal_konfirmasi'] = now_text
                changed = True
            for row in allocations:
                if row.get('id_shu') == id_shu:
                    row['status'] = 'Disetujui'
            if changed:
                tulis_csv(FILE_SHU_TAHUNAN, records, SHU_TAHUNAN_FIELDNAMES)
                tulis_csv(FILE_SHU_ALOKASI, allocations, SHU_ALOKASI_FIELDNAMES)
                flash('Data SHU berhasil dikonfirmasi.', 'success')
            else:
                flash('Data SHU tidak ditemukan.', 'warning')
        return redirect(url_for('halaman_shu', id_shu=id_shu))

    # file fallback
    records = _load_shu_records()
    allocations = _load_shu_allocation()
    now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    changed = False
    for row in records:
        if row.get('id_shu') != id_shu:
            continue
        row['status'] = 'Disetujui'
        row['dikonfirmasi_oleh'] = session.get('user', '')
        row['tanggal_konfirmasi'] = now_text
        changed = True
    for row in allocations:
        if row.get('id_shu') == id_shu:
            row['status'] = 'Disetujui'
    if changed:
        tulis_csv(FILE_SHU_TAHUNAN, records, SHU_TAHUNAN_FIELDNAMES)
        tulis_csv(FILE_SHU_ALOKASI, allocations, SHU_ALOKASI_FIELDNAMES)
        flash('Data SHU berhasil dikonfirmasi.', 'success')
    else:
        flash('Data SHU tidak ditemukan.', 'warning')
    return redirect(url_for('halaman_shu', id_shu=id_shu))


@app.route('/shu/hapus/<id_shu>', methods=['POST'])
@permission_required('shu.manage')
@csrf_protect
def hapus_shu(id_shu: str):
    ensure_shu_schema()
    if DATABASE_URL:
        try:
            with db_session() as conn:
                conn.execute(text("DELETE FROM shu_anggota WHERE id_shu = :id_shu"), {'id_shu': id_shu})
                res = conn.execute(text("DELETE FROM shu_tahunan WHERE id_shu = :id_shu"), {'id_shu': id_shu})
                if res.rowcount and res.rowcount > 0:
                    flash('Data SHU berhasil dihapus.', 'success')
                else:
                    flash('Data SHU tidak ditemukan.', 'warning')
        except Exception:
            records = [row for row in _load_shu_records() if row.get('id_shu') != id_shu]
            allocations = [row for row in _load_shu_allocation() if row.get('id_shu') != id_shu]
            tulis_csv(FILE_SHU_TAHUNAN, records, SHU_TAHUNAN_FIELDNAMES)
            tulis_csv(FILE_SHU_ALOKASI, allocations, SHU_ALOKASI_FIELDNAMES)
            flash('Data SHU berhasil dihapus.', 'success')
        return redirect(url_for('halaman_shu'))

    records = [row for row in _load_shu_records() if row.get('id_shu') != id_shu]
    allocations = [row for row in _load_shu_allocation() if row.get('id_shu') != id_shu]
    tulis_csv(FILE_SHU_TAHUNAN, records, SHU_TAHUNAN_FIELDNAMES)
    tulis_csv(FILE_SHU_ALOKASI, allocations, SHU_ALOKASI_FIELDNAMES)
    flash('Data SHU berhasil dihapus.', 'success')
    return redirect(url_for('halaman_shu'))


def generate_excel_laporan_terpadu():
    """Buat workbook Excel berisi 2 tabel: pinjaman + simpanan, plus tabel SHU pinjaman."""
    ensure_pinjaman_plafon_schema()
    ensure_simpanan_schema()
    ensure_pinjaman_cicilan_schema()

    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            raise RuntimeError("Fitur export membutuhkan openpyxl.")

    anggota_full = baca_csv(FILE_ANGGOTA)
    pinjaman_t = enrich_pinjaman_untuk_tampilan(baca_csv(FILE_PINJAMAN), anggota_full)
    pinjaman_t.reverse()

    wb = wb_cls()
    ws = wb.active
    ws.title = 'Laporan Terpadu'

    # -------------------
    # Tabel: PINJAMAN
    # -------------------
    ws.append(['DATA PINJAMAN'])
    ws.append(['No', 'Jenis', 'No Anggota', 'Nama', 'Plafon', 'Tenor', 'Bunga %', 'Cicilan/bln', 'Sisa', 'Status'])
    for i, p in enumerate(pinjaman_t, 1):
        ws.append([
            i,
            p.get('jenis_pinjaman', ''),
            p.get('no_anggota', ''),
            p.get('nama_anggota', ''),
            float(p.get('plafon') or p.get('total_pinjaman') or 0),
            int(float(p.get('tenor_bulan') or p.get('tenor') or 0)),
            float(p.get('bunga_persen') or 0),
            float(p.get('cicilan_per_bulan') or 0),
            saldo_pinjaman_aktual(p),
            p.get('status', ''),
        ])

    # -------------------
    # SHU (65% untuk anggota)
    # -------------------
    shu = hitung_shu_pinjaman()
    ws.append([])
    ws.append(['SHU PINJAMAN'])
    ws.append(['SHU Kotor (bunga dibayar)', float(shu['shu_kotor'])])
    ws.append(['SHU Anggota (65%)', float(shu['shu_anggota_65'])])
    ws.append(['SHU Lainnya (35%)', float(shu['shu_lain'])])
    ws.append([])

    ws.append(['Alokasi SHU Anggota 65% (proporsional bunga dibayar)'])
    ws.append(['No', 'No Anggota', 'Nama', 'Bunga dibayar', 'SHU anggota'])
    for i, a in enumerate(shu['alokasi'], 1):
        ws.append([i, a['no_anggota'], a['nama_anggota'], float(a['bunga_dibayar']), float(a['shu_anggota'])])

    # -------------------
    # Tabel: SIMPANAN
    # -------------------
    simpanan_t = enrich_simpanan_untuk_tampilan(baca_csv(FILE_SIMPANAN), anggota_full)
    simpanan_t.reverse()
    ws.append([])
    ws.append(['DATA SIMPANAN (saldo per anggota)'])
    ws.append(['No', 'No Anggota', 'Nama', 'Total Simpanan'])
    for i, s in enumerate(simpanan_t, 1):
        ws.append([
            i,
            s.get('no_anggota', ''),
            s.get('nama_anggota', ''),
            float(s.get('total_simpanan') or 0),
        ])

    return wb


@app.route('/laporan')
@permission_required('reports.self.view', 'reports.strategic.view')
def halaman_laporan():
    ensure_simpanan_schema()
    simpanan = baca_csv(FILE_SIMPANAN)
    pinjaman = baca_csv(FILE_PINJAMAN)
    anggota = baca_csv(FILE_ANGGOTA)
    if not is_current_user_admin():
        id_anggota = get_current_user_id_anggota()
        simpanan = [s for s in simpanan if s.get('id_anggota') == id_anggota]
        pinjaman = [p for p in pinjaman if p.get('id_anggota') == id_anggota]
        anggota = [a for a in anggota if a.get('id_anggota') == id_anggota]
    return render_template('laporan.html', simpanan=simpanan, pinjaman=pinjaman, anggota=anggota)


@app.route('/export/laporan-terpadu')
@permission_required('reports.export')
def export_laporan_terpadu():
    """Upsert anggota dari data pinjaman/simpanan lalu siapkan download Excel 1 file."""
    ensure_simpanan_schema()
    ensure_pinjaman_plafon_schema()

    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    simpanan_rows = baca_csv(FILE_SIMPANAN)
    hasil = upsert_anggota_from_riwayat(pinjaman_rows, simpanan_rows)

    return render_template(
        'export_laporan_terpadu.html',
        added=hasil['added'],
        updated=hasil['updated'],
        failed=hasil.get('failed', 0),
        download_url=url_for('export_laporan_terpadu_unduh'),
    )


@app.route('/export/laporan-terpadu/unduh')
@permission_required('reports.export')
def export_laporan_terpadu_unduh():
    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            return f"Fitur export membutuhkan openpyxl. Jalankan: \"{sys.executable}\" -m pip install openpyxl", 500

    wb = generate_excel_laporan_terpadu()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='laporan_terpadu_pinjam_simpan.xlsx',
        as_attachment=True
    )


@app.route('/export/simpanan')
@permission_required('reports.export')
def export_simpanan():
    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    ensure_iuran_sosial_schema()
    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            return f"Fitur export membutuhkan openpyxl. Jalankan: \"{sys.executable}\" -m pip install openpyxl", 500
    wb = wb_cls()
    ws = wb.active
    ws.title = "Data Simpanan"
    ws.append(['No', 'No Anggota', 'Nama', 'Total Simpanan', 'Jenis Simpanan Anggota', 'Jenis Pinjaman Anggota'])
    anggota_rows = baca_csv(FILE_ANGGOTA)
    simpanan = enrich_simpanan_untuk_tampilan(baca_csv(FILE_SIMPANAN), anggota_rows)
    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    simpanan_transaksi_rows = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    iuran_sosial_rows = baca_csv(FILE_IURAN_SOSIAL)
    riwayat_simpanan_map = _riwayat_jenis_simpanan_per_anggota(simpanan_transaksi_rows, iuran_sosial_rows)
    riwayat_pinjaman_map = _riwayat_jenis_pinjaman_per_anggota(pinjaman_rows)
    for i, s in enumerate(simpanan, 1):
        id_anggota = s.get('id_anggota', '')
        ws.append([
            i,
            s['no_anggota'],
            s['nama_anggota'],
            float(s.get('total_simpanan') or 0),
            riwayat_simpanan_map.get(id_anggota, '-'),
            riwayat_pinjaman_map.get(id_anggota, '-'),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='laporan_simpanan.xlsx', as_attachment=True)


@app.route('/export/pinjaman')
@permission_required('reports.export')
def export_pinjaman():
    ensure_simpanan_transaksi_schema()
    ensure_iuran_sosial_schema()
    wb_cls = Workbook
    if wb_cls is None:
        try:
            from openpyxl import Workbook as wb_cls  # type: ignore
        except ModuleNotFoundError:
            return f"Fitur export membutuhkan openpyxl. Jalankan: \"{sys.executable}\" -m pip install openpyxl", 500
    wb = wb_cls()
    ws = wb.active
    ws.title = "Data Pinjaman"
    ws.append([
        'No', 'Jenis', 'Riwayat Jenis Pinjaman', 'No Anggota', 'Nama',
        'Plafon', 'Tenor', 'Cicilan/Bulan', 'Sisa', 'Status', 'Jenis Simpanan Anggota'
    ])
    anggota_rows = baca_csv(FILE_ANGGOTA)
    pinjaman_rows = baca_csv(FILE_PINJAMAN)
    pinjaman = enrich_pinjaman_untuk_tampilan(pinjaman_rows, anggota_rows)
    simpanan_transaksi_rows = baca_csv(FILE_SIMPANAN_TRANSAKSI)
    iuran_sosial_rows = baca_csv(FILE_IURAN_SOSIAL)
    riwayat_simpanan_map = _riwayat_jenis_simpanan_per_anggota(simpanan_transaksi_rows, iuran_sosial_rows)
    for i, p in enumerate(pinjaman, 1):
        id_anggota = p.get('id_anggota', '')
        ws.append([
            i,
            p.get('jenis_pinjaman', ''),
            p.get('riwayat_jenis_pinjaman', p.get('jenis_pinjaman', '')),
            p['no_anggota'],
            p['nama_anggota'],
            float(p.get('plafon') or 0),
            int(float(p.get('tenor_bulan') or 0)),
            float(p.get('cicilan_per_bulan') or 0),
            saldo_pinjaman_aktual(p),
            p.get('status', ''),
            riwayat_simpanan_map.get(id_anggota, '-'),
        ])

    # Tambahkan SHU Pinjaman
    shu = hitung_shu_pinjaman()
    ws.append([])
    ws.append(['SHU PINJAMAN'])
    ws.append(['SHU Kotor (bunga dibayar)', float(shu['shu_kotor'])])
    ws.append(['SHU Anggota (65%)', float(shu['shu_anggota_65'])])
    ws.append(['SHU Lainnya (35%)', float(shu['shu_lain'])])
    ws.append([])
    ws.append(['Alokasi SHU Anggota 65% (proporsional bunga dibayar)'])
    ws.append(['No', 'No Anggota', 'Nama', 'Bunga dibayar', 'SHU anggota'])
    for i, a in enumerate(shu['alokasi'], 1):
        ws.append([i, a['no_anggota'], a['nama_anggota'], float(a['bunga_dibayar']), float(a['shu_anggota'])])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='laporan_pinjaman.xlsx', as_attachment=True)


@app.route('/laporan/anggota/<id_anggota>')
@login_required
def laporan_anggota(id_anggota):
    """Detail laporan per anggota."""
    restrict_id_anggota_or_forbid(id_anggota)
    anggota = baca_csv(FILE_ANGGOTA)
    anggota_data = next((a for a in anggota if a['id_anggota'] == id_anggota), None)
    if not anggota_data:
        return "Anggota tidak ditemukan", 404

    ensure_simpanan_schema()
    ensure_simpanan_transaksi_schema()
    anggota_all = baca_csv(FILE_ANGGOTA)
    simpanan_raw = [s for s in baca_csv(FILE_SIMPANAN) if s['id_anggota'] == id_anggota]
    simpanan_transaksi = [s for s in baca_csv(FILE_SIMPANAN_TRANSAKSI) if s.get('id_anggota') == id_anggota]
    for t in simpanan_transaksi:
        ket = (t.get('keterangan') or '').strip()
        if ket:
            t['keterangan'] = ket.replace('Setoran', 'Simpanan').replace('setoran', 'simpanan')
    pinjaman_raw = [p for p in baca_csv(FILE_PINJAMAN) if p['id_anggota'] == id_anggota]
    simpanan = enrich_simpanan_untuk_tampilan(simpanan_raw, anggota_all)
    pinjaman = enrich_pinjaman_untuk_tampilan(pinjaman_raw, anggota_all)

    pinjaman_laporan = []
    total_saldo_diterima = 0.0
    for p in pinjaman:
        row = dict(p)
        plafon = float(row.get('plafon') or 0)
        tenor = int(float(row.get('tenor_bulan') or 0))
        status = (row.get('status') or '').strip()
        provisi = plafon * PROVISI_RATE_LONG_TENOR if tenor > 12 else 0.0
        saldo_diterima = max(plafon - provisi, 0.0)
        if status == 'Disetujui':
            row['saldo_diterima'] = str(round(saldo_diterima, 2))
            total_saldo_diterima += saldo_diterima
        else:
            row['saldo_diterima'] = ''
        pinjaman_laporan.append(row)

    total_simpanan = sum(float(s.get('total_simpanan') or 0) for s in simpanan_raw)
    total_pinjaman = sum(
        saldo_pinjaman_aktual(p)
        for p in pinjaman_raw
        if p.get('status') == 'Disetujui'
    )

    return render_template('cetak_struk.html',
                           anggota=anggota_data,
                           simpanan=simpanan,
                           simpanan_transaksi=sorted(simpanan_transaksi, key=lambda x: x.get('tanggal', ''), reverse=True),
                           pinjaman=pinjaman_laporan,
                           total_simpanan=total_simpanan,
                           total_pinjaman=total_pinjaman,
                           total_saldo_diterima=total_saldo_diterima)


# ══════════════════════════════════════════════
#  JALANKAN APLIKASI
# ══════════════════════════════════════════════
from koperasi_system.route_aliases import register_route_aliases

bootstrap_storage_files()
register_route_aliases(app)

if __name__ == '__main__':
    print("=" * 50)
    print("  Kepemilikan dan Pengelolaan Aplikasi:")
    print("  KPRI BLK BANDUNG")
    print("  Login : http://localhost:5000")
    print("=" * 50)
    app.run(debug=False, host='127.0.0.1', port=5000)