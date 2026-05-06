import argparse
import os
import uuid
from datetime import datetime
from pathlib import Path

import sys

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    try:
        for raw_line in path.read_text(encoding='utf-8').splitlines():
            line = raw_line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
    except Exception:
        pass


_load_env_file(ROOT_DIR / '.env')

from sqlalchemy import text

from koperasi_system.db import db_session, init_db_schema
from koperasi_system.settings import (
    BASE_DIR,
    FILE_IMPORT_LOG,
    FILE_PENDAFTARAN_ANGGOTA,
    PENDAFTARAN_FIELDNAMES,
)


def _read_xlsx_as_dicts(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as e:  # pragma: no cover
        raise RuntimeError("openpyxl belum terpasang. Jalankan: pip install openpyxl") from e

    if not os.path.exists(path):
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in (rows[0] or [])]
    out = []
    for rr in rows[1:]:
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = rr[i] if i < len(rr) else None
            item[h] = "" if v is None else str(v).strip()
        if any(str(v or "").strip() for v in item.values()):
            out.append(item)
    return out


def _migrate_pendaftaran_anggota(rows: list[dict]) -> int:
    if not rows:
        return 0

    cols = PENDAFTARAN_FIELDNAMES
    placeholders = ", ".join(f":{c}" for c in cols)
    updatable = [c for c in cols if c != "id_pengajuan"]
    set_clause = ", ".join(f"{c}=EXCLUDED.{c}" for c in updatable)
    stmt = text(
        f"""
        INSERT INTO pendaftaran_anggota ({", ".join(cols)})
        VALUES ({placeholders})
        ON CONFLICT (id_pengajuan) DO UPDATE SET {set_clause}
        """
    )

    migrated = 0
    with db_session() as conn:
        for r in rows:
            payload = {c: (r.get(c) if r.get(c) is not None else "") for c in cols}
            if not str(payload.get("id_pengajuan") or "").strip():
                payload["id_pengajuan"] = str(uuid.uuid4())
            conn.execute(stmt, payload)
            migrated += 1
    return migrated


def _migrate_import_log(rows: list[dict]) -> int:
    if not rows:
        return 0

    stmt = text(
        """
        INSERT INTO import_log (waktu, "user", mode, nama_file, berhasil, gagal, catatan)
        VALUES (:waktu, :user, :mode, :nama_file, :berhasil, :gagal, :catatan)
        """
    )

    migrated = 0
    with db_session() as conn:
        for r in rows:
            payload = {
                "waktu": (r.get("waktu") or "")[:19],
                "user": (r.get("user") or ""),
                "mode": (r.get("mode") or ""),
                "nama_file": (r.get("nama_file") or ""),
                "berhasil": int(float(r.get("berhasil") or 0) or 0),
                "gagal": int(float(r.get("gagal") or 0) or 0),
                "catatan": (r.get("catatan") or "")[:2000],
            }
            # Jika format waktu tidak sesuai, fallback ke now agar insert tidak gagal.
            try:
                if payload["waktu"]:
                    datetime.strptime(payload["waktu"], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                payload["waktu"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(stmt, payload)
            migrated += 1
    return migrated


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Migrasi pendaftaran_anggota dan import_log dari Excel ke PostgreSQL."
    )
    parser.add_argument(
        "--schema",
        default=os.path.join(BASE_DIR, "db", "schema.sql"),
        help="Path schema SQL untuk inisialisasi tabel (default: db/schema.sql).",
    )
    parser.add_argument(
        "--no-init-schema",
        action="store_true",
        help="Jangan jalankan inisialisasi schema (asumsikan tabel sudah ada).",
    )
    args = parser.parse_args()

    if not args.no_init_schema:
        init_db_schema(args.schema)

    pendaftaran_rows = _read_xlsx_as_dicts(FILE_PENDAFTARAN_ANGGOTA)
    import_log_rows = _read_xlsx_as_dicts(FILE_IMPORT_LOG)

    n_pendaftaran = _migrate_pendaftaran_anggota(pendaftaran_rows)
    n_import_log = _migrate_import_log(import_log_rows)

    print(f"OK. Migrated pendaftaran_anggota: {n_pendaftaran} baris.")
    print(f"OK. Migrated import_log: {n_import_log} baris.")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())

